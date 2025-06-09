import os
from datetime import datetime, timedelta, date
from flask import Flask, render_template, redirect, url_for, flash, request, current_app, send_from_directory, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, current_user, UserMixin, login_required
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from urllib.parse import urlparse, quote_plus, parse_qs
from functools import wraps
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect, generate_csrf
from wtforms import StringField, PasswordField, BooleanField, SubmitField, TextAreaField, SelectField, IntegerField, MultipleFileField
from wtforms.validators import DataRequired, Email, EqualTo, ValidationError, Length, Optional, NumberRange, InputRequired, Regexp
import logging
from markupsafe import escape, Markup
import re
import uuid
from flask_mail import Mail, Message
import sys
from itertools import groupby
from sqlalchemy import func, Date, cast, or_ # For date-based aggregation & OR conditions
from sqlalchemy.orm import aliased
from wtforms import SelectField, TextAreaField, StringField, SubmitField # Ensure these are imported
from wtforms.validators import DataRequired, Length, Optional
import csv
import io # For in-memory file handling
from flask import make_response # For sending files
import openpyxl # For XLSX
from openpyxl.utils import get_column_letter
from wtforms import StringField, PasswordField, BooleanField, SubmitField, SelectField, IntegerField, MultipleFileField # Ensure SelectField is imported
from wtforms.validators import DataRequired, Email, EqualTo, ValidationError, Length, Optional, NumberRange, InputRequired, Regexp # Ensure ValidationError is 



# --- New Imports for Integrated Features ---
import google.generativeai as genai
import requests
from bs4 import BeautifulSoup
import json

from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate # Add this import


# Twilio Integration
from twilio.rest import Client as TwilioClient
from twilio.base.exceptions import TwilioRestException

# --- Global Constants for Form Choices ---
AWS_SERVICE_CHOICES = [
    ('', '--- Select AWS Service (if AWS) ---'),
    ('EC2', 'EC2 - Elastic Compute Cloud'), ('S3', 'S3 - Simple Storage Service'),
    ('RDS', 'RDS - Relational Database Service'), ('Lambda', 'Lambda - Serverless Compute'),
    ('VPC', 'VPC - Virtual Private Cloud'), ('Route53', 'Route 53 - DNS Service'),
    ('IAM', 'IAM - Identity & Access Management'), ('CloudFront', 'CloudFront - CDN'),
    ('Other', 'Other AWS Service'),
]
TICKET_STATUS_CHOICES = [('Open', 'Open'), ('In Progress', 'In Progress'), ('On Hold', 'On Hold'), ('Resolved', 'Resolved'), ('Closed', 'Closed')]
TICKET_PRIORITY_CHOICES = [('Low', 'Low'), ('Medium', 'Medium'), ('High', 'High'), ('Urgent', 'Urgent')]
PRIORITY_ORDER_MAP = {
    'Urgent': 0,
    'High': 1,
    'Medium': 2,
    'Low': 3
}

TICKET_STATUS_CHOICES_FLAT = [s[0].lower() for s in TICKET_STATUS_CHOICES]

REQUEST_CALL_BACK_CHOICES = [('', '-'), ('Yes', 'Yes'), ('No', 'No')]
EFFORT_CHOICES = [
    ('', '-'), ('0', 'Not Applicable (0 min)'),
    ('15', '15 min'), ('30', '30 min'), ('45', '45 min'),
    ('60', '1 hour (60 min)'), ('90', '1.5 hours (90 min)'), ('120', '2 hours (120 min)'),
    ('180', '3 hours (180 min)'), ('240', '4 hours (240 min)'), ('300', '5 hours (300 min)'),
    ('360', '6 hours (360 min)'), ('420', '7 hours (420 min)'), ('480', '8 hours (1 day)'),
    ('960', '16 hours (2 days)'), ('1440', '24 hours (3 days)')
]

def to_snake_case(name):
    name = re.sub(r'(.)([A-Z][a-z]+)', r'\1_\2', name)
    name = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', name).lower()
    if name.endswith("_option"): name = name[:-7]
    elif name.endswith("option"): name = name[:-6]
    return name

class Config:
    SECRET_KEY = os.environ.get('SECRET_KEY') or 'ticket-cms-agent-views-final-key-secure'
    MYSQL_USER = os.environ.get('MYSQL_USER_TICKET_CMS') or 'ticket_user'
    _raw_mysql_password = os.environ.get('MYSQL_PASSWORD_TICKET_CMS') or 'Jodha@123'
    MYSQL_PASSWORD_ENCODED = quote_plus(_raw_mysql_password) if _raw_mysql_password else ''
    MYSQL_HOST = os.environ.get('MYSQL_HOST_TICKET_CMS') or 'localhost'
    MYSQL_DB = os.environ.get('MYSQL_DB_TICKET_CMS') or 'ticket_cms_db'
    MYSQL_CHARSET = 'utf8mb4'
    APPLICATION_ROOT = '/'

    if not all([MYSQL_USER, _raw_mysql_password, MYSQL_HOST, MYSQL_DB]):
        print("FATAL ERROR: Missing critical MySQL configuration.")
        SQLALCHEMY_DATABASE_URI = None
    else:
        SQLALCHEMY_DATABASE_URI = (
            f"mysql+pymysql://{MYSQL_USER}:{MYSQL_PASSWORD_ENCODED}@"
            f"{MYSQL_HOST}/{MYSQL_DB}?charset={MYSQL_CHARSET}"
        )

    SQLALCHEMY_TRACK_MODIFICATIONS = False
    SQLALCHEMY_ECHO = os.environ.get('SQLALCHEMY_ECHO', 'False').lower() in ['true', '1', 't']

    MAIL_SERVER = os.environ.get('MAIL_SERVER_TICKET_CMS') or 'smtp.gmail.com'
    MAIL_PORT = int(os.environ.get('MAIL_PORT_TICKET_CMS') or 587)
    MAIL_USE_TLS = os.environ.get('MAIL_USE_TLS_TICKET_CMS', 'true').lower() in ['true', '1', 't']
    MAIL_USE_SSL = os.environ.get('MAIL_USE_SSL_TICKET_CMS', 'false').lower() in ['true', '1', 't']
    MAIL_USERNAME = os.environ.get('MAIL_USERNAME_TICKET_CMS')
    MAIL_PASSWORD = os.environ.get('MAIL_PASSWORD_TICKET_CMS')
    MAIL_DEFAULT_SENDER_EMAIL = os.environ.get('MAIL_DEFAULT_SENDER_EMAIL_TICKET_CMS') or MAIL_USERNAME
    MAIL_DEFAULT_SENDER = ('TicketSys Admin', MAIL_DEFAULT_SENDER_EMAIL or 'noreply@example.com')
    ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL_TICKET_CMS')

    IMAP_SERVER = os.environ.get('IMAP_SERVER_TICKET_CMS_FETCH') or 'imap.gmail.com'
    IMAP_USERNAME = os.environ.get('IMAP_USERNAME_TICKET_CMS_FETCH')
    IMAP_PASSWORD = os.environ.get('IMAP_PASSWORD_TICKET_CMS_FETCH') or 'placeholder16apppass'
    IMAP_MAILBOX_FOLDER = os.environ.get('IMAP_MAILBOX_FOLDER_TICKET_CMS_FETCH') or 'INBOX'
    EMAIL_TICKET_DEFAULT_CATEGORY_NAME = os.environ.get('EMAIL_TICKET_DEFAULT_CATEGORY_NAME') or 'General Inquiry'
    EMAIL_TICKET_DEFAULT_SEVERITY_NAME = os.environ.get('EMAIL_TICKET_DEFAULT_SEVERITY_NAME') or 'Severity 3 (Medium)'

    BASE_URL = os.environ.get('BASE_URL') or 'http://localhost:5000'
    _parsed_base = urlparse(BASE_URL)
    SERVER_NAME = os.environ.get('SERVER_NAME') or _parsed_base.netloc
    PREFERRED_URL_SCHEME = os.environ.get('PREFERRED_URL_SCHEME') or _parsed_base.scheme or 'http'

    _default_upload_folder = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'uploads')
    UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', _default_upload_folder)
    ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'log', 'csv'}
    MAX_CONTENT_LENGTH = int(os.environ.get('MAX_CONTENT_LENGTH') or 16 * 1000 * 1000)

    TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID_TICKET_CMS')
    TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN_TICKET_CMS')
    TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER_TICKET_CMS')
    EMERGENCY_CALL_RECIPIENT_PHONE_NUMBER = os.environ.get('EMERGENCY_CALL_RECIPIENT_PHONE_NUMBER_TICKET_CMS')
    SEVERITIES_FOR_CALL_ALERT = ["Severity 1 (Critical)", "Severity 2 (High)"]

    GEMINI_API_KEY = os.getenv('GOOGLE_API_KEY')


app = Flask(__name__)
app.config.from_object(Config)

db = SQLAlchemy(app)
migrate = Migrate(app, db)

app.jinja_env.add_extension('jinja2.ext.do')

logging.basicConfig(level=logging.INFO)
app.logger.setLevel(logging.INFO)

app.logger.info(f"--- App Initialization ---")
app.logger.info(f"Configured UPLOAD_FOLDER: {app.config['UPLOAD_FOLDER']}")
if app.config['UPLOAD_FOLDER'] == "/path/to/your/uploads":
    app.logger.warning("UPLOAD_FOLDER is set to '/path/to/your/uploads'. This is a placeholder. Unsetting it to use default.")
    app.config['UPLOAD_FOLDER'] = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'uploads')
    app.logger.info(f"UPLOAD_FOLDER reset to default: {app.config['UPLOAD_FOLDER']}")

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    try:
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        app.logger.info(f"Upload folder ensured/created: {app.config['UPLOAD_FOLDER']}")
    except OSError as e:
        app.logger.error(f"Could not create upload folder {app.config['UPLOAD_FOLDER']}: {e}. Check permissions and path.")

if not app.config['SQLALCHEMY_DATABASE_URI']:
    raise RuntimeError("SQLALCHEMY_DATABASE_URI is not configured. Application cannot start.")

csrf = CSRFProtect(app)
mail = Mail(app)

if not app.config['GEMINI_API_KEY']:
    app.logger.warning("CRITICAL WARNING: GOOGLE_API_KEY (GEMINI_API_KEY in Config) environment variable not set. AI features will fail.")
else:
    try:
        genai.configure(api_key=app.config['GEMINI_API_KEY'])
        app.logger.info("Gemini API Key configured successfully.")
    except Exception as e:
        app.logger.error(f"Error configuring Gemini API: {e}")
        app.config['GEMINI_API_KEY'] = None


def nl2br_filter(value):
    if not isinstance(value, str): value = str(value)
    escaped_value = escape(value)
    br_value = re.sub(r'(\r\n|\r|\n)', '<br>\n', escaped_value)
    return Markup(br_value)
app.jinja_env.filters['nl2br'] = nl2br_filter

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# --- Models ---
# In app.py

# Ensure these imports are at the top of your app.py
from itsdangerous import URLSafeTimedSerializer as Serializer 
from flask import current_app # Should already be there
# from flask_sqlalchemy import SQLAlchemy # Should already be there
# from flask_login import UserMixin # Should already be there
# from werkzeug.security import generate_password_hash, check_password_hash # Should already be there
# from . import db # Assuming db is initialized in __init__.py or globally in app.py

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), index=True, unique=True, nullable=False)
    email = db.Column(db.String(120), index=True, unique=True, nullable=False)
    password_hash = db.Column(db.String(256))
    role = db.Column(db.String(20), default='client', nullable=False) # client, organization_client, agent, admin
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    organization_id = db.Column(db.Integer, db.ForeignKey('organization_options.id'), nullable=True)
    organization = db.relationship('OrganizationOption', backref=db.backref('users', lazy='dynamic'))

    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=True)
    department = db.relationship('Department', backref=db.backref('users', lazy='dynamic'))

    tickets_created = db.relationship('Ticket', foreign_keys='Ticket.created_by_id', backref='creator', lazy='dynamic')
    tickets_assigned = db.relationship('Ticket', foreign_keys='Ticket.assigned_to_id', backref='assignee', lazy='dynamic')
    comments = db.relationship('Comment', backref='author', lazy='dynamic')

    def set_password(self, password): 
        self.password_hash = generate_password_hash(password)

    def check_password(self, password): 
        return check_password_hash(self.password_hash or "", password)

    @property
    def is_admin(self): return self.role == 'admin'
    @property
    def is_agent(self): return self.role == 'agent'
    @property
    def is_client(self): return self.role in ['client', 'organization_client']
    @property
    def is_department_client(self): return self.role == 'client' and self.department_id is not None
    @property
    def is_organization_client(self): 
        # An OrgClient has role 'organization_client' OR they are 'client' with an org but no specific dept (less common setup)
        return self.role == 'organization_client' or \
               (self.role == 'client' and self.organization_id and not self.department_id)


    def get_organization_name(self): return self.organization.name if self.organization else None
    def get_department_name(self): return self.department.name if self.department else None
    
    def __repr__(self): return f'<User {self.username} ({self.role})>'

    # --- Password Reset Token Methods ---
    def get_reset_password_token(self, expires_sec=1800): # Token expires in 30 minutes
        s = Serializer(current_app.config['SECRET_KEY'])
        return s.dumps({'user_id': self.id})

    @staticmethod
    def verify_reset_password_token(token, expires_sec=1800):
        s = Serializer(current_app.config['SECRET_KEY'])
        try:
            data = s.loads(token, max_age=expires_sec)
            user_id = data.get('user_id')
        except Exception as e: 
            current_app.logger.warning(f"Password reset token verification failed: {e}")
            return None
        return db.session.get(User, user_id)

class Category(db.Model):
    __tablename__ = 'categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.String(200), nullable=True)
    tickets = db.relationship('Ticket', backref='category_ref', lazy='dynamic')
    def __repr__(self): return f'<Category {self.name}>'


class CloudProviderOption(db.Model):
    __tablename__ = 'cloud_provider_options'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False, index=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    def __repr__(self): return f'<CloudProviderOption {self.name}>'


class SeverityOption(db.Model):
    __tablename__ = 'severity_options'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False, index=True)
    description = db.Column(db.String(100), nullable=True)
    order = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    def __repr__(self): return f'<SeverityOption {self.name}>'


class EnvironmentOption(db.Model):
    __tablename__ = 'environment_options'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False, index=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    def __repr__(self): return f'<EnvironmentOption {self.name}>'


class OrganizationOption(db.Model):
    __tablename__ = 'organization_options'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False, index=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    # `users` backref created by User.organization relationship
    tickets = db.relationship('Ticket', backref='organization_option_ref', lazy='dynamic')
    departments = db.relationship('Department', backref='organization_option_ref', lazy='dynamic', cascade="all, delete-orphan")
    def __repr__(self): return f'<OrganizationOption {self.name}>'

# NEW Department Model
class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    organization_id = db.Column(db.Integer, db.ForeignKey('organization_options.id'), nullable=False, index=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    # `users` backref will be automatically created by User.department relationship
    # `tickets` backref will be automatically created by Ticket.department_ref relationship

    __table_args__ = (db.UniqueConstraint('name', 'organization_id', name='uq_department_name_organization'),)

    def __repr__(self):
        org_name = self.organization_option_ref.name if self.organization_option_ref else "Unknown Org"
        return f'<Department {self.name} (Org: {org_name})>'


class FormTypeOption(db.Model):
    __tablename__ = 'form_type_options'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False, index=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    tickets = db.relationship('Ticket', backref='form_type_option_ref', lazy='dynamic')
    def __repr__(self): return f'<FormTypeOption {self.name}>'


class APNOpportunityOption(db.Model):
    __tablename__ = 'apn_opportunity_options'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False, index=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    tickets = db.relationship('Ticket', backref='apn_opportunity_option_ref', lazy='dynamic')
    def __repr__(self): return f'<APNOpportunityOption {self.name}>'


class SupportModalOption(db.Model):
    __tablename__ = 'support_modal_options'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False, index=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    tickets = db.relationship('Ticket', backref='support_modal_option_ref', lazy='dynamic')
    def __repr__(self): return f'<SupportModalOption {self.name}>'


class Ticket(db.Model):
    __tablename__ = 'tickets'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='Open', nullable=False)
    priority = db.Column(db.String(20), default='Medium', nullable=False)
    created_at = db.Column(db.DateTime, index=True, default=datetime.utcnow)
    resolved_at = db.Column(db.DateTime, nullable=True)
    updated_at = db.Column(db.DateTime, index=True, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    assigned_to_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    category_id = db.Column(db.Integer, db.ForeignKey('categories.id'), nullable=False)
    cloud_provider = db.Column(db.String(50), nullable=True)
    severity = db.Column(db.String(50), nullable=False)
    aws_service = db.Column(db.String(100), nullable=True)
    aws_account_id = db.Column(db.String(20), nullable=True)
    environment = db.Column(db.String(50), nullable=True)
    organization_id = db.Column(db.Integer, db.ForeignKey('organization_options.id'), nullable=True, index=True)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=True, index=True)
    department_ref = db.relationship('Department', backref=db.backref('tickets', lazy='dynamic')) # Note: "ref" matches model name convention
    form_type_id = db.Column(db.Integer, db.ForeignKey('form_type_options.id'), nullable=True)
    tags = db.Column(db.Text, nullable=True)
    additional_email_recipients = db.Column(db.Text, nullable=True)
    request_call_back = db.Column(db.String(10), nullable=True)
    contact_details = db.Column(db.String(255), nullable=True)
    aws_support_case_id = db.Column(db.String(50), nullable=True)
    effort_required_to_resolve_min = db.Column(db.Integer, nullable=True)
    customer_name = db.Column(db.String(100), nullable=False) # Consider if this should default to Org or Dept name for clients
    apn_opportunity_id = db.Column(db.Integer, db.ForeignKey('apn_opportunity_options.id'), nullable=True)
    apn_opportunity_description = db.Column(db.Text, nullable=True)
    support_modal_id = db.Column(db.Integer, db.ForeignKey('support_modal_options.id'), nullable=True)
    first_response_at = db.Column(db.DateTime, nullable=True)
    first_response_duration_minutes = db.Column(db.Integer, nullable=True)
    total_resolution_duration_minutes = db.Column(db.Integer, nullable=True)

    comments = db.relationship('Comment', backref='ticket_ref', lazy='dynamic', cascade="all, delete-orphan")

    cloud_provider_obj = db.relationship('CloudProviderOption', foreign_keys=[cloud_provider], primaryjoin='Ticket.cloud_provider == CloudProviderOption.name', viewonly=True)
    severity_obj = db.relationship('SeverityOption', foreign_keys=[severity], primaryjoin='Ticket.severity == SeverityOption.name', viewonly=True)
    environment_obj = db.relationship('EnvironmentOption', foreign_keys=[environment], primaryjoin='Ticket.environment == EnvironmentOption.name', viewonly=True)

    def __repr__(self): return f'<Ticket {self.id}: {self.title}>'


class Comment(db.Model):
    __tablename__ = 'comments'
    id = db.Column(db.Integer, primary_key=True)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, index=True, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False)
    is_internal = db.Column(db.Boolean, default=False)
    def __repr__(self): return f'<Comment {self.id} on Ticket {self.ticket_id}>'


class Attachment(db.Model):
    __tablename__ = 'attachments'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    stored_filename = db.Column(db.String(255), unique=True, nullable=False)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    content_type = db.Column(db.String(100), nullable=True)
    ticket = db.relationship('Ticket', backref=db.backref('ticket_attachments', lazy='dynamic', cascade="all, delete-orphan"))
    uploader = db.relationship('User', backref='uploaded_attachments_ref')
    def __repr__(self): return f'<Attachment {self.filename} for Ticket {self.ticket_id}>'


class Interaction(db.Model):
    __tablename__ = 'interactions'
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    interaction_type = db.Column(db.String(50), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    details = db.Column(db.JSON, nullable=True)
    ticket = db.relationship('Ticket', backref=db.backref('interactions_rel', lazy='dynamic', cascade="all, delete-orphan"))
    user = db.relationship('User', backref='interactions_rel')
    def __repr__(self): return f'<Interaction {self.id} on Ticket {self.ticket_id} - {self.interaction_type}>'


# --- Forms ---
class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(max=64)])
    password = PasswordField('Password', validators=[DataRequired()])
    remember_me = BooleanField('Remember Me')
    submit = SubmitField('Login')

class AdminUserForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=64)])
    email = StringField('Email', validators=[DataRequired(), Email(), Length(max=120)])
    password = PasswordField('Password (leave blank to keep current)', validators=[Optional(), Length(min=6)])
    password2 = PasswordField('Confirm Password', 
                              validators=[Optional(), EqualTo('password', message='Passwords must match if new password provided.')]) # Added Optional here too
    role = SelectField('Role', 
                       choices=[('client', 'Client (Company Specific)'), 
                                ('organization_client', 'Org Client (Sees all Companies in their Org)'), 
                                ('agent', 'Agent'), 
                                ('admin', 'Admin')], 
                       validators=[DataRequired()])
    organization_id = SelectField('Organization', coerce=int, validators=[Optional()])
    # Updated label for clarity based on your requirement
    department_id = SelectField('Company / Department (Required for "Client" role)', 
                                coerce=int, validators=[Optional()]) 
    submit = SubmitField('Save User')

    def validate_department_id(self, field):
        # This validation runs after individual field validators pass
        role_data = self.role.data # Get the submitted role
        org_id_data = self.organization_id.data # Get submitted org_id
        dept_id_data = field.data # This is self.department_id.data

        if role_data == 'client':
            if not org_id_data or org_id_data == 0:
                # If role is client, an organization must also be selected to assign a department
                # This might be redundant if department choices are already filtered by org,
                # but good as a server-side check.
                self.organization_id.errors.append('An Organization is required to assign a Company/Department to a "Client".')
                # We could also add an error to department_id itself, but org is the prerequisite.
            elif not dept_id_data or dept_id_data == 0:
                raise ValidationError('A specific Company/Department must be assigned to users with the "Client (Company Specific)" role.')
        
        if role_data == 'organization_client':
            if dept_id_data and dept_id_data != 0:
                raise ValidationError('Org Clients should not be assigned to a specific Company/Department; their scope is the entire Organization.')
            # Also ensure an organization_id is selected for organization_client
            if not org_id_data or org_id_data == 0:
                 self.organization_id.errors.append('An Organization must be assigned to "Org Client" roles.')


    def validate_password2(self, field):
        # Custom validation for password2 to make it required only if password is set
        # This is often better handled in the route, but can be done here too.
        # The existing EqualTo validator with Optional on both fields usually works well.
        # The route logic already handles making password fields required for new users.
        if self.password.data and not field.data:
            # This specific case might be if editing and only password is filled
            # For new users, password and password2 are made required by the route.
            user_id = request.form.get('_user_id_for_edit_check') # You'd need to pass this or similar
            if not user_id: # If it's a new user, this is already covered by DataRequired on password in route
                pass
            elif self.password.data: # If editing and password has data
                 raise ValidationError('Please confirm the new password.')


class UserSelfRegistrationForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=64)])
    email = StringField('Email Address', validators=[DataRequired(), Email(), Length(max=120)])
    password = PasswordField('Password', validators=[DataRequired(), Length(min=6)])
    password2 = PasswordField('Confirm Password', validators=[DataRequired(), EqualTo('password')])
    submit = SubmitField('Create Account')
    def validate_username(self, username):
        if User.query.filter_by(username=username.data).first(): raise ValidationError('Username taken.')
    def validate_email(self, email_field):
        if User.query.filter_by(email=email_field.data.lower()).first(): raise ValidationError('Email registered.')

# In app.py Forms section
class CreateTicketForm(FlaskForm):
    title = StringField('Subject*', validators=[DataRequired(), Length(max=100)])
    description = TextAreaField('Description*', validators=[DataRequired()])
    organization_id = SelectField('Organization', coerce=int, validators=[Optional()])
    department_id = SelectField('Company / Department', coerce=int, validators=[Optional()]) 
    form_type_id = SelectField('Form Type', coerce=int, validators=[Optional()])
    tags = StringField('Tags (comma-separated)', validators=[Optional(), Length(max=255)])
    category = SelectField('Issue Category*', coerce=int, validators=[InputRequired(message="Category is required.")])
    cloud_provider = SelectField('Cloud Provider', coerce=str, validators=[Optional()])
    severity = SelectField('Severity Level*', coerce=str, validators=[InputRequired(message="Severity is required.")])
    aws_service = SelectField('AWS Service (if AWS)', choices=AWS_SERVICE_CHOICES, validators=[Optional()])
    aws_account_id = StringField('AWS Account ID', validators=[
        Optional(), 
        Length(min=12, max=12, message="AWS Account ID must be 12 digits if provided."),
        Regexp(r'^\d{12}$', message="AWS Account ID must consist of 12 digits if provided.")
    ])
    environment = SelectField('Environment', coerce=str, validators=[Optional()])
    request_call_back = SelectField('Request Call Back?', choices=REQUEST_CALL_BACK_CHOICES, validators=[Optional()])
    contact_details = StringField('Contact Details for Callback', validators=[Optional(), Length(max=255)])
    # customer_name = StringField('Customer Company Name*', validators=[DataRequired(), Length(max=100)]) # REMOVED
    # support_modal_id = SelectField('Support Modals by Plan', coerce=int, validators=[Optional()]) # REMOVED
    additional_recipients = TextAreaField('Additional Email Recipients (comma-separated)', validators=[Optional()])
    attachments = MultipleFileField('Attachments', validators=[Optional()])
    submit = SubmitField('Submit Ticket')

    # validate_additional_recipients method remains the same
    def validate_additional_recipients(form, field):
        if field.data:
            emails = [email.strip() for email in field.data.split(',') if email.strip()]
            for email_str in emails:
                if '@' not in email_str or '.' not in email_str.split('@')[1]:
                    raise ValidationError(f"Invalid email address format: {email_str}")

class CommentForm(FlaskForm):
    content = TextAreaField('Your Comment', validators=[DataRequired()])
    is_internal = BooleanField('Internal Note (Agents/Admins Only)')
    submit = SubmitField('Add Comment')

# In app.py Forms section
# In app.py

class AgentUpdateTicketForm(FlaskForm):
    status = SelectField('Status', choices=TICKET_STATUS_CHOICES, validators=[DataRequired()])
    priority = SelectField('Priority', choices=TICKET_PRIORITY_CHOICES, validators=[DataRequired()])
    assigned_to_id = SelectField('Assign To Agent', coerce=int, validators=[Optional()])
    organization_id = SelectField('Organization', coerce=int, validators=[Optional()])
    # department_id = SelectField('Company / Department', coerce=int, validators=[Optional()]) 
    form_type_id = SelectField('Form Type', coerce=int, validators=[Optional()])
    tags = StringField('Tags (comma-separated)', validators=[Optional(), Length(max=255)])
    category_id = SelectField('Category', coerce=int, validators=[Optional()])
    cloud_provider = SelectField('Cloud Provider', coerce=str, validators=[Optional()]) # String is fine
    severity = SelectField('Severity', coerce=str, validators=[Optional()]) # String is fine
    aws_service = SelectField('AWS Service', choices=AWS_SERVICE_CHOICES, validators=[Optional()])
    aws_account_id = StringField('AWS Account ID', validators=[
        Optional(), 
        Length(min=12, max=12, message="AWS Account ID must be 12 digits if provided."),
        Regexp(r'^\d{12}$', message="AWS Account ID must consist of 12 digits if provided.")
    ])
    environment = SelectField('Environment', coerce=str, validators=[Optional()]) # String is fine
    request_call_back = SelectField('Request Call Back?', choices=REQUEST_CALL_BACK_CHOICES, validators=[Optional()])
    contact_details = StringField('Contact Details for Callback', validators=[Optional(), Length(max=255)])
    aws_support_case_id = StringField('AWS Support Case ID', validators=[Optional(), Length(max=50)])
    
    effort_required_to_resolve_min = SelectField(
        'Effort (min)', 
        choices=EFFORT_CHOICES, 
        validators=[Optional()],
        # Custom coerce to handle empty string from "---" option
        coerce=lambda x: int(x) if x is not None and x.strip() != '' and x.isdigit() else None
    )
    
    # customer_name = StringField('Customer Company Name', validators=[Optional(), Length(max=100)]) # REMOVED
    apn_opportunity_id = SelectField('APN Opportunities', coerce=int, validators=[Optional()])
    apn_opportunity_description = TextAreaField('APN Opportunities Description', validators=[Optional()])
    support_modal_id = SelectField('Support Modals by Plan', coerce=int, validators=[Optional()]) 
    additional_email_recipients = TextAreaField('Additional Email Recipients (comma-separated)', validators=[Optional()])
    submit = SubmitField('Update Ticket')

class CategoryForm(FlaskForm):
    name = StringField('Category Name', validators=[DataRequired(), Length(max=50)])
    description = StringField('Description', validators=[Optional(), Length(max=200)])
    submit = SubmitField('Save Category')

class CloudProviderOptionForm(FlaskForm):
    name = StringField('Cloud Provider Name', validators=[DataRequired(), Length(max=50)])
    is_active = BooleanField('Active', default=True)
    submit = SubmitField('Save Cloud Provider')

class SeverityOptionForm(FlaskForm):
    name = StringField('Severity Name', validators=[DataRequired(), Length(max=50)])
    description = StringField('Description (Optional)', validators=[Optional(), Length(max=100)])
    order = IntegerField('Sort Order (Optional)', validators=[Optional(), NumberRange(min=0)])
    is_active = BooleanField('Active', default=True)
    submit = SubmitField('Save Severity')

class EnvironmentOptionForm(FlaskForm):
    name = StringField('Environment Name', validators=[DataRequired(), Length(max=50)])
    is_active = BooleanField('Active', default=True)
    submit = SubmitField('Save Environment')

class OrganizationOptionForm(FlaskForm):
    name = StringField('Organization Name', validators=[DataRequired(), Length(max=100)])
    is_active = BooleanField('Active', default=True)
    submit = SubmitField('Save Organization')

class FormTypeOptionForm(FlaskForm):
    name = StringField('Form Type Name', validators=[DataRequired(), Length(max=50)])
    is_active = BooleanField('Active', default=True)
    submit = SubmitField('Save Form Type')

class APNOpportunityOptionForm(FlaskForm):
    name = StringField('APN Opportunity Name', validators=[DataRequired(), Length(max=100)])
    is_active = BooleanField('Active', default=True)
    submit = SubmitField('Save APN Opportunity')

class SupportModalOptionForm(FlaskForm):
    name = StringField('Support Modal Name', validators=[DataRequired(), Length(max=100)])
    is_active = BooleanField('Active', default=True)
    submit = SubmitField('Save Support Modal')

class ShareCredentialsForm(FlaskForm):
    recipient_email = StringField('Recipient Email', validators=[DataRequired(), Email()])
    
    
    
    
class AssignAgentOnlyForm(FlaskForm):
    assigned_to_id = SelectField('Assign To Agent', coerce=int, 
                                 validators=[InputRequired(message="Please select an agent or choose 'Unassign'.")])
    submit = SubmitField('Assign Agent')
    
    
# In app.py, within the --- Forms --- section

class DepartmentForm(FlaskForm):
    name = StringField('Department Name', validators=[DataRequired(), Length(max=100)])
    organization_id = SelectField('Parent Organization*', coerce=int, 
                                  validators=[InputRequired(message="Organization is required.")]) # Made it InputRequired
    is_active = BooleanField('Active', default=True)
    # Optional: Add a description field if you want it for departments
    # description = TextAreaField('Description (Optional)', validators=[Optional(), Length(max=255)])
    submit = SubmitField('Save Department')

    def __init__(self, *args, **kwargs):
        super(DepartmentForm, self).__init__(*args, **kwargs)
        # Populate organization choices, ensuring a valid selection is made
        # The value 0 is "--- Select Organization ---"
        self.organization_id.choices = get_active_organization_choices() # Uses your existing helper

    def validate_name(self, name):
        # Check for uniqueness within the selected organization
        organization_id = self.organization_id.data
        item_id = None
        if self.obj and hasattr(self.obj, 'id'): # self.obj is set by WTForms when editing
            item_id = self.obj.id
        
        if organization_id and organization_id != 0: # Ensure an organization is selected
            query = Department.query.filter_by(name=name.data, organization_id=organization_id)
            if item_id:
                query = query.filter(Department.id != item_id)
            if query.first():
                org = OrganizationOption.query.get(organization_id)
                org_name = org.name if org else "the selected organization"
                raise ValidationError(f'The department name "{name.data}" already exists in {org_name}.')
        elif not organization_id or organization_id == 0:
            # This case should be caught by InputRequired on organization_id,
            # but good to have a fallback or be aware of it.
            pass # Let InputRequired handle it

    def validate_organization_id(self, organization_id):
        if organization_id.data == 0: # 0 is the placeholder value
            raise ValidationError('Please select a parent organization.')

# --- Flask-Login, Context Processors, Decorators ---
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'
login_manager.login_message = "Please log in to access this page."












from flask import Blueprint, jsonify, request
from datetime import datetime

analytics_api_bp = Blueprint('analytics_api', __name__, url_prefix='/api/analytics')













# In app.py, near other admin create/edit routes
    
    





# Add this to your app.py, likely near other ticket-related routes

# In app.py

@app.route('/ticket/<int:ticket_id>/assign', methods=['GET', 'POST'])
@login_required 
def assign_ticket_page(ticket_id):
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        flash('Ticket not found.', 'danger')
        return redirect(url_for('dashboard'))

    # Use the new, simpler form specifically for assignment
    form = AssignAgentOnlyForm() 

    # Populate choices for the assigned_to_id dropdown
    agent_choices = [(u.id, u.username) for u in User.query.filter(User.role.in_(['agent', 'admin'])).order_by('username').all()]
    form.assigned_to_id.choices = [(0, '--- Unassign/Select Agent ---')] + agent_choices
    
    if request.method == 'GET':
        # Populate the form with the current assignee
        form.assigned_to_id.data = ticket.assigned_to_id or 0 # Use 0 for 'Unassign/Select Agent' if no one is assigned

    if form.validate_on_submit():
        old_assignee_id = ticket.assigned_to_id
        old_assignee_name = ticket.assignee.username if ticket.assignee else "Unassigned"
        
        new_assignee_id_from_form = form.assigned_to_id.data
        
        # Convert form value (0 for Unassign) to None for DB storage
        new_assignee_id_for_db = new_assignee_id_from_form if new_assignee_id_from_form != 0 else None

        if old_assignee_id != new_assignee_id_for_db:
            ticket.assigned_to_id = new_assignee_id_for_db
            
            new_assignee_obj = db.session.get(User, new_assignee_id_for_db) if new_assignee_id_for_db else None
            new_assignee_name = new_assignee_obj.username if new_assignee_obj else "Unassigned"
            
            log_interaction(ticket.id, 'ASSIGNMENT_CHANGE', user_id=current_user.id, 
                            details={'old_value': old_assignee_name, 'new_value': new_assignee_name, 'field_display_name': 'Assignee'})
            
            # If a ticket is assigned (not unassigned) and was 'Open', change its status to 'In Progress'
            if new_assignee_id_for_db and ticket.status == 'Open':
                old_status = ticket.status
                ticket.status = 'In Progress'
                log_interaction(ticket.id, 'STATUS_CHANGE', user_id=current_user.id,
                                details={'old_value': old_status, 'new_value': ticket.status, 'field_display_name': 'Status'})

            try:
                db.session.commit()
                flash(f'Ticket #{ticket.id} assignment updated to {new_assignee_name}.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating ticket assignment: {str(e)}', 'danger')
                app.logger.error(f"Error updating assignment for ticket {ticket.id}: {e}", exc_info=True)
        else:
            flash('No change in assignment.', 'info')
        
        # Always redirect to the view_ticket page after processing (success, error, or no change)
        return redirect(url_for('view_ticket', ticket_id=ticket.id))

    # If form validation fails on POST (e.g., somehow no agent selected, though InputRequired should catch)
    # or if it's a GET request, render the template.
    return render_template('agent/assign_ticket.html', 
                           title=f"Assign Ticket #{ticket.id}", 
                           ticket=ticket,
                           form=form)





# In app.py - --- Routes --- section

# I



@login_manager.user_loader
def load_user(user_id): 
    return db.session.get(User, int(user_id))

@app.context_processor
def inject_global_vars():
    return {
        'current_year': datetime.utcnow().year,
        'app': app,
        'to_snake_case': to_snake_case,
        'EFFORT_CHOICES': EFFORT_CHOICES,
        'active_category_choices_gdoc': get_active_category_choices(),
        'active_severity_choices_gdoc': get_active_severity_choices(),
    }

def admin_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            flash('Admin access is required to view this page.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def agent_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not (current_user.is_agent or current_user.is_admin):
            flash('Agent or Admin access is required to view this page.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# --- Helper functions ---
DOMAIN_TO_ORGANIZATION_MAP = {
    'cloudkeeper.com': 'CloudKeeper (CK)',
    'jietjodhpur.ac.in': 'JIET Jodhpur' 
}

def get_organization_by_email_domain(email, auto_create=True):
    if '@' not in email: return None
    domain = email.split('@')[-1].lower()
    
    org_name_from_map = DOMAIN_TO_ORGANIZATION_MAP.get(domain)
    if org_name_from_map:
        organization = OrganizationOption.query.filter_by(name=org_name_from_map, is_active=True).first()
        if organization: return organization
        elif auto_create: 
            app.logger.info(f"Mapped organization '{org_name_from_map}' not found. Creating for domain '{domain}'.")
            new_org = OrganizationOption(name=org_name_from_map, is_active=True); db.session.add(new_org)
            db.session.flush() 
            return new_org 
    
    if auto_create:
        parts = domain.split('.'); potential_org_name_base = parts[-2] if len(parts) > 1 else parts[0]
        if len(parts) > 2 and parts[-2] in ['ac', 'co', 'com', 'org', 'gov', 'edu']: potential_org_name_base = parts[-3]
        potential_org_name = potential_org_name_base.replace('-', ' ').title()
        existing_org_by_derived_name = OrganizationOption.query.filter(OrganizationOption.name.ilike(potential_org_name)).first()
        if existing_org_by_derived_name: return existing_org_by_derived_name
        app.logger.info(f"No mapping/existing org for domain '{domain}'. Auto-creating '{potential_org_name}'.")
        new_org = OrganizationOption(name=potential_org_name, is_active=True); db.session.add(new_org)
        db.session.flush() 
        return new_org
    return None

def get_active_choices(model_class, placeholder_text_id_0=None, placeholder_text_str_empty=None, order_by_attr='name'):
    query = model_class.query.filter_by(is_active=True).order_by(getattr(model_class, order_by_attr))
    choices = []
    if placeholder_text_id_0: choices.append((0, placeholder_text_id_0))
    elif placeholder_text_str_empty: choices.append(('', placeholder_text_str_empty))
    if placeholder_text_id_0: choices.extend([(opt.id, opt.name) for opt in query.all()])
    else: choices.extend([(opt.name, opt.name) for opt in query.all()])
    return choices


def get_active_department_choices_for_org(organization_id=None):
    choices = [(0, '--- Select Department (Requires Org) ---')]
    if organization_id and organization_id != 0:
        departments = Department.query.filter_by(organization_id=organization_id, is_active=True).order_by(Department.name).all()
        choices.extend([(dept.id, dept.name) for dept in departments])
    # If no organization is selected, or if selected org has no departments, the list will just have the placeholder.
    # Or, you could fetch all active departments if no org is selected, but that might be confusing.
    # This approach makes it clear a department belongs to an organization.
    return choices

def get_active_category_choices():
    return [(0, '--- Select Issue Category* ---')] + [(c.id, c.name) for c in Category.query.order_by('name').all()]
def get_active_cloud_provider_choices():
    return get_active_choices(CloudProviderOption, placeholder_text_str_empty='--- Select Cloud Provider ---')
def get_active_severity_choices():
    return get_active_choices(SeverityOption, placeholder_text_str_empty='--- Select Severity* ---', order_by_attr='order')
def get_active_environment_choices():
    return get_active_choices(EnvironmentOption, placeholder_text_str_empty='--- Select Environment ---')
def get_active_organization_choices():
    return get_active_choices(OrganizationOption, placeholder_text_id_0='--- Select Organization ---')
def get_active_form_type_choices():
    return get_active_choices(FormTypeOption, placeholder_text_id_0='--- Select Dificulty-Level ---')
def get_active_apn_opportunity_choices():
    return get_active_choices(APNOpportunityOption, placeholder_text_id_0='--- Select APN Opportunity ---')
def get_active_support_modal_choices():
    return get_active_choices(SupportModalOption, placeholder_text_id_0='--- Select Support Modal ---')

def log_interaction(ticket_id, interaction_type, user_id=None, details=None, timestamp_override=None, commit_now=False):
    actual_user_id = user_id
    if actual_user_id is None and current_user and current_user.is_authenticated:
        actual_user_id = current_user.id
    interaction_timestamp = timestamp_override if timestamp_override else datetime.utcnow()
    interaction = Interaction(ticket_id=ticket_id, user_id=actual_user_id, interaction_type=interaction_type, details=details or {}, timestamp=interaction_timestamp)
    db.session.add(interaction)
    if commit_now:
        try: db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Failed to commit interaction log for ticket {ticket_id}: {e}", exc_info=True)

def trigger_priority_call_alert(ticket, old_severity=None):
    account_sid = app.config.get('TWILIO_ACCOUNT_SID')
    auth_token = app.config.get('TWILIO_AUTH_TOKEN')
    twilio_phone_number = app.config.get('TWILIO_PHONE_NUMBER')
    recipient_phone_number = app.config.get('EMERGENCY_CALL_RECIPIENT_PHONE_NUMBER')
    alert_severities = app.config.get('SEVERITIES_FOR_CALL_ALERT', [])
    current_ticket_severity = ticket.severity

    app.logger.info(f"--- trigger_priority_call_alert (Using simpler logic) ---")
    app.logger.info(f"Ticket ID: {ticket.id}, Current Severity: '{current_ticket_severity}', Old Severity: '{old_severity}'")
    app.logger.info(f"Alert Severities List: {alert_severities}")
    app.logger.info(f"Twilio Config: SID set: {bool(account_sid)}, Token set: {bool(auth_token)}, From #: {twilio_phone_number}, To #: {recipient_phone_number}")

    if not all([account_sid, auth_token, twilio_phone_number, recipient_phone_number]):
        app.logger.warning(f"Twilio not fully configured. Skipping call alert for ticket #{ticket.id}.")
        return

    if current_ticket_severity not in alert_severities:
        app.logger.info(f"Ticket #{ticket.id} severity '{current_ticket_severity}' is not in the alert list {alert_severities}. Skipping call.")
        return
    
    alert_description = f"Ticket {ticket.id} ({ticket.title}) has severity {current_ticket_severity}."
    if old_severity and old_severity != current_ticket_severity:
        # This case means it was already alertable and changed, or escalated into alertable
        if current_ticket_severity in alert_severities and old_severity not in alert_severities:
            alert_description = f"Ticket {ticket.id} ({ticket.title}) severity escalated from {old_severity} to {current_ticket_severity}."
        elif current_ticket_severity in alert_severities and old_severity in alert_severities : # Changed between alertable severities
            alert_description = f"Ticket {ticket.id} ({ticket.title}) severity changed from {old_severity} to {current_ticket_severity}."
        # else: # de-escalated from alertable to non-alertable, or changed between non-alertable (already handled by first if)
    elif not old_severity: # New ticket
        alert_description = f"New high priority ticket {ticket.id} ({ticket.title}) created with severity {current_ticket_severity}."
    else: # old_severity == new_severity and new_severity is alertable (no change, was already high)
        app.logger.info(f"Ticket #{ticket.id} severity '{current_ticket_severity}' unchanged and already alertable. No new call.")
        return


    app.logger.info(f"Proceeding to make Twilio call for ticket #{ticket.id}. Description: {alert_description}")
    try:
        client = TwilioClient(account_sid, auth_token)
        sanitized_title = re.sub(r'[^\w\s,.-]', '', ticket.title) 
        
        message_to_say = (
            f"Hello. This is an urgent alert from the Ticket System. "
            f"{alert_description} "
            f"Please check the system immediately."
        )
        twiml_instruction = f'<Response><Say>{escape(message_to_say)}</Say></Response>'
        
        call = client.calls.create(
            twiml=twiml_instruction,
            to=recipient_phone_number,
            from_=twilio_phone_number
        )
        app.logger.info(f"Twilio call initiated for ticket #{ticket.id} to {recipient_phone_number}. Call SID: {call.sid}")
        flash(f'High priority ticket #{ticket.id} alerted via call to {recipient_phone_number}. ({alert_description})', 'success') 
    except TwilioRestException as e:
        app.logger.error(f"Twilio API error for ticket #{ticket.id}: {e}")
        error_message_detail = getattr(e, 'msg', str(e)) # Safely get msg or fallback to str(e)
        flash(f'Error initiating Twilio call for ticket #{ticket.id}: {error_message_detail}', 'danger')
    except Exception as e:
        app.logger.error(f"Unexpected error during Twilio call for ticket #{ticket.id}: {e}", exc_info=True)
        flash(f'An unexpected error occurred while trying to initiate a call for ticket #{ticket.id}.', 'danger')

# --- Routes ---
@app.route('/')
@app.route('/index')
def index():
    if current_user.is_authenticated: return redirect(url_for('dashboard'))
    return render_template('index.html', title='Welcome')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        next_page = request.args.get('next')
        if next_page and not (next_page.startswith('//') or '://' in next_page): return redirect(next_page)
        return redirect(url_for('dashboard'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember_me.data); next_page = request.args.get('next')
            if next_page and not (next_page.startswith('//') or '://' in next_page): return redirect(next_page)
            return redirect(url_for('dashboard'))
        else: flash('Invalid username or password.', 'danger')
    return render_template('login.html', title='Login', form=form)

@app.route('/logout')
@login_required
def logout(): logout_user(); flash('You have been logged out.', 'info'); return redirect(url_for('index'))

@app.after_request
def add_header(response):
    if '/static/' not in request.path: response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, private, max-age=0'; response.headers['Pragma'] = 'no-cache'; response.headers['Expires'] = '0'
    return response

@app.route('/register/client', methods=['GET', 'POST'])
def register_client():
    if current_user.is_authenticated: return redirect(url_for('dashboard'))
    form = UserSelfRegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data.lower(), role='client'); user.set_password(form.password.data)
        organization = get_organization_by_email_domain(user.email, auto_create=True) 
        if organization: 
            if not organization.id: 
                db.session.flush() 
            user.organization_id = organization.id 
            app.logger.info(f"User '{user.username}' associated with organization '{organization.name}'.")
        else: app.logger.info(f"No organization could be determined or created for domain of '{user.email}'.")
        db.session.add(user)
        try: 
            db.session.commit() 
            flash('Client account created successfully! Please log in.', 'success'); return redirect(url_for('login'))
        except Exception as e: db.session.rollback(); flash('Error during registration. Please try again.', 'danger'); app.logger.error(f"Client registration error: {e}", exc_info=True)
    return render_template('register_user.html', title='Register as Client', form=form, registration_type='Client', info_text='Submit and track your support tickets.')

@app.route('/register/agent', methods=['GET', 'POST'])
@admin_required
def register_agent():
    form = UserSelfRegistrationForm(); 
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data.lower(), role='agent'); user.set_password(form.password.data)
        db.session.add(user)
        try: db.session.commit(); flash('Agent account created successfully!', 'success'); return redirect(url_for('admin_user_list'))
        except Exception as e: db.session.rollback(); flash('Error during agent registration. Please try again.', 'danger'); app.logger.error(f"Admin agent registration error: {e}", exc_info=True)
    return render_template('register_user.html', title='Register New Agent', form=form, registration_type='Agent', info_text='Register new support agents to assist clients.')












# ... (other imports and code) ...

@app.route('/dashboard')
@login_required
def dashboard():
    page_title = "My Dashboard"
    
    if current_user.is_admin: 
        page_title = 'Admin Dashboard'
        # ... (Admin dashboard logic as previously defined) ...
        total_tickets_count = Ticket.query.count()
        open_tickets_count = Ticket.query.filter_by(status='Open').count()
        inprogress_tickets_count = Ticket.query.filter_by(status='In Progress').count()
        resolved_tickets_count = Ticket.query.filter_by(status='Resolved').count()
        on_hold_tickets_count = Ticket.query.filter_by(status='On Hold').count()
        closed_tickets_count = Ticket.query.filter_by(status='Closed').count()
        unassigned_tickets_count = Ticket.query.filter_by(assigned_to_id=None, status='Open').count()
        stats = {
            'total_tickets': total_tickets_count, 
            'open_tickets': open_tickets_count, 
            'inprogress_tickets': inprogress_tickets_count, 
            'resolved_tickets': resolved_tickets_count,
            'on_hold_tickets': on_hold_tickets_count, 
            'closed_tickets': closed_tickets_count,   
            'unassigned_open_tickets': unassigned_tickets_count, 
            'total_users': User.query.count(),
            'total_agents': User.query.filter(User.role.in_(['agent', 'admin'])).count(), 
            'total_clients': User.query.filter_by(role='client').count(), 
            'active_categories_count': Category.query.count() 
        }
        recent_interactions = Interaction.query.order_by(Interaction.timestamp.desc()).limit(7).all()
        return render_template('dashboard.html', title=page_title, recent_interactions=recent_interactions, **stats)

    elif current_user.is_agent: 
        page_title = 'Agent Dashboard'
        agent_id = current_user.id
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=today_start.weekday())

        # Agent's Personal Stats
        resolved_today_count = Ticket.query.filter(
            Ticket.assigned_to_id == agent_id,
            Ticket.status.in_(['Resolved', 'Closed']),
            Ticket.resolved_at >= today_start 
        ).count()
        resolved_week_count = Ticket.query.filter(
            Ticket.assigned_to_id == agent_id,
            Ticket.status.in_(['Resolved', 'Closed']),
            Ticket.resolved_at >= week_start
        ).count()

        responded_today_ticket_ids = db.session.query(Comment.ticket_id).distinct().filter(
            Comment.user_id == agent_id,
            Comment.is_internal == False,
            Comment.created_at >= today_start,
            Comment.ticket_id.in_(db.session.query(Ticket.id).filter(Ticket.assigned_to_id == agent_id))
        ).all()
        responded_week_ticket_ids = db.session.query(Comment.ticket_id).distinct().filter(
            Comment.user_id == agent_id,
            Comment.is_internal == False,
            Comment.created_at >= week_start,
            Comment.ticket_id.in_(db.session.query(Ticket.id).filter(Ticket.assigned_to_id == agent_id))
        ).all()
        responded_today_count = len(responded_today_ticket_ids)
        responded_week_count = len(responded_week_ticket_ids)

        agent_stats = {
            'resolved_today': resolved_today_count,
            'resolved_week': resolved_week_count,
            'responded_today': responded_today_count,
            'responded_week': responded_week_count,
        }

        # Recent Client Replies on Agent's Tickets
        LatestCommentSubquery = db.session.query(
            Comment.ticket_id,
            func.max(Comment.created_at).label('latest_comment_at')
        ).filter(
            Comment.ticket_id.in_(
                db.session.query(Ticket.id).filter(Ticket.assigned_to_id == agent_id)
            )
        ).group_by(Comment.ticket_id).subquery()

        tickets_with_client_reply_data = db.session.query(
            Ticket, 
            User.username.label('last_commenter_username'),
            Comment.created_at.label('last_comment_created_at') # Fetch the timestamp of the specific comment
        ) \
            .join(LatestCommentSubquery, Ticket.id == LatestCommentSubquery.c.ticket_id) \
            .join(Comment, (Comment.ticket_id == LatestCommentSubquery.c.ticket_id) & \
                           (Comment.created_at == LatestCommentSubquery.c.latest_comment_at)) \
            .join(User, Comment.user_id == User.id) \
            .filter(Ticket.assigned_to_id == agent_id) \
            .filter(User.role == 'client') \
            .filter(Ticket.status.notin_(['Resolved', 'Closed'])) \
            .order_by(LatestCommentSubquery.c.latest_comment_at.desc()) \
            .limit(5).all() # This returns Row objects (like tuples)

        agent_data = {
            'my_assigned_tickets': Ticket.query.filter_by(assigned_to_id=agent_id).filter(Ticket.status.notin_(['Resolved', 'Closed'])).order_by(Ticket.updated_at.desc()).limit(5).all(), 
            'unassigned_tickets': Ticket.query.filter_by(assigned_to_id=None, status='Open').order_by(Ticket.created_at.desc()).limit(5).all(),
            'agent_stats': agent_stats,
            'tickets_with_client_reply_data': tickets_with_client_reply_data # Use the new variable
        }
        return render_template('dashboard.html', title=page_title, **agent_data)
    
    else: # Client
        page_title = 'My Dashboard'
        my_tickets = Ticket.query.filter_by(created_by_id=current_user.id).order_by(Ticket.updated_at.desc()).limit(10).all()
        return render_template('dashboard.html', title=page_title, my_tickets=my_tickets)







# In app.py - --- Forms --- section

class RequestPasswordResetForm(FlaskForm):
    email = StringField('Email Address', validators=[DataRequired(), Email(), Length(max=120)])
    submit = SubmitField('Request Password Reset')

    def validate_email(self, email):
        user = User.query.filter_by(email=email.data.lower()).first()
        if not user:
            # Don't reveal if email exists for security, but log it for admin
            current_app.logger.info(f"Password reset request for non-existent email: {email.data}")
            # Raise a generic error or just let it pass and show a generic success message
            # For better UX, sometimes a generic "If that email is in our system..." message is shown
            # For now, we'll just proceed and show success if no errors
            pass 
            # raise ValidationError('There is no account with that email. You must register first.')


class PasswordResetForm(FlaskForm):
    password = PasswordField('New Password', validators=[DataRequired(), Length(min=6)])
    password2 = PasswordField('Confirm New Password', validators=[DataRequired(), EqualTo('password')])
    submit = SubmitField('Reset Password')




def send_password_reset_email(user):
    token = user.get_reset_password_token()
    # Ensure you have MAIL_DEFAULT_SENDER_EMAIL configured or fallback
    sender_email = current_app.config.get('MAIL_DEFAULT_SENDER_EMAIL') or current_app.config.get('MAIL_USERNAME')
    if not sender_email:
        current_app.logger.error("MAIL_DEFAULT_SENDER_EMAIL or MAIL_USERNAME not configured. Cannot send password reset email.")
        return False

    msg = Message('Password Reset Request - Ticket CMS',
                  sender=('TicketSys Admin', sender_email), # Use tuple for sender name and email
                  recipients=[user.email])
    reset_url = url_for('reset_password_with_token', token=token, _external=True)
    msg.body = f'''To reset your password, visit the following link:
{reset_url}

If you did not make this request then simply ignore this email and no changes will be made.
This link will expire in 30 minutes.
'''
    # Optional: Create an HTML version (msg.html)
    # msg.html = render_template('email/reset_password_email.html', user=user, reset_url=reset_url)
    try:
        mail.send(msg)
        current_app.logger.info(f"Password reset email sent to {user.email}")
        return True
    except Exception as e:
        current_app.logger.error(f"Failed to send password reset email to {user.email}: {e}", exc_info=True)
        return False


@app.route('/reset_password_request', methods=['GET', 'POST'])
def reset_password_request():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    form = RequestPasswordResetForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data.lower()).first()
        if user:
            if send_password_reset_email(user):
                flash('An email has been sent with instructions to reset your password.', 'info')
            else:
                flash('There was an issue sending the password reset email. Please try again later or contact support.', 'danger')
        else:
            # To prevent email enumeration, show a generic success message even if email not found.
            # Log the attempt for admin review.
            current_app.logger.info(f"Password reset requested for non-existent or unverified email: {form.email.data}")
            flash('If an account with that email exists, instructions to reset your password have been sent.', 'info')
        return redirect(url_for('login')) # Always redirect to login to obscure if email was valid
    return render_template('auth/request_reset.html', title='Reset Password Request', form=form)


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password_with_token(token):
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    user = User.verify_reset_password_token(token)
    if not user:
        flash('That is an invalid or expired token. Please request a new one.', 'warning')
        return redirect(url_for('reset_password_request'))
    
    form = PasswordResetForm()
    if form.validate_on_submit():
        user.set_password(form.password.data)
        db.session.commit()
        log_interaction(None, 'PASSWORD_RESET_COMPLETED', user_id=user.id, details={'username': user.username}) # Log to general system log or user activity if ticket_id is None
        flash('Your password has been reset successfully! You can now log in.', 'success')
        return redirect(url_for('login'))
    return render_template('auth/reset_password.html', title='Reset Your Password', form=form, token=token)







# In app.py, near other admin create/edit routes

@app.route('/admin/department/new', methods=['GET', 'POST'])
@app.route('/admin/department/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_department(item_id=None):
    item = db.session.get(Department, item_id) if item_id else None
    form = DepartmentForm(obj=item) # Pass obj for pre-population on GET

    legend = 'New Department' if not item else f'Edit Department: {item.name}'

    if form.validate_on_submit():
        is_new = (item is None)
        department_instance = item or Department()

        # Check uniqueness (name within organization_id) - already handled by form.validate_name

        department_instance.name = form.name.data
        department_instance.organization_id = form.organization_id.data
        department_instance.is_active = form.is_active.data
        # if hasattr(form, 'description'): # If you added description to the form
        #     department_instance.description = form.description.data
            
        if is_new:
            db.session.add(department_instance)
        
        try:
            db.session.commit()
            flash(f'Department "{department_instance.name}" saved.', 'success')
            return redirect(url_for('admin_department_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Database error saving department: {str(e)}', 'danger')
            current_app.logger.error(f"Error saving Department {department_instance.name}: {e}", exc_info=True)
            # Repopulate choices if commit fails and form is re-rendered (though less common here)
            form.organization_id.choices = get_active_organization_choices()


    # If GET request or validation failed, re-populate choices if necessary (though form init does it)
    if not form.organization_id.choices or len(form.organization_id.choices) <= 1 : # Check if choices are missing/minimal
         form.organization_id.choices = get_active_organization_choices()
    
    # If editing, ensure the current organization is selected in the dropdown
    if request.method == 'GET' and item and item.organization_id:
        form.organization_id.data = item.organization_id

    return render_template('admin/create_edit_option.html', 
                           title=legend, 
                           form=form, 
                           legend=legend, 
                           item_type_name="Department",  # For display in template
                           list_url=url_for('admin_department_list'))
    
    
    
# In app.py

@app.route('/admin/department/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_department(item_id):
    department = db.session.get(Department, item_id)
    if not department:
        flash('Department not found.', 'danger')
        return redirect(url_for('admin_department_list'))

    department_name = department.name
    
    # Check for dependencies before deleting
    if department.users.count() > 0:
        flash(f'Cannot delete department "{department_name}" as it has associated users.', 'danger')
        return redirect(url_for('admin_department_list'))
    if department.tickets.count() > 0:
        flash(f'Cannot delete department "{department_name}" as it has associated tickets.', 'danger')
        return redirect(url_for('admin_department_list'))

    try:
        db.session.delete(department)
        db.session.commit()
        flash(f'Department "{department_name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting department "{department_name}": {str(e)}', 'danger')
        current_app.logger.error(f"Error deleting department {item_id}: {e}", exc_info=True)
    
    return redirect(url_for('admin_department_list'))




# In app.py

@app.route('/reports/overview')
@admin_required 
def reports_overview():
    all_statuses_data = [{'value': s[0], 'display': s[1]} for s in TICKET_STATUS_CHOICES]
    all_priorities_data = [{'value': p[0], 'display': p[1]} for p in TICKET_PRIORITY_CHOICES]
    all_categories_data = Category.query.order_by(Category.name).all()
    all_organizations_data = OrganizationOption.query.filter_by(is_active=True).order_by(OrganizationOption.name).all()
    all_agents_data = User.query.filter(User.role.in_(['agent', 'admin'])).order_by(User.username).all()

    # For pre-populating department filter if an org is already selected via query params for the page load
    selected_organization_id_str = request.args.get('organization_id')
    departments_for_filter_initial = []
    if selected_organization_id_str and selected_organization_id_str.isdigit() and selected_organization_id_str != 'all': # 'all' or '0' could be used for "All Orgs"
        try:
            org_id = int(selected_organization_id_str)
            if org_id != 0: # Check it's not the "All Organizations" placeholder value if you use 0 for that
                departments_for_filter_initial = Department.query.filter_by(
                    organization_id=org_id, 
                    is_active=True
                ).order_by(Department.name).all()
        except ValueError:
            pass # Invalid org_id in query param

    return render_template('reports/overview.html', 
                           title="Reports Overview",
                           all_statuses=all_statuses_data,
                           all_priorities=all_priorities_data,
                           all_categories=all_categories_data,
                           all_organizations=all_organizations_data,
                           all_agents=all_agents_data,
                           initial_departments_for_filter=departments_for_filter_initial # Pass for initial load
                           )
# Example: API endpoint for a specific predefined report (Ticket List Report)
# In app.py

@app.route('/api/reports/ticket_list_data')
@admin_required
def api_report_ticket_list_data():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)

    query = Ticket.query
    
    # _apply_common_filters will now also handle department_id if present in request.args
    query = _apply_common_filters(query, request.args, model_to_filter=Ticket)

    sort_by = request.args.get('sort_by', 'created_at')
    sort_order = request.args.get('sort_order', 'desc')

    if hasattr(Ticket, sort_by):
        column_to_sort = getattr(Ticket, sort_by)
        if sort_order == 'asc':
            query = query.order_by(column_to_sort.asc())
        else:
            query = query.order_by(column_to_sort.desc())
    else: 
        query = query.order_by(Ticket.created_at.desc())

    tickets_pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    tickets_data = []
    for ticket in tickets_pagination.items:
        tickets_data.append({
            'id': ticket.id,
            'title': ticket.title,
            'status': ticket.status,
            'priority': ticket.priority,
            'category': ticket.category_ref.name if ticket.category_ref else 'N/A',
            'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M'),
            'updated_at': ticket.updated_at.strftime('%Y-%m-%d %H:%M'),
            'resolved_at': ticket.resolved_at.strftime('%Y-%m-%d %H:%M') if ticket.resolved_at else 'N/A',
            'created_by': ticket.creator.username if ticket.creator else 'N/A',
            'assigned_to': ticket.assignee.username if ticket.assignee else 'Unassigned',
            'organization': ticket.organization_option_ref.name if ticket.organization_option_ref else 'N/A',
            'department': ticket.department_ref.name if ticket.department_ref else 'N/A' # <-- ADDED DEPARTMENT
        })

    return jsonify({
        'tickets': tickets_data,
        'total': tickets_pagination.total,
        'pages': tickets_pagination.pages,
        'current_page': tickets_pagination.page,
        'has_next': tickets_pagination.has_next,
        'has_prev': tickets_pagination.has_prev,
        'per_page': tickets_pagination.per_page
    })




# In app.py
# In app.py

# ... (Ensure these helpers are correctly defined as before)
# def get_active_organization_choices(): ...
# def get_active_department_choices_for_org(organization_id=None): ...

# In app.py

# Ensure all necessary imports are present at the top of your app.py:
# from .models import User, OrganizationOption, Department, Ticket, Category, Attachment # etc.
# from .forms import CreateTicketForm
# from .helpers import get_active_category_choices, get_active_cloud_provider_choices, # etc.
#                      get_active_organization_choices, get_active_department_choices_for_org, log_interaction
# from flask import render_template, redirect, url_for, flash, request, current_app
# from flask_login import login_required, current_user
# from werkzeug.utils import secure_filename
# import uuid
# from datetime import datetime
# import os

@app.route('/tickets/new', methods=['GET', 'POST'])
@login_required
def create_ticket():
    form = CreateTicketForm() # Uses updated form without direct customer_name/support_modal input

    # --- ALWAYS POPULATE CHOICES for fields not dependent on other form inputs directly for their *choices list* ---
    form.category.choices = get_active_category_choices()
    form.cloud_provider.choices = get_active_cloud_provider_choices()
    form.severity.choices = get_active_severity_choices()
    form.environment.choices = get_active_environment_choices()
    form.form_type_id.choices = get_active_form_type_choices()
    # form.aws_service.choices are static in its definition.
    # form.request_call_back.choices are static in its definition.

    user_org = current_user.organization
    user_dept = current_user.department

    # --- Dynamically set choices for organization_id and department_id ---
    # This needs to happen for ALL request methods (GET/POST) before validation.

    # 1. Set choices for organization_id
    if current_user.role == 'client' and user_dept and user_org:
        form.organization_id.choices = [(user_org.id, user_org.name)]
    elif current_user.role == 'organization_client' and user_org:
        form.organization_id.choices = [(user_org.id, user_org.name)]
    else: # Admin, Agent, or client not fully associated
        form.organization_id.choices = get_active_organization_choices()

    # 2. Set choices for department_id based on the context (user role and organization selection)
    org_id_for_department_choices = None
    if current_user.role == 'client' and user_dept and user_org:
        form.department_id.choices = [(user_dept.id, user_dept.name)]
        # org_id_for_department_choices is implicitly user_org.id here
    elif current_user.role == 'organization_client' and user_org:
        form.department_id.choices = get_active_department_choices_for_org(user_org.id)
        org_id_for_department_choices = user_org.id
    else: # Admin/Agent: department choices depend on the selected organization_id
        # For POST: use the submitted value of organization_id from request.form
        # For GET: use form.organization_id.data if pre-filled, else None
        if request.method == 'POST':
            submitted_org_id_str = request.form.get(form.organization_id.name)
            if submitted_org_id_str and submitted_org_id_str.isdigit() and int(submitted_org_id_str) != 0:
                org_id_for_department_choices = int(submitted_org_id_str)
        elif request.method == 'GET': # If form was pre-populated (e.g. edit, or failed POST re-render)
            if form.organization_id.data and form.organization_id.data != 0:
                org_id_for_department_choices = form.organization_id.data
        
        form.department_id.choices = get_active_department_choices_for_org(org_id_for_department_choices)

    # --- User-specific context for form rendering and initial data population on GET ---
    render_org_as_readonly = False
    render_dept_as_readonly = False
    template_org_name = None
    template_dept_name = None
    derived_customer_name_for_template = None

    if request.method == 'GET':
        if current_user.role == 'client' and user_dept and user_org:
            form.organization_id.data = user_org.id # Pre-fill data
            form.department_id.data = user_dept.id   # Pre-fill data
            render_org_as_readonly = True; template_org_name = user_org.name
            render_dept_as_readonly = True; template_dept_name = user_dept.name
            derived_customer_name_for_template = user_dept.name
        elif current_user.role == 'organization_client' and user_org:
            form.organization_id.data = user_org.id # Pre-fill data
            # department_id.data is not pre-set; they choose.
            render_org_as_readonly = True; template_org_name = user_org.name
            render_dept_as_readonly = False
            derived_customer_name_for_template = user_org.name
        else: # Admin, Agent, or a Client not yet fully associated
            render_org_as_readonly = False 
            render_dept_as_readonly = False
            # `derived_customer_name_for_template` will be based on selection or prompt in template.
            # No .data pre-fill for org/dept for Admin/Agent on new GET

    if form.validate_on_submit():
        ticket_org_id = None
        ticket_dept_id = None
        final_customer_name_for_ticket = None

        # 1. Determine Organization ID for the ticket (using form.organization_id.data after validation)
        if current_user.role == 'client' and user_org:
            ticket_org_id = user_org.id
        elif current_user.role == 'organization_client' and user_org:
            ticket_org_id = user_org.id
        elif form.organization_id.data and form.organization_id.data != 0:
            ticket_org_id = form.organization_id.data
        
        # 2. Determine Department ID for the ticket (using form.department_id.data after validation)
        if current_user.role == 'client' and user_dept:
            ticket_dept_id = user_dept.id
        elif form.department_id.data and form.department_id.data != 0:
            # Validation that dept belongs to org (if org is also set)
            if ticket_org_id:
                dept_check = db.session.get(Department, form.department_id.data)
                if dept_check and dept_check.organization_id == ticket_org_id:
                    ticket_dept_id = dept_check.id
                else:
                    form.department_id.errors.append("Selected Company/Department does not belong to the ticket's Organization.")
            # else: # Department selected without organization (should be caught by form logic for dependent fields)
                 # If ticket_org_id is None but form.department_id.data is present, it's an issue.
                 # This case is tricky because a department *must* have an org.
                 # If form.organization_id.data was '0' or None, then ticket_org_id would be None.
                 # If department was still selected, this implies an issue in form state or choices.
                 # The choices for department should have been empty if no org was selected.
                 # So if form.department_id.data has a value here, org must have been selected.
        
        # 3. Derive final_customer_name_for_ticket
        if ticket_dept_id:
            dept_obj_for_name = db.session.get(Department, ticket_dept_id)
            if dept_obj_for_name: final_customer_name_for_ticket = dept_obj_for_name.name
        elif ticket_org_id:
            org_obj_for_name = db.session.get(OrganizationOption, ticket_org_id)
            if org_obj_for_name: final_customer_name_for_ticket = org_obj_for_name.name
        
        if not final_customer_name_for_ticket:
            err_msg = "Customer association unclear. Please select an Organization or Company/Department."
            can_select_org_or_dept = not ((current_user.role == 'client' and user_dept and user_org) or \
                                        (current_user.role == 'organization_client' and user_org and not form.department_id.choices)) # Check if fields were actually selectable
            if can_select_org_or_dept:
                 form.organization_id.errors.append(err_msg) # Add to a visible field if selectable
            else:
                app.logger.error(f"Could not derive customer_name for user {current_user.id} creating ticket. This is unexpected.")
                flash("Error: Could not determine customer association for this ticket.", "danger")

        # Attachment handling logic
        uploaded_files_info = []
        if not form.errors: 
            if form.attachments.data:
                for file_storage in form.attachments.data:
                    if file_storage and file_storage.filename:
                        # ... (rest of attachment saving logic - unchanged) ...
                        filename = secure_filename(file_storage.filename)
                        if allowed_file(filename):
                            unique_suffix = uuid.uuid4().hex[:8]
                            stored_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{unique_suffix}_{filename}"
                            file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], stored_filename)
                            try:
                                file_storage.save(file_path)
                                uploaded_files_info.append({
                                    'original_filename': filename,
                                    'stored_filename': stored_filename,
                                    'content_type': file_storage.content_type
                                })
                            except Exception as e:
                                current_app.logger.error(f"Failed to save attachment {filename}: {e}", exc_info=True)
                                form.attachments.errors.append(f"Could not save file: {filename}. {e}")
                        else: 
                            form.attachments.errors.append(f"File type not allowed: {filename}")

        if form.errors: 
            flash('Please correct the errors in the form.', 'danger')
            # Re-populate dynamic choices and template vars for re-render
            # The choices for category, cloud_provider, etc., are already set at the top.
            # Org/Dept choices were also set at the top. What needs re-evaluation here are the
            # readonly flags and template names for display based on the *current* user state.
            # The `form.field.data` values will be the submitted (invalid) ones.
            
            # Re-evaluate render flags and template names for re-displaying the form
            if current_user.role == 'client' and user_dept and user_org:
                render_org_as_readonly = True; template_org_name = user_org.name
                render_dept_as_readonly = True; template_dept_name = user_dept.name
                derived_customer_name_for_template = user_dept.name
            elif current_user.role == 'organization_client' and user_org:
                render_org_as_readonly = True; template_org_name = user_org.name
                render_dept_as_readonly = False 
                # Derive customer name based on submitted (but potentially invalid) dept or org
                submitted_dept_id = form.department_id.data if form.department_id.data and form.department_id.data != 0 else None
                if submitted_dept_id:
                    dept_obj_on_err = db.session.get(Department, submitted_dept_id)
                    if dept_obj_on_err and dept_obj_on_err.organization_id == user_org.id: # Check if submitted dept is valid for this org client
                         derived_customer_name_for_template = dept_obj_on_err.name
                    else: # Submitted dept was invalid or not for their org
                         derived_customer_name_for_template = user_org.name # Fallback to org name
                else: # No department was selected by org client
                    derived_customer_name_for_template = user_org.name
            else: # Admin/Agent
                render_org_as_readonly = False
                render_dept_as_readonly = False
                # Re-derive customer name for template based on submitted org/dept data
                submitted_dept_id_on_err = form.department_id.data if form.department_id.data and form.department_id.data !=0 else None
                submitted_org_id_on_err = form.organization_id.data if form.organization_id.data and form.organization_id.data !=0 else None
                if submitted_dept_id_on_err:
                    dept_obj_err = db.session.get(Department, submitted_dept_id_on_err)
                    if dept_obj_err : derived_customer_name_for_template = dept_obj_err.name
                elif submitted_org_id_on_err:
                    org_obj_err = db.session.get(OrganizationOption, submitted_org_id_on_err)
                    if org_obj_err: derived_customer_name_for_template = org_obj_err.name

            return render_template('client/create_ticket.html', title='Submit New Support Request', form=form,
                                   render_org_as_readonly=render_org_as_readonly,
                                   render_dept_as_readonly=render_dept_as_readonly,
                                   template_org_name=template_org_name,
                                   template_dept_name=template_dept_name,
                                   derived_customer_name_for_template=derived_customer_name_for_template)

        # If all validations passed
        new_ticket = Ticket(
            # ... (ticket creation logic as before - unchanged) ...
            title=form.title.data,
            description=form.description.data,
            created_by_id=current_user.id,
            organization_id=ticket_org_id,
            department_id=ticket_dept_id,
            customer_name=final_customer_name_for_ticket,
            category_id=form.category.data,
            severity=form.severity.data,
            cloud_provider=form.cloud_provider.data or None,
            aws_service=form.aws_service.data if form.cloud_provider.data == 'AWS' and form.aws_service.data else None,
            aws_account_id=form.aws_account_id.data.strip() if form.aws_account_id.data else None,
            environment=form.environment.data or None,
            form_type_id=form.form_type_id.data if form.form_type_id.data and form.form_type_id.data != 0 else None,
            tags=form.tags.data.strip() if form.tags.data else None,
            additional_email_recipients=form.additional_recipients.data.strip() if form.additional_recipients.data else None,
            request_call_back=form.request_call_back.data or None,
            contact_details=form.contact_details.data.strip() if form.contact_details.data else None,
        )
        db.session.add(new_ticket)
        try:
            # ... (commit logic, flash, redirect - unchanged) ...
            db.session.flush() 
            for file_info in uploaded_files_info: 
                attachment = Attachment(filename=file_info['original_filename'], 
                                        stored_filename=file_info['stored_filename'], 
                                        ticket_id=new_ticket.id, 
                                        uploaded_by_id=current_user.id,
                                        content_type=file_info['content_type'])
                db.session.add(attachment)
            log_interaction(new_ticket.id, 'TICKET_CREATED', user_id=current_user.id, details={'title': new_ticket.title}, timestamp_override=new_ticket.created_at)
            db.session.commit()
            flash('Ticket created successfully!', 'success')
            return redirect(url_for('view_ticket', ticket_id=new_ticket.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Database error during ticket creation: {str(e)[:150]}', 'danger')
            current_app.logger.error(f"Ticket creation error (commit stage): {e}", exc_info=True)
            # Re-render form, re-populating choices & template vars (similar to form.errors block)
            if current_user.role == 'client' and user_dept and user_org:
                render_org_as_readonly = True; template_org_name = user_org.name
                render_dept_as_readonly = True; template_dept_name = user_dept.name
                derived_customer_name_for_template = user_dept.name
            elif current_user.role == 'organization_client' and user_org:
                render_org_as_readonly = True; template_org_name = user_org.name
                render_dept_as_readonly = False 
                # Similar logic for derived_customer_name_for_template as in form.errors block
                submitted_dept_id_commit_err = form.department_id.data if form.department_id.data and form.department_id.data != 0 else None
                if submitted_dept_id_commit_err:
                    dept_obj_on_commit_err = db.session.get(Department, submitted_dept_id_commit_err)
                    if dept_obj_on_commit_err and dept_obj_on_commit_err.organization_id == user_org.id:
                         derived_customer_name_for_template = dept_obj_on_commit_err.name
                    else: derived_customer_name_for_template = user_org.name
                else: derived_customer_name_for_template = user_org.name
            else: # Admin/Agent
                render_org_as_readonly = False
                render_dept_as_readonly = False
                temp_dept_id_on_err_commit = form.department_id.data if form.department_id.data and form.department_id.data !=0 else None
                temp_org_id_on_err_commit = form.organization_id.data if form.organization_id.data and form.organization_id.data !=0 else None
                if temp_dept_id_on_err_commit:
                    dept_obj_err_commit = db.session.get(Department, temp_dept_id_on_err_commit)
                    if dept_obj_err_commit : derived_customer_name_for_template = dept_obj_err_commit.name
                elif temp_org_id_on_err_commit:
                    org_obj_err_commit = db.session.get(OrganizationOption, temp_org_id_on_err_commit)
                    if org_obj_err_commit: derived_customer_name_for_template = org_obj_err_commit.name
            
            return render_template('client/create_ticket.html', title='Submit New Support Request', form=form,
                                   render_org_as_readonly=render_org_as_readonly,
                                   render_dept_as_readonly=render_dept_as_readonly,
                                   template_org_name=template_org_name,
                                   template_dept_name=template_dept_name,
                                   derived_customer_name_for_template=derived_customer_name_for_template)

    # Final render for GET request or if POST failed very early (before form.validate_on_submit())
    # Ensure render flags and derived_customer_name are set for the template.
    # The GET logic above should have already set these.
    if request.method == 'GET' and not derived_customer_name_for_template: # Fallback for admin/agent GET
        if not (render_org_as_readonly or render_dept_as_readonly): # Admin/Agent view
             derived_customer_name_for_template = "Select Org/Company to associate ticket"

    return render_template('client/create_ticket.html', title='Submit New Support Request', form=form,
                           render_org_as_readonly=render_org_as_readonly,
                           render_dept_as_readonly=render_dept_as_readonly,
                           template_org_name=template_org_name,
                           template_dept_name=template_dept_name,
                           derived_customer_name_for_template=derived_customer_name_for_template)
    

# In app.py
# In app.py

@app.route('/tickets/overview') # This is for Company/Organization-wide tickets for clients
@login_required
def tickets_overview():
    query = Ticket.query
    page_title = "Tickets Overview"
    department_filter_choices = [] 
    current_department_filter_val = request.args.get('department_filter', '0', type=str)


    if current_user.role == 'client' and current_user.department_id:
        # Client sees tickets for their specific department
        user_dept = db.session.get(Department, current_user.department_id)
        if user_dept:
            query = query.filter(Ticket.department_id == current_user.department_id)
            page_title = f"Tickets for {user_dept.name}"
        else:
            flash("Your department association is missing or invalid. Please contact support.", "warning")
            return redirect(url_for('dashboard'))

    elif current_user.role == 'organization_client' and current_user.organization_id:
        # Organization Client sees tickets for their entire organization
        user_org = db.session.get(OrganizationOption, current_user.organization_id)
        if user_org:
            query = query.filter(Ticket.organization_id == current_user.organization_id)
            page_title = f"Tickets for {user_org.name}"
            
            org_departments = Department.query.filter_by(organization_id=current_user.organization_id, is_active=True).order_by(Department.name).all()
            department_filter_choices = [(0, "All Companies / Departments")] + [(d.id, d.name) for d in org_departments]
            
            if current_department_filter_val != '0' and current_department_filter_val.isdigit():
                query = query.filter(Ticket.department_id == int(current_department_filter_val))
        else:
            flash("Your organization association is missing or invalid. Please contact support.", "warning")
            return redirect(url_for('dashboard'))

    elif current_user.is_agent or current_user.is_admin:
        # Agents/Admins use their own dedicated views
        flash("Agents and Admins should use their dedicated ticket list views.", "info")
        if current_user.is_admin:
            return redirect(url_for('admin_all_tickets'))
        else: # Agent
            return redirect(url_for('agent_ticket_list')) # Default agent view
    else: 
        flash("You do not have permission to view this page or are not fully configured.", "warning")
        return redirect(url_for('dashboard'))

    page = request.args.get('page', 1, type=int)
    tickets_pagination = query.order_by(Ticket.updated_at.desc()).paginate(page=page, per_page=10, error_out=False)
    
    # Use client/tickets_overview.html template for this view
    return render_template('client/tickets_overview.html', 
                           title=page_title, 
                           tickets_pagination=tickets_pagination,
                           department_filter_choices=department_filter_choices,
                           current_department_filter=current_department_filter_val) 

    
# In app.py


# In app.py (near other _admin_list_options routes)

# --- Department Management (NEW if not already existing) ---
@app.route('/admin/departments')
@admin_required
def admin_department_list():
    # Explicitly set model_name_slug for the template context
    return _admin_list_options(Department, 'admin/list_options.html', 'Manage Departments', order_by_attr='name') # model_name_slug will be 'department'

# You would also need create/edit/delete routes for Departments
# Example for create/edit:
# class DepartmentForm(FlaskForm): # Define this in your Forms section
#     name = StringField('Department Name', validators=[DataRequired(), Length(max=100)])
#     organization_id = SelectField('Parent Organization', coerce=int, validators=[DataRequired(message="Organization is required.")])
#     is_active = BooleanField('Active', default=True)
#     description = TextAreaField('Description (Optional)', validators=[Optional(), Length(max=255)]) # If you add description
#     submit = SubmitField('Save Department')

#     def __init__(self, *args, **kwargs):
#         super(DepartmentForm, self).__init__(*args, **kwargs)
#         self.organization_id.choices = [(0, '-- Select Organization --')] + \
#                                        [(org.id, org.name) for org in OrganizationOption.query.filter_by(is_active=True).order_by('name').all()]

# @app.route('/admin/department/new', methods=['GET', 'POST'])
# @app.route('/admin/department/<int:item_id>/edit', methods=['GET', 'POST'])
# @admin_required
# def admin_create_edit_department(item_id=None):
#     # Custom logic might be needed here if _admin_create_edit_option is too generic
#     # For now, assuming _admin_create_edit_option could be adapted or a specific function written.
#     # This is a placeholder to show where it would go.
#     return _admin_create_edit_option(Department, DepartmentForm, 'admin_department_list', item_id)


# @app.route('/admin/department/<int:item_id>/delete', methods=['POST'])
# @admin_required
# def admin_delete_department(item_id):
#     # Custom logic for deleting departments, e.g., checking if users or tickets are assigned
#     return _admin_delete_option(Department, item_id, 'admin_department_list', related_ticket_attr_id='department_id')







@app.route('/tickets/my')
@login_required
def my_tickets(): 
    page = request.args.get('page', 1, type=int)
    # Query for tickets created by the current user
    tickets_query = Ticket.query.filter_by(created_by_id=current_user.id)\
                                .order_by(Ticket.updated_at.desc())
    
    tickets_pagination = tickets_query.paginate(page=page, per_page=10, error_out=False)
    
    # The 'current_department_filter' is not strictly needed for this page 
    # as it's not filtering by department, but a pagination macro might expect it.
    # We can pass '0' or None.
    return render_template('client/my_tickets.html', 
                           title='My Submitted Tickets', 
                           tickets_pagination=tickets_pagination,
                           current_department_filter='0' # Or None, depending on your pagination macro needs
                           )

# app.py

# ... (other imports and code) ...
from datetime import datetime # Ensure datetime is imported

# ... (Models, Forms, etc.) ...

# In app.py

# Ensure all necessary imports are at the top of your app.py:
# from .models import User, OrganizationOption, Department, Ticket, Category, Attachment, Interaction, SeverityOption # etc.
# from .forms import CommentForm, AgentUpdateTicketForm 
# from .helpers import (get_active_category_choices, get_active_cloud_provider_choices, 
#                       get_active_severity_choices, get_active_environment_choices, 
#                       get_active_organization_choices, get_active_form_type_choices, 
#                       get_active_apn_opportunity_choices, get_active_support_modal_choices,
#                       log_interaction, trigger_priority_call_alert, EFFORT_CHOICES) # EFFORT_CHOICES for context
# from flask import render_template, redirect, url_for, flash, request, current_app
# from flask_login import login_required, current_user
# from datetime import datetime, date
# from itertools import groupby

@app.route('/ticket/<int:ticket_id>', methods=['GET', 'POST'])
@login_required
def view_ticket(ticket_id):
    ticket = db.session.get(Ticket, ticket_id) 
    if not ticket:
        flash('Ticket not found.', 'danger')
        return redirect(url_for('dashboard'))
        
    # --- Permission Check ---
    can_view = False
    if current_user.is_admin or current_user.is_agent:
        can_view = True
    elif ticket.created_by_id == current_user.id: 
        can_view = True
    elif current_user.role == 'client' and current_user.department_id and \
         ticket.department_id == current_user.department_id:
        can_view = True
    elif current_user.role == 'organization_client' and current_user.organization_id and \
         ticket.organization_id == current_user.organization_id:
        can_view = True
    
    if not can_view:
        flash('You do not have permission to view this ticket.', 'danger')
        current_app.logger.warning(
            f"Unauthorized attempt to view ticket ID: {ticket_id} by user {current_user.username} "
            f"(Role: {current_user.role}, OrgID: {current_user.organization_id}, DeptID: {current_user.department_id}). "
            f"Ticket OrgID: {ticket.organization_id}, Ticket DeptID: {ticket.department_id}"
        )
        return redirect(url_for('dashboard')) 

    comment_form = CommentForm()
    agent_update_form = None 
    attachments = ticket.ticket_attachments.order_by(Attachment.uploaded_at.desc()).all()
    is_privileged_user = current_user.is_agent or current_user.is_admin

    sorted_interaction_dates = []
    interactions_by_date = {}
    today_date_obj = date.today()
    yesterday_date_obj = today_date_obj - timedelta(days=1)

    if is_privileged_user:
        # Initialize form. If GET, populate with ticket object. If POST, WTForms handles submitted data.
        agent_update_form = AgentUpdateTicketForm(obj=ticket if request.method == 'GET' else None, prefix="agent_update") # Added prefix
        
        # Populate choices for SelectFields
        agent_choices = [(u.id, u.username) for u in User.query.filter(User.role.in_(['agent', 'admin'])).order_by(User.username).all()]
        agent_update_form.assigned_to_id.choices = [(0, '--- Unassign/Select Agent ---')] + agent_choices
        agent_update_form.category_id.choices = get_active_category_choices() 
        agent_update_form.cloud_provider.choices = get_active_cloud_provider_choices()
        agent_update_form.severity.choices = get_active_severity_choices()
        agent_update_form.environment.choices = get_active_environment_choices()
        agent_update_form.organization_id.choices = get_active_organization_choices()
        agent_update_form.form_type_id.choices = get_active_form_type_choices()
        agent_update_form.apn_opportunity_id.choices = get_active_apn_opportunity_choices()
        agent_update_form.support_modal_id.choices = get_active_support_modal_choices()
        # EFFORT_CHOICES is already set in the form definition

        if request.method == 'GET': # Pre-fill form fields on GET
            agent_update_form.status.data = ticket.status
            agent_update_form.priority.data = ticket.priority
            agent_update_form.assigned_to_id.data = ticket.assigned_to_id or 0
            agent_update_form.category_id.data = ticket.category_id or 0 
            agent_update_form.cloud_provider.data = ticket.cloud_provider or ''
            agent_update_form.severity.data = ticket.severity or ''
            agent_update_form.aws_service.data = ticket.aws_service or ''
            agent_update_form.aws_account_id.data = ticket.aws_account_id or ''
            agent_update_form.environment.data = ticket.environment or ''
            agent_update_form.organization_id.data = ticket.organization_id or 0
            agent_update_form.form_type_id.data = ticket.form_type_id or 0
            agent_update_form.tags.data = ticket.tags or ''
            agent_update_form.additional_email_recipients.data = ticket.additional_email_recipients or ''
            agent_update_form.request_call_back.data = ticket.request_call_back or ''
            agent_update_form.contact_details.data = ticket.contact_details or ''
            agent_update_form.aws_support_case_id.data = ticket.aws_support_case_id or ''
            # For SelectField with custom coerce, ensure data is string if it comes from int
            agent_update_form.effort_required_to_resolve_min.data = str(ticket.effort_required_to_resolve_min) if ticket.effort_required_to_resolve_min is not None else ''
            agent_update_form.apn_opportunity_id.data = ticket.apn_opportunity_id or 0
            agent_update_form.apn_opportunity_description.data = ticket.apn_opportunity_description or ''
            agent_update_form.support_modal_id.data = ticket.support_modal_id or 0

    if request.method == 'POST':
        if 'submit_comment' in request.form and comment_form.validate_on_submit():
            is_internal_comment = is_privileged_user and hasattr(comment_form, 'is_internal') and comment_form.is_internal.data
            comment = Comment(content=comment_form.content.data, user_id=current_user.id, ticket_id=ticket.id, is_internal=is_internal_comment)
            db.session.add(comment)
            db.session.flush() # To get comment.id for logging
            log_interaction(ticket.id, 'COMMENT_ADDED', user_id=current_user.id, details={'comment_id': comment.id, 'is_internal': is_internal_comment})
            
            if not is_internal_comment and is_privileged_user and not ticket.first_response_at:
                ticket.first_response_at = comment.created_at
                if ticket.created_at: 
                    ticket.first_response_duration_minutes = int((ticket.first_response_at - ticket.created_at).total_seconds() / 60)
                log_interaction(ticket.id, 'FIRST_RESPONSE_RECORDED', user_id=current_user.id, 
                                details={'responded_at': ticket.first_response_at.isoformat() if ticket.first_response_at else None, 
                                         'duration_minutes': ticket.first_response_duration_minutes})
            db.session.commit()
            flash('Your comment has been added.', 'success')
            return redirect(url_for('view_ticket', ticket_id=ticket.id, _anchor='comments_section'))

        elif 'submit_update' in request.form and is_privileged_user and agent_update_form:
            # Ensure choices are populated before validation if it's a POST request
            # This is important if validation fails and the form is re-rendered
            if not agent_update_form.assigned_to_id.choices: # Check if choices were already populated
                 agent_choices = [(u.id, u.username) for u in User.query.filter(User.role.in_(['agent', 'admin'])).order_by(User.username).all()]
                 agent_update_form.assigned_to_id.choices = [(0, '--- Unassign/Select Agent ---')] + agent_choices
            if not agent_update_form.category_id.choices: agent_update_form.category_id.choices = get_active_category_choices()
            if not agent_update_form.cloud_provider.choices: agent_update_form.cloud_provider.choices = get_active_cloud_provider_choices()
            if not agent_update_form.severity.choices: agent_update_form.severity.choices = get_active_severity_choices()
            if not agent_update_form.environment.choices: agent_update_form.environment.choices = get_active_environment_choices()
            if not agent_update_form.organization_id.choices: agent_update_form.organization_id.choices = get_active_organization_choices()
            if not agent_update_form.form_type_id.choices: agent_update_form.form_type_id.choices = get_active_form_type_choices()
            if not agent_update_form.apn_opportunity_id.choices: agent_update_form.apn_opportunity_id.choices = get_active_apn_opportunity_choices()
            if not agent_update_form.support_modal_id.choices: agent_update_form.support_modal_id.choices = get_active_support_modal_choices()

            if agent_update_form.validate_on_submit():
                old_values = {
                    'status': ticket.status, 'priority': ticket.priority,
                    'assignee_name': ticket.assignee.username if ticket.assignee else "Unassigned",
                    'category_name': ticket.category_ref.name if ticket.category_ref else "None",
                    'organization_name': ticket.organization_option_ref.name if ticket.organization_option_ref else "None",
                    'department_name': ticket.department_ref.name if ticket.department_ref else "None", # Will log if it changes due to org change
                    'customer_name': ticket.customer_name or "None",
                    'severity': ticket.severity or "None",
                    'cloud_provider': ticket.cloud_provider or "None",
                    'aws_service': ticket.aws_service or "None",
                    'aws_account_id': ticket.aws_account_id or "None",
                    'environment': ticket.environment or "None",
                    'form_type_name': ticket.form_type_option_ref.name if ticket.form_type_option_ref else "None",
                    'tags': ticket.tags or "None",
                    'additional_email_recipients': ticket.additional_email_recipients or "None",
                    'request_call_back': ticket.request_call_back or "None",
                    'contact_details': ticket.contact_details or "None",
                    'aws_support_case_id': ticket.aws_support_case_id or "None",
                    'effort_required_to_resolve_min': str(ticket.effort_required_to_resolve_min) if ticket.effort_required_to_resolve_min is not None else "None",
                    'apn_opportunity_name': ticket.apn_opportunity_option_ref.name if ticket.apn_opportunity_option_ref else "None",
                    'apn_opportunity_description': ticket.apn_opportunity_description or "None",
                    'support_modal_name': ticket.support_modal_option_ref.name if ticket.support_modal_option_ref else "None",
                    'total_resolution_duration_minutes': ticket.total_resolution_duration_minutes,
                }
                old_severity_for_alert_trigger = ticket.severity
                old_ticket_org_id = ticket.organization_id # For checking if org changed

                # Assign form data to ticket object
                ticket.status = agent_update_form.status.data
                ticket.priority = agent_update_form.priority.data
                ticket.tags = agent_update_form.tags.data.strip() if agent_update_form.tags.data else None
                ticket.cloud_provider = agent_update_form.cloud_provider.data if agent_update_form.cloud_provider.data else None
                ticket.aws_service = agent_update_form.aws_service.data if ticket.cloud_provider == 'AWS' and agent_update_form.aws_service.data else None
                ticket.aws_account_id = agent_update_form.aws_account_id.data.strip() if agent_update_form.aws_account_id.data else None
                ticket.environment = agent_update_form.environment.data if agent_update_form.environment.data else None
                ticket.additional_email_recipients = agent_update_form.additional_email_recipients.data.strip() if agent_update_form.additional_email_recipients.data else None
                ticket.request_call_back = agent_update_form.request_call_back.data if agent_update_form.request_call_back.data else None
                ticket.contact_details = agent_update_form.contact_details.data.strip() if agent_update_form.contact_details.data else None
                ticket.aws_support_case_id = agent_update_form.aws_support_case_id.data.strip() if agent_update_form.aws_support_case_id.data else None
                ticket.apn_opportunity_description = agent_update_form.apn_opportunity_description.data.strip() if agent_update_form.apn_opportunity_description.data else None
                
                # Handle fields that might be 0 for "no selection" or require specific handling
                ticket.assigned_to_id = agent_update_form.assigned_to_id.data if agent_update_form.assigned_to_id.data != 0 else None
                ticket.category_id = agent_update_form.category_id.data # Assuming category is required and 0 means a valid placeholder was selected
                ticket.form_type_id = agent_update_form.form_type_id.data if agent_update_form.form_type_id.data != 0 else None
                ticket.apn_opportunity_id = agent_update_form.apn_opportunity_id.data if agent_update_form.apn_opportunity_id.data != 0 else None
                ticket.support_modal_id = agent_update_form.support_modal_id.data if agent_update_form.support_modal_id.data != 0 else None
                
                ticket.effort_required_to_resolve_min = agent_update_form.effort_required_to_resolve_min.data # Already coerced to int or None by the form
                ticket.severity = agent_update_form.severity.data # String

                # Organization and Department logic
                new_org_id = agent_update_form.organization_id.data
                ticket.organization_id = new_org_id if new_org_id != 0 else None

                if ticket.organization_id != old_ticket_org_id:
                    ticket.department_id = None # Clear department if organization changes as it's not re-selected in this form
                    current_app.logger.info(f"Ticket {ticket.id} organization changed by agent from {old_ticket_org_id} to {ticket.organization_id}. Department cleared.")
                
                # Re-derive customer_name based on the (potentially new) org/dept
                if ticket.department_id:
                    dept = db.session.get(Department, ticket.department_id)
                    if dept and dept.organization_id == ticket.organization_id: # Check if dept belongs to current ticket org
                        ticket.customer_name = dept.name
                    else: # Department is invalid for this org, or no department
                        ticket.department_id = None # Ensure it's cleared if invalid
                        if ticket.organization_id:
                           org = db.session.get(OrganizationOption, ticket.organization_id)
                           ticket.customer_name = org.name if org else (ticket.creator.username if ticket.creator else "Undefined Customer")
                        else: # No org either
                           ticket.customer_name = ticket.creator.username if ticket.creator else "Undefined Customer"
                elif ticket.organization_id:
                    org = db.session.get(OrganizationOption, ticket.organization_id)
                    ticket.customer_name = org.name if org else (ticket.creator.username if ticket.creator else "Undefined Customer")
                else: # No organization and no department associated with the ticket
                    ticket.customer_name = ticket.creator.username if ticket.creator else "Undefined Customer"


                # Resolved_at and Total Duration logic
                if ticket.status in ['Resolved', 'Closed'] and old_values['status'] not in ['Resolved', 'Closed']:
                    if not ticket.resolved_at:
                        ticket.resolved_at = datetime.utcnow()
                    # Calculate total resolution time when first moving to a resolved/closed state
                    if ticket.created_at and ticket.resolved_at:
                        delta = ticket.resolved_at - ticket.created_at
                        ticket.total_resolution_duration_minutes = int(delta.total_seconds() / 60)
                elif ticket.status not in ['Resolved', 'Closed'] and old_values['status'] in ['Resolved', 'Closed']:
                    if ticket.resolved_at:
                        ticket.resolved_at = None
                    # Clear total resolution time when reopening
                    ticket.total_resolution_duration_minutes = None
                
                # Interaction Logging
                changed_fields_map = {
                    'Status': (old_values['status'], ticket.status), 
                    'Priority': (old_values['priority'], ticket.priority),
                    'Assignee': (old_values['assignee_name'], ticket.assignee.username if ticket.assignee else "Unassigned"),
                    'Category': (old_values['category_name'], ticket.category_ref.name if ticket.category_ref else "None"),
                    'Organization': (old_values['organization_name'], ticket.organization_option_ref.name if ticket.organization_option_ref else "None"),
                    'Department': (old_values['department_name'], ticket.department_ref.name if ticket.department_ref else "None"), # Will show change if org change cleared it
                    'Customer Name': (old_values['customer_name'], ticket.customer_name or "None"),
                    'Severity': (old_values['severity'], ticket.severity or "None"),
                    'Cloud Provider': (old_values['cloud_provider'], ticket.cloud_provider or "None"), 
                    'AWS Service': (old_values['aws_service'], ticket.aws_service or "None"), 
                    'AWS Account ID': (old_values['aws_account_id'], ticket.aws_account_id or "None"),
                    'Environment': (old_values['environment'], ticket.environment or "None"),
                    'Form Type': (old_values['form_type_name'], ticket.form_type_option_ref.name if ticket.form_type_option_ref else "None"),
                    'Tags': (old_values['tags'], ticket.tags or "None"),
                    'Additional Email Recipients': (old_values['additional_email_recipients'], ticket.additional_email_recipients or "None"),
                    'Request Call Back': (old_values['request_call_back'], ticket.request_call_back or "None"),
                    'Contact Details': (old_values['contact_details'], ticket.contact_details or "None"),
                    'AWS Support Case ID': (old_values['aws_support_case_id'], ticket.aws_support_case_id or "None"),
                    'Effort (min)': (old_values['effort_required_to_resolve_min'], str(ticket.effort_required_to_resolve_min) if ticket.effort_required_to_resolve_min is not None else "None"),
                    'APN Opportunity': (old_values['apn_opportunity_name'], ticket.apn_opportunity_option_ref.name if ticket.apn_opportunity_option_ref else "None"),
                    'APN Opportunity Description': (old_values['apn_opportunity_description'], ticket.apn_opportunity_description or "None"),
                    'Support Modal': (old_values['support_modal_name'], ticket.support_modal_option_ref.name if ticket.support_modal_option_ref else "None"),
                    'Total Resolution (min)': (old_values['total_resolution_duration_minutes'], ticket.total_resolution_duration_minutes)
                }
                for field_name, (old_val, new_val) in changed_fields_map.items():
                    if str(old_val) != str(new_val): 
                        interaction_type_suffix = field_name.upper().replace(" ", "_").replace("(", "").replace(")", "") + "_CHANGE"
                        log_interaction(ticket.id, interaction_type_suffix, user_id=current_user.id, details={'old_value': str(old_val), 'new_value': str(new_val), 'field_display_name': field_name})
                
                try: 
                    db.session.commit()
                    flash('Ticket details updated successfully.', 'success')
                    if ticket.severity != old_severity_for_alert_trigger or \
                       (ticket.severity in app.config.get("SEVERITIES_FOR_CALL_ALERT", []) and old_severity_for_alert_trigger not in app.config.get("SEVERITIES_FOR_CALL_ALERT", [])):
                        trigger_priority_call_alert(ticket, old_severity_for_alert_trigger) 
                    return redirect(url_for('view_ticket', ticket_id=ticket.id))
                except Exception as e: 
                    db.session.rollback()
                    flash(f'Database error during ticket update: {str(e)[:150]}', 'danger')
                    current_app.logger.error(f"Ticket update DB error for #{ticket.id}: {e}", exc_info=True)
            else: # agent_update_form.validate_on_submit() failed
                 flash('Error updating ticket. Please check the form values.', 'danger')
                 # Choices were already re-populated before validation attempt
        
        elif request.method == 'POST' and (comment_form.errors or (agent_update_form and agent_update_form.errors)):
            flash('Please correct the errors in the form.', 'danger') # General error if a form was submitted but didn't hit specific handlers or failed validation early

    # --- GET Request or POST with errors: Prepare data for template (interaction log, comments, etc.) ---
    comments_query = ticket.comments
    if not is_privileged_user: 
        comments_query = comments_query.filter_by(is_internal=False)
    comments = comments_query.order_by(Comment.created_at.asc()).all()
    
    if is_privileged_user: 
        raw_interactions = ticket.interactions_rel.order_by(Interaction.timestamp.desc()).all()
        processed_interactions = []
        for interaction in raw_interactions:
            actor_name = "System"
            if interaction.user: actor_name = interaction.user.username
            elif interaction.interaction_type.startswith('TICKET_CREATED') and ticket.creator: actor_name = ticket.creator.username
            
            p_interaction = {
                'obj': interaction, 'actor_name': actor_name, 
                'timestamp_obj': interaction.timestamp, 
                'time_str': interaction.timestamp.strftime('%H:%M'), 
                'date_str_short': interaction.timestamp.strftime('%b %d'), 
                'datetime_str_full': interaction.timestamp.strftime('%b %d, %Y %H:%M:%S'), 
                'message': "", 'title_for_display': actor_name
            }
            details = interaction.details or {}
            field_display_name = details.get('field_display_name')
            
            if field_display_name: 
                old_val_display = details.get('old_value', "not set")
                new_val_display = details.get('new_value', "not set")
                if old_val_display == "None" and new_val_display == "None": p_interaction['message'] = f"verified <strong>{field_display_name}</strong> (remained not set)."
                elif old_val_display == "None" or old_val_display == "not set" or old_val_display == "" or old_val_display is None: p_interaction['message'] = f"set <strong>{field_display_name}</strong> to <strong>{new_val_display}</strong>."
                elif new_val_display == "None" or new_val_display == "not set" or new_val_display == "" or new_val_display is None: p_interaction['message'] = f"cleared <strong>{field_display_name}</strong> (was <strong>{old_val_display}</strong>)."
                else: p_interaction['message'] = f"changed <strong>{field_display_name}</strong> from <strong>{old_val_display}</strong> to <strong>{new_val_display}</strong>."
            
            # ... (rest of interaction message formatting, same as before) ...
            if interaction.interaction_type == 'TICKET_CREATED': p_interaction['message'] = f"created this ticket."
            elif interaction.interaction_type == 'EMAIL_TICKET_CREATED': p_interaction['message'] = f"created this ticket via email (Subject: {details.get('subject','N/A')}, From: {details.get('sender','N/A')})."
            elif interaction.interaction_type == 'TICKET_CREATED_GDOC': p_interaction['message'] = f"created this ticket from Google Doc (Title: {details.get('title','N/A')})."
            elif interaction.interaction_type == 'COMMENT_ADDED':
                comment_id_from_details = details.get('comment_id')
                comment_obj = db.session.get(Comment, comment_id_from_details) if comment_id_from_details else None
                comment_type = "internal" if details.get('is_internal') else "public"
                p_interaction['message'] = f"added a {comment_type} comment."
                if comment_obj and (not comment_obj.is_internal or is_privileged_user): 
                    p_interaction['comment_preview'] = comment_obj.content
            elif interaction.interaction_type == 'FIRST_RESPONSE_RECORDED': 
                duration_min = details.get('duration_minutes', 'N/A')
                p_interaction['message'] = f"logged the first agent response. Duration: {duration_min} minutes."
            elif interaction.interaction_type == 'TICKET_RESOLVED_TIMESTAMPED':
                resolved_at_str = details.get('resolved_at')
                status_changed_to_str = details.get('status_changed_to')
                p_interaction['message'] = f"marked ticket as <strong>{status_changed_to_str}</strong> and recorded resolution time."
                if resolved_at_str:
                    try: p_interaction['message'] += f" (at {datetime.fromisoformat(resolved_at_str).strftime('%b %d, %Y %H:%M')})"
                    except: pass 
            elif interaction.interaction_type == 'TICKET_REOPENED_RESOLUTION_CLEARED':
                status_changed_to_str = details.get('status_changed_to')
                p_interaction['message'] = f"reopened ticket (status to <strong>{status_changed_to_str}</strong>), cleared previous resolution time."
            
            if not p_interaction['message'] and not field_display_name: 
                p_interaction['message'] = f"performed action: {interaction.interaction_type.replace('_', ' ').title()}. Details: {str(details)[:100]}"


            processed_interactions.append(p_interaction)
        
        interactions_by_date = {
            k: sorted(list(g), key=lambda i: i['timestamp_obj'], reverse=True) 
            for k, g in groupby(processed_interactions, key=lambda i: i['timestamp_obj'].date())
        }
        sorted_interaction_dates = sorted(interactions_by_date.keys(), reverse=True)
    
    page_title_val = f'Ticket #{ticket.id}: {ticket.title}'
    return render_template('client/view_ticket.html', 
                           title=page_title_val, 
                           ticket=ticket, 
                           comments=comments, 
                           comment_form=comment_form, 
                           agent_update_form=agent_update_form, 
                           attachments=attachments, 
                           is_privileged_user=is_privileged_user,
                           sorted_interaction_dates=sorted_interaction_dates, 
                           interactions_by_date=interactions_by_date,     
                           today_date=today_date_obj,                   
                           yesterday_date=yesterday_date_obj,
                           EFFORT_CHOICES=EFFORT_CHOICES)
#kanban
# ... (near other agent_required routes) ...

# app.py

# ... (other imports and code) ...

# In app.py, find and REPLACE the entire agent_kanban_board function with this new version.

@app.route('/agent/kanban_board')
@agent_required
def agent_kanban_board():

    filters = {k: v for k, v in request.args.items() if v and v.lower() != 'all'}
    
    # Fetch data for filter dropdowns
    categories_for_filter = Category.query.order_by('name').all()
    agents_for_filter = User.query.filter(User.role.in_(['agent', 'admin'])).order_by('username').all()
    organizations_for_filter = OrganizationOption.query.filter_by(is_active=True).order_by('name').all()
    
    # For pre-populating department filter if an org is already selected on page load
    departments_for_filter = []
    organization_id_filter = filters.get('organization_id')
    if organization_id_filter and organization_id_filter.isdigit() and organization_id_filter != '0':
        depts_query = Department.query.filter_by(
            organization_id=int(organization_id_filter), 
            is_active=True
        ).order_by(Department.name).all()
        departments_for_filter = [{'id': dept.id, 'name': dept.name} for dept in depts_query]

    # --- Ticket Query and Filtering ---
    tickets_query = Ticket.query

    # Apply common filters from the form using the 'filters' dictionary
    tickets_query = _apply_common_filters(tickets_query, filters)

    if 'assigned_to_id' not in filters:
        if current_user.is_agent and not current_user.is_admin:
            tickets_query = tickets_query.filter(Ticket.assigned_to_id == current_user.id)
    
    # Sort tickets for display within columns
    tickets_query = tickets_query.order_by(
        db.case(
            PRIORITY_ORDER_MAP, 
            value=Ticket.priority, 
            else_=len(PRIORITY_ORDER_MAP) 
        ).asc(), 
        Ticket.updated_at.desc()
    )
    
    all_tickets_for_board = tickets_query.all()

    # --- Data Structuring for Kanban ---
    kanban_statuses_ordered = [s[0] for s in TICKET_STATUS_CHOICES]
    tickets_by_status = {status: [] for status in kanban_statuses_ordered}
    for ticket in all_tickets_for_board:
        if ticket.status in tickets_by_status:
            tickets_by_status[ticket.status].append(ticket)

    page_title = "Kanban Board"

    return render_template(
        'agent/kanban_board.html', 
        title=page_title,
        tickets_by_status=tickets_by_status,
        kanban_statuses=kanban_statuses_ordered,
        TICKET_PRIORITY_CHOICES_DICT=dict(TICKET_PRIORITY_CHOICES),
        # Pass all filter-related data to the template
        priorities=TICKET_PRIORITY_CHOICES,
        categories=categories_for_filter,
        agents=agents_for_filter,
        organizations_for_filter=organizations_for_filter,
        departments_for_filter=departments_for_filter,
        # FIX: Ensure the 'filters' dictionary is always passed to the template.
        current_filters=filters
    )

# ... (rest of your app.py) ...



@app.route('/api/ai/suggest_comment/<int:ticket_id>', methods=['POST'])
@agent_required
def ai_suggest_comment(ticket_id):
    if not app.config['GEMINI_API_KEY']:
        return jsonify({"error": "AI service unavailable. Key not configured."}), 503

    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        return jsonify({"error": "Ticket not found."}), 404

    # Assemble the prompt context
    comments_query = ticket.comments.filter_by(is_internal=False).order_by(Comment.created_at.asc())
    comments = comments_query.all()

    conversation_history = []
    for comment in comments:
        author_name = comment.author.username if comment.author else "Unknown"
        conversation_history.append(f"{author_name}: {comment.content}")

    formatted_history = "\n\n".join(conversation_history)

    prompt = f"""
You are a highly-skilled and empathetic customer support agent AI. Your task is to draft a reply for a support agent to send to a customer.

Analyze the entire ticket history provided below to understand the context. The history includes the original ticket description and the conversation so far. The last comment is often from the customer, so pay close attention to it.

Your generated response should be:
- Professional and courteous.
- Empathetic to the user's issue.
- Directly addressing the last question or comment from the customer, if applicable.
- Clear and concise.
- Proposing a next step, asking for clarifying information, or providing a solution if evident from the context.

Do NOT add a generic greeting if a conversation is already underway. Do NOT sign off with a name. Output only the body of the reply.

--- TICKET HISTORY ---

Ticket Title: {ticket.title}

Original Description by {ticket.creator.username if ticket.creator else 'User'}:
{ticket.description}

--- CONVERSATION ---
{formatted_history if formatted_history else "No comments have been made yet."}

--- DRAFT YOUR REPLY BELOW ---
"""

    app.logger.info(f"AI Comment Suggestion Prompt for ticket {ticket_id} (first 150 chars): {prompt[:150]}...")
    
    try:
        model = genai.GenerativeModel('gemini-1.5-flash-latest') 
        response = model.generate_content(
            prompt,
            generation_config=genai.types.GenerationConfig(
                max_output_tokens=800,
                temperature=0.6 # Slightly more creative for drafting replies
            )
        )
        generated_text = "".join(part.text for part in response.candidates[0].content.parts).strip() if response.candidates and response.candidates[0].content.parts else ""
        
        if not generated_text:
            if response.prompt_feedback and response.prompt_feedback.block_reason:
                block_msg = response.prompt_feedback.block_reason_message or "Content generation was limited."
                app.logger.warning(f"Gemini blocked AI comment suggestion for ticket {ticket_id}: {block_msg}")
                return jsonify({"error": f"AI content generation was blocked: {block_msg}. Please rephrase or try again."}), 400
            app.logger.warning(f"Gemini returned empty content for AI comment suggestion on ticket {ticket_id}.")
            return jsonify({"error": "AI could not generate a suggestion for this ticket."})
            
        return jsonify({"suggested_comment": generated_text})

    except Exception as e:
        app.logger.error(f"Gemini API error during AI comment suggestion for ticket {ticket_id}: {str(e)}", exc_info=True)
        return jsonify({"error": "An error occurred while communicating with the AI service."}), 500





# API Endpoint for status updates (Phase 2 will use this)
@app.route('/api/ticket/<int:ticket_id>/update_status_kanban', methods=['POST'])
@agent_required
def update_ticket_status_kanban(ticket_id):
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        return jsonify({'success': False, 'message': 'Ticket not found'}), 404

    data = request.get_json()
    new_status = data.get('new_status')
    # new_order = data.get('new_order') # For ordering within column, future use

    if not new_status or new_status not in [s[0] for s in TICKET_STATUS_CHOICES]:
        return jsonify({'success': False, 'message': 'Invalid status provided'}), 400

    old_status = ticket.status
    ticket.status = new_status
    
    # Update resolved_at logic (same as in view_ticket)
    if new_status in ['Resolved', 'Closed'] and old_status not in ['Resolved', 'Closed']:
        if not ticket.resolved_at:
            ticket.resolved_at = datetime.utcnow()
    elif new_status not in ['Resolved', 'Closed'] and old_status in ['Resolved', 'Closed']:
        if ticket.resolved_at:
            ticket.resolved_at = None
    
    # ticket.kanban_order = new_order # Future use

    log_interaction(ticket.id, 'STATUS_CHANGE_KANBAN', user_id=current_user.id,
                    details={'old_value': old_status, 'new_value': new_status, 'field_display_name': 'Status (Kanban)'})
    
    try:
        db.session.commit()
        return jsonify({'success': True, 'message': f'Ticket #{ticket.id} status updated to {new_status}'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Kanban API Error updating ticket {ticket_id}: {e}")
        return jsonify({'success': False, 'message': 'Error updating ticket status'}), 500
    
    





@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    attachment = Attachment.query.filter_by(stored_filename=filename).first_or_404(); ticket = attachment.ticket
    if not (current_user.is_admin or current_user.is_agent or current_user.id == attachment.uploaded_by_id or current_user.id == ticket.created_by_id or (ticket.assigned_to_id and current_user.id == ticket.assigned_to_id)): flash("You do not have permission to download this file.", "danger"); return redirect(request.referrer or url_for('dashboard'))
    try:
        upload_dir = app.config['UPLOAD_FOLDER']
        if not os.path.isabs(upload_dir): upload_dir = os.path.join(current_app.root_path, upload_dir)
        return send_from_directory(upload_dir, attachment.stored_filename, as_attachment=True, download_name=attachment.filename)
    except FileNotFoundError: app.logger.error(f"Physical file not found: {attachment.stored_filename} in directory {upload_dir}"); flash("File not found on server. Please contact support.", "danger"); return redirect(request.referrer or url_for('dashboard'))

@app.route('/agent/tickets/')
@app.route('/agent/tickets/view/<view_name>')
@agent_required
def agent_ticket_list(view_name=None):
    page = request.args.get('page', 1, type=int); query = Ticket.query; list_title = "Agent Tickets"
    if view_name is None: view_name = 'my_unsolved'
    if view_name == 'my_unsolved': query = query.filter(Ticket.assigned_to_id == current_user.id, Ticket.status.notin_(['Resolved', 'Closed'])); list_title = "Your Unsolved Tickets"
    elif view_name == 'unassigned': query = query.filter(Ticket.assigned_to_id.is_(None), Ticket.status == 'Open'); list_title = "Unassigned Open Tickets"
    elif view_name == 'all_unsolved': query = query.filter(Ticket.status.notin_(['Resolved', 'Closed'])); list_title = "All Unsolved Tickets"
    elif view_name == 'recently_updated': list_title = "Recently Updated Tickets" 
    elif view_name == 'pending': query = query.filter(Ticket.status == 'On Hold'); list_title = "Pending (On Hold) Tickets"
    elif view_name == 'recently_solved': query = query.filter(Ticket.status == 'Resolved'); list_title = "Recently Solved Tickets"
    elif view_name == 'current_tasks': query = query.filter(Ticket.assigned_to_id == current_user.id, Ticket.status == 'In Progress'); list_title = "Your Current In-Progress Tickets"
    else: flash(f"Unknown ticket view: '{view_name}'. Defaulting to 'Your Unsolved Tickets'.", "warning"); query = query.filter(Ticket.assigned_to_id == current_user.id, Ticket.status.notin_(['Resolved', 'Closed'])); list_title = "Your Unsolved Tickets"; view_name = 'my_unsolved'
    
    if view_name in ['recently_updated', 'recently_solved']: 
        ordered_query = query.order_by(Ticket.updated_at.desc())
    elif view_name in ['all_unsolved', 'unassigned', 'pending']:
        ordered_query = query.order_by(Ticket.created_at.desc()) 
    else: 
        priority_order = db.case({'Urgent': 1, 'High': 2, 'Medium': 3, 'Low': 4}, value=Ticket.priority, else_=5)
        ordered_query = query.order_by(priority_order.asc(), Ticket.updated_at.desc())
        
    tickets_pagination = ordered_query.paginate(page=page, per_page=10, error_out=False)
    return render_template('agent/ticket_list.html', title=list_title, tickets_pagination=tickets_pagination, current_view=view_name)

@app.route('/ticket/<int:ticket_id>/assign_to_me')
@agent_required
def assign_ticket_to_me(ticket_id):
    ticket = db.session.get(Ticket, ticket_id) 
    if not ticket:
        flash('Ticket not found.', 'danger')
        return redirect(url_for('agent_ticket_list'))
        
    if ticket.assigned_to_id is None or ticket.assigned_to_id != current_user.id:
        old_assignee_name = ticket.assignee.username if ticket.assignee else "Unassigned"; old_status = ticket.status; status_changed = False
        ticket.assigned_to_id = current_user.id
        if ticket.status == 'Open': ticket.status = 'In Progress'; status_changed = True
        log_interaction(ticket.id, 'ASSIGNMENT_CHANGE', user_id=current_user.id, details={'old_value': old_assignee_name, 'new_value': current_user.username, 'field_display_name': 'Assignee'})
        if status_changed: log_interaction(ticket.id, 'STATUS_CHANGE', user_id=current_user.id, details={'old_value': old_status, 'new_value': ticket.status, 'field_display_name': 'Status'})
        db.session.commit(); flash(f'Ticket #{ticket.id} has been assigned to you.', 'success')
    elif ticket.assigned_to_id == current_user.id: flash(f'Ticket #{ticket.id} is already assigned to you.', 'info')
    else: flash(f'Ticket #{ticket.id} is already assigned to another agent ({ticket.assignee.username if ticket.assignee else "Unknown"}).', 'warning')
    return redirect(request.referrer or url_for('agent_ticket_list', view_name='unassigned'))

# --- Admin Routes ---
def _admin_list_options(model_class, template_name, title, order_by_attr='name'):
    items = model_class.query.order_by(getattr(model_class, order_by_attr)).all(); model_name_slug = to_snake_case(model_class.__name__)
    return render_template(template_name, title=title, items=items, model_name=model_name_slug)

def _admin_create_edit_option(model_class, form_class, list_url_func_name, item_id=None):
    item = db.session.get(model_class, item_id) if item_id else None
    form = form_class(obj=item if request.method == 'GET' and item else None)
    type_name_raw = model_class.__name__.replace("Option",""); type_name_display = " ".join(re.findall('[A-Z][^A-Z]*', type_name_raw) or [type_name_raw]);
    if not type_name_display.strip(): type_name_display = model_class.__name__ 
    legend = f'New {type_name_display}' if not item else f'Edit {type_name_display}: {getattr(item, "name", "Item")}'
    if form.validate_on_submit():
        is_new = (item is None); option_instance = item or model_class(); existing_item_with_name_query = model_class.query.filter(model_class.name == form.name.data)
        if option_instance.id: existing_item_with_name_query = existing_item_with_name_query.filter(model_class.id != option_instance.id)
        existing_item_with_name = existing_item_with_name_query.first()
        if existing_item_with_name: form.name.errors.append(f'The name "{form.name.data}" already exists for {type_name_display}.')
        if not form.errors:
            form.populate_obj(option_instance);
            if is_new: db.session.add(option_instance)
            try: db.session.commit(); flash(f'{type_name_display} "{option_instance.name}" saved.', 'success'); return redirect(url_for(list_url_func_name))
            except Exception as e: db.session.rollback(); flash(f'DB error saving {type_name_display}: {e}', 'danger'); app.logger.error(f"Error saving option {type_name_display} {option_instance.name}: {e}", exc_info=True)
    return render_template('admin/create_edit_option.html', title=legend, form=form, legend=legend,item_type_name=type_name_display, list_url=url_for(list_url_func_name))

def _admin_delete_option(model_class, item_id, list_url_func_name, related_ticket_attr_id=None, related_ticket_attr_name=None):
    item = db.session.get(model_class, item_id)
    if not item:
        flash(f'{model_class.__name__} not found.', 'danger')
        return redirect(url_for(list_url_func_name))
    item_name = getattr(item, "name", "Item"); type_name_raw = model_class.__name__.replace("Option",""); type_name_display = " ".join(re.findall('[A-Z][^A-Z]*', type_name_raw) or [type_name_raw])
    if not type_name_display.strip(): type_name_display = model_class.__name__
    can_delete = True; query_filter = None
    if related_ticket_attr_id and hasattr(Ticket, related_ticket_attr_id): query_filter = (getattr(Ticket, related_ticket_attr_id) == item.id)
    elif related_ticket_attr_name and hasattr(Ticket, related_ticket_attr_name): query_filter = (getattr(Ticket, related_ticket_attr_name) == item.name)
    if query_filter is not None and Ticket.query.filter(query_filter).first(): can_delete = False; flash(f'Cannot delete "{item_name}", as it is associated with one or more tickets.', 'danger')    
    if can_delete:
        try: db.session.delete(item); db.session.commit(); flash(f'{type_name_display} "{item_name}" deleted.', 'success')
        except Exception as e: db.session.rollback(); flash(f'Error deleting {type_name_display} "{item_name}": {e}', 'danger'); app.logger.error(f"Error deleting option {type_name_display} {item_name}: {e}", exc_info=True)
    return redirect(url_for(list_url_func_name))

@app.route('/admin/users')
@admin_required
def admin_user_list(): users = User.query.order_by(User.username).all(); share_form = ShareCredentialsForm(); return render_template('admin/user_list.html', title='Manage Users', users=users, share_form=share_form)

@app.route('/admin/user/new', methods=['GET', 'POST'])
@app.route('/admin/user/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_user(user_id=None):
    user_to_edit = db.session.get(User, user_id) if user_id else None
    form = AdminUserForm(obj=user_to_edit if request.method == 'GET' and user_to_edit else None)
    
    form.organization_id.choices = get_active_organization_choices() # Existing
    
    # Populate department choices based on current/selected organization
    # If editing, use the user's current org. If new, or if org changes via JS, this needs to be dynamic.
    # For now, initial population based on current user's org (if editing) or selected org on POST.
    # A full dynamic update would require JavaScript.
    current_org_id_for_dept_choices = None
    if request.method == 'POST': # Use submitted org_id if available
        current_org_id_for_dept_choices = form.organization_id.data
    elif user_to_edit and user_to_edit.organization_id: # Use user's current org if editing
        current_org_id_for_dept_choices = user_to_edit.organization_id
    
    form.department_id.choices = get_active_department_choices_for_org(current_org_id_for_dept_choices)

    legend = 'Create New User' if not user_to_edit else f'Edit User: {user_to_edit.username}'
    
    # Password validation logic (no changes needed here)
    original_password_validators = list(form.password.validators); original_password2_validators = list(form.password2.validators)
    if not user_to_edit: form.password.validators = [DataRequired(message="Password is required for new users.")] + [v for v in original_password_validators if not isinstance(v, Optional)]; form.password2.validators = [DataRequired(message="Please confirm the password.")] + [v for v in original_password2_validators if not isinstance(v, Optional)]
    else: form.password.validators = [Optional()] + [v for v in original_password_validators if not isinstance(v, (DataRequired, Optional))]; form.password2.validators = [EqualTo('password', message='Passwords must match if new password provided.')] + [Optional()] + [v for v in original_password2_validators if not isinstance(v, (DataRequired, Optional, EqualTo))]

    if form.validate_on_submit():
        is_new_user = (user_to_edit is None); user = user_to_edit or User(); user_id_to_exclude = user.id if user.id else -1
        if User.query.filter(User.username == form.username.data, User.id != user_id_to_exclude).first(): form.username.errors.append('This username is already taken.')
        if User.query.filter(User.email == form.email.data.lower(), User.id != user_id_to_exclude).first(): form.email.errors.append('This email address is already registered.')
        if not is_new_user and form.password.data and not form.password2.data: form.password2.errors.append("Please confirm the new password if you are changing it.")
        
        # Validate department only if an organization is selected
        selected_org_id = form.organization_id.data if form.organization_id.data != 0 else None
        selected_dept_id = form.department_id.data if form.department_id.data != 0 else None

        if selected_dept_id and not selected_org_id:
            form.department_id.errors.append("A department cannot be assigned without an organization.")
        
        if selected_dept_id and selected_org_id:
            dept_check = Department.query.filter_by(id=selected_dept_id, organization_id=selected_org_id).first()
            if not dept_check:
                form.department_id.errors.append("Selected department does not belong to the selected organization.")


        if not form.errors:
            user.username = form.username.data; user.email = form.email.data.lower(); user.role = form.role.data
            user.organization_id = selected_org_id # Already handled
            user.department_id = selected_dept_id # <-- ASSIGN DEPARTMENT

            if form.password.data: user.set_password(form.password.data)
            if is_new_user: db.session.add(user)
            try: 
                db.session.commit()
                flash(f'User "{user.username}" has been {"created" if is_new_user else "updated"} successfully.', 'success')
                return redirect(url_for('admin_user_list'))
            except Exception as e: 
                db.session.rollback()
                flash(f'Database error: Could not save user. {str(e)}', 'danger')
                app.logger.error(f"Admin user save error for '{form.username.data}': {e}", exc_info=True)
    
    elif request.method == 'GET' and user_to_edit:
        form.organization_id.data = user_to_edit.organization_id or 0
        # Ensure department choices are re-populated based on the user's current org
        form.department_id.choices = get_active_department_choices_for_org(user_to_edit.organization_id)
        form.department_id.data = user_to_edit.department_id or 0 # Populate current department

    # Fallback if it's a GET for a new user and no org_id has been selected yet for dept choices
    elif request.method == 'GET' and not user_to_edit:
         form.department_id.choices = get_active_department_choices_for_org(None)


    return render_template('admin/create_edit_user.html', title=legend, form=form, legend=legend, user=user_to_edit)

@app.route('/admin/user/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    user_to_delete = db.session.get(User, user_id)
    if not user_to_delete:
        flash('User not found.', 'danger')
        return redirect(url_for('admin_user_list'))
    if user_to_delete.id == current_user.id: flash('You cannot delete your own account.', 'danger')
    elif user_to_delete.is_admin and User.query.filter_by(role='admin').count() <= 1: flash('Cannot delete the only remaining administrator account.', 'danger')
    else:
        deleted_username = user_to_delete.username
        try:
            Ticket.query.filter_by(assigned_to_id=user_id).update({'assigned_to_id': None}); Attachment.query.filter_by(uploaded_by_id=user_id).delete(synchronize_session='fetch')
            Comment.query.filter_by(user_id=user_id).delete(synchronize_session='fetch'); Interaction.query.filter_by(user_id=user_id).update({'user_id': None}, synchronize_session='fetch')
            db.session.delete(user_to_delete); db.session.commit(); flash(f'User "{deleted_username}" and their associated comments/attachments deleted. Interactions anonymized.', 'success'); app.logger.info(f"Admin '{current_user.username}' deleted user '{deleted_username}'.")
        except Exception as e: db.session.rollback(); flash(f'Error deleting user "{deleted_username}": {e}', 'danger'); app.logger.error(f"Error deleting user {user_id}: {e}", exc_info=True)
    return redirect(url_for('admin_user_list'))

@app.route('/admin/user/<int:user_id>/share_credentials', methods=['POST'])
@admin_required
def admin_share_credentials(user_id):
    user_to_share = db.session.get(User, user_id)
    if not user_to_share:
        flash('User not found.', 'danger')
        return redirect(url_for('admin_user_list'))
    admin_user = current_user; form = ShareCredentialsForm(request.form)
    if form.validate():
        recipient_email = form.recipient_email.data; subject = f"Account Information for Ticket System: {user_to_share.username}"; body_text = render_template('email/share_credentials_email.txt', user_being_shared=user_to_share, admin_user=admin_user)
        msg = Message(subject, recipients=[recipient_email], body=body_text)
        try: mail.send(msg); flash(f'Account info for "{user_to_share.username}" sent to {recipient_email}.', 'success'); app.logger.info(f"Admin '{admin_user.username}' shared credentials info for '{user_to_share.username}' with '{recipient_email}'.")
        except Exception as e: flash(f'Failed to send email: {e}', 'danger'); app.logger.error(f"Failed to send share credentials email for {user_to_share.username}: {e}", exc_info=True)
    else:
        for field_name, errors in form.errors.items():
            label = getattr(getattr(form, field_name), 'label', None); label_text = label.text if label else field_name.replace("_", " ").title(); flash(f"Error in sharing form ({label_text}): {', '.join(errors)}", 'danger')
    return redirect(url_for('admin_user_list'))





@app.route('/api/ai/summarize_ticket/<int:ticket_id>', methods=['POST'])
@agent_required
def ai_summarize_ticket(ticket_id):
    if not app.config['GEMINI_API_KEY']:
        return jsonify({"error": "AI service unavailable. Key not configured."}), 503

    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        return jsonify({"error": "Ticket not found."}), 404

    # Assemble the prompt context from the ticket and public comments
    comments = ticket.comments.filter_by(is_internal=False).order_by(Comment.created_at.asc()).all()
    conversation_history = [f"{c.author.username}: {c.content}" for c in comments if c.author]
    formatted_history = "\n".join(conversation_history)

    prompt = f"""
You are an expert AI assistant for a technical support team. Your task is to summarize a support ticket to help an agent quickly understand the situation.

Analyze the entire ticket history provided below, including the original description and the conversation.

Create a concise summary that includes:
1.  **The Core Problem:** What is the main issue the user is facing?
2.  **Key Information Provided:** Mention any critical details, error messages, or specific configurations the user has shared.
3.  **Current Status/Last Action:** What was the last thing that happened? Is the agent waiting for the customer, or vice-versa?

Format the output clearly using Markdown for headings and bullet points.

--- TICKET DETAILS ---

**Ticket Title:** {ticket.title}

**Original Description by {ticket.creator.username if ticket.creator else 'User'}:**
{ticket.description}

--- CONVERSATION HISTORY ---
{formatted_history if formatted_history else "No public comments have been made."}

--- SUMMARY ---
"""

    app.logger.info(f"AI Summary Prompt for ticket {ticket_id} (first 150 chars): {prompt[:150]}...")
    
    try:
        model = genai.GenerativeModel('gemini-1.5-flash-latest')
        response = model.generate_content(
            prompt,
            generation_config=genai.types.GenerationConfig(
                max_output_tokens=800,
                temperature=0.3 # Lower temperature for factual summarization
            )
        )
        generated_text = "".join(part.text for part in response.candidates[0].content.parts).strip() if response.candidates and response.candidates[0].content.parts else ""
        
        if not generated_text:
            # Handle cases where the AI returns no content
            return jsonify({"error": "AI could not generate a summary for this ticket."})
            
        return jsonify({"summary": generated_text})

    except Exception as e:
        app.logger.error(f"Gemini API error during AI summary for ticket {ticket_id}: {str(e)}", exc_info=True)
        return jsonify({"error": "An error occurred while communicating with the AI service."}), 500





@app.route('/ticket/<int:ticket_id>/client_close', methods=['POST'])
@login_required
def client_close_ticket(ticket_id):
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        flash('Ticket not found.', 'danger')
        return redirect(url_for('dashboard'))

    # --- Permission Check ---
    can_modify = False
    if ticket.created_by_id == current_user.id:
        can_modify = True
    elif current_user.role == 'client' and ticket.department_id and ticket.department_id == current_user.department_id:
        can_modify = True
    elif current_user.role == 'organization_client' and ticket.organization_id and ticket.organization_id == current_user.organization_id:
        can_modify = True
    
    if not can_modify:
        flash('You do not have permission to modify this ticket.', 'danger')
        return redirect(url_for('view_ticket', ticket_id=ticket.id))

    if ticket.status in ['Resolved', 'Closed']:
        flash('This ticket is already resolved or closed.', 'info')
        return redirect(url_for('view_ticket', ticket_id=ticket.id))

    old_status = ticket.status
    ticket.status = 'Closed'
    if not ticket.resolved_at:
        ticket.resolved_at = datetime.utcnow()
    if ticket.created_at and ticket.resolved_at:
        delta = ticket.resolved_at - ticket.created_at
        ticket.total_resolution_duration_minutes = int(delta.total_seconds() / 60)
    
    log_interaction(ticket.id, 'STATUS_CHANGE_BY_CLIENT', user_id=current_user.id,
                    details={'old_value': old_status, 'new_value': 'Closed', 'field_display_name': 'Status'})
    
    db.session.commit()
    flash(f'Ticket #{ticket.id} has been closed.', 'success')
    return redirect(url_for('view_ticket', ticket_id=ticket.id))


@app.route('/ticket/<int:ticket_id>/client_reopen', methods=['POST'])
@login_required
def client_reopen_ticket(ticket_id):
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        flash('Ticket not found.', 'danger')
        return redirect(url_for('dashboard'))

    # --- Permission Check (same as close) ---
    can_modify = False
    if ticket.created_by_id == current_user.id:
        can_modify = True
    elif current_user.role == 'client' and ticket.department_id and ticket.department_id == current_user.department_id:
        can_modify = True
    elif current_user.role == 'organization_client' and ticket.organization_id and ticket.organization_id == current_user.organization_id:
        can_modify = True

    if not can_modify:
        flash('You do not have permission to modify this ticket.', 'danger')
        return redirect(url_for('view_ticket', ticket_id=ticket.id))

    if ticket.status not in ['Resolved', 'Closed']:
        flash('This ticket is not resolved or closed.', 'info')
        return redirect(url_for('view_ticket', ticket_id=ticket.id))

    old_status = ticket.status
    ticket.status = 'Open' # Reopen to 'Open' status
    ticket.resolved_at = None
    ticket.total_resolution_duration_minutes = None

    log_interaction(ticket.id, 'STATUS_CHANGE_BY_CLIENT', user_id=current_user.id,
                    details={'old_value': old_status, 'new_value': 'Open', 'field_display_name': 'Status'})
    
    db.session.commit()
    flash(f'Ticket #{ticket.id} has been reopened.', 'success')
    return redirect(url_for('view_ticket', ticket_id=ticket.id))



@app.route('/admin/categories')
@admin_required
def admin_category_list(): return _admin_list_options(Category, 'admin/list_options.html', 'Manage Categories')
@app.route('/admin/category/new', methods=['GET', 'POST'])
@app.route('/admin/category/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_category(item_id=None): return _admin_create_edit_option(Category, CategoryForm, 'admin_category_list', item_id)
@app.route('/admin/category/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_category(item_id): return _admin_delete_option(Category, item_id, 'admin_category_list', related_ticket_attr_id='category_id')

@app.route('/admin/cloud_providers')
@admin_required
def admin_cloud_provider_list(): return _admin_list_options(CloudProviderOption, 'admin/list_options.html', 'Manage Cloud Providers')
@app.route('/admin/cloud_provider/new', methods=['GET', 'POST'])
@app.route('/admin/cloud_provider/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_cloud_provider(item_id=None): return _admin_create_edit_option(CloudProviderOption, CloudProviderOptionForm, 'admin_cloud_provider_list', item_id)
@app.route('/admin/cloud_provider/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_cloud_provider(item_id): return _admin_delete_option(CloudProviderOption, item_id, 'admin_cloud_provider_list', related_ticket_attr_name='cloud_provider')

@app.route('/admin/severities')
@admin_required
def admin_severity_list(): return _admin_list_options(SeverityOption, 'admin/list_options.html', 'Manage Severity Levels', 'order')
@app.route('/admin/severity/new', methods=['GET', 'POST'])
@app.route('/admin/severity/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_severity(item_id=None): return _admin_create_edit_option(SeverityOption, SeverityOptionForm, 'admin_severity_list', item_id)
@app.route('/admin/severity/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_severity(item_id): return _admin_delete_option(SeverityOption, item_id, 'admin_severity_list', related_ticket_attr_name='severity')

@app.route('/admin/environments')
@admin_required
def admin_environment_list(): return _admin_list_options(EnvironmentOption, 'admin/list_options.html', 'Manage Environments')
@app.route('/admin/environment/new', methods=['GET', 'POST'])
@app.route('/admin/environment/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_environment(item_id=None): return _admin_create_edit_option(EnvironmentOption, EnvironmentOptionForm, 'admin_environment_list', item_id)
@app.route('/admin/environment/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_environment(item_id): return _admin_delete_option(EnvironmentOption, item_id, 'admin_environment_list', related_ticket_attr_name='environment')

@app.route('/admin/organizations')
@admin_required
def admin_organization_list(): return _admin_list_options(OrganizationOption, 'admin/list_options.html', 'Manage Organizations')
@app.route('/admin/organization/new', methods=['GET', 'POST'])
@app.route('/admin/organization/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_organization(item_id=None): return _admin_create_edit_option(OrganizationOption, OrganizationOptionForm, 'admin_organization_list', item_id)
@app.route('/admin/organization/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_organization(item_id): return _admin_delete_option(OrganizationOption, item_id, 'admin_organization_list', related_ticket_attr_id='organization_id')

@app.route('/admin/form_types')
@admin_required
def admin_form_type_list(): return _admin_list_options(FormTypeOption, 'admin/list_options.html', 'Manage Form Types')
@app.route('/admin/form_type/new', methods=['GET', 'POST'])
@app.route('/admin/form_type/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_form_type(item_id=None): return _admin_create_edit_option(FormTypeOption, FormTypeOptionForm, 'admin_form_type_list', item_id)
@app.route('/admin/form_type/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_form_type(item_id): return _admin_delete_option(FormTypeOption, item_id, 'admin_form_type_list', related_ticket_attr_id='form_type_id')

@app.route('/admin/apn_opportunities')
@admin_required
def admin_apn_opportunity_list(): return _admin_list_options(APNOpportunityOption, 'admin/list_options.html', 'Manage APN Opportunities')
@app.route('/admin/apn_opportunity/new', methods=['GET', 'POST'])
@app.route('/admin/apn_opportunity/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_apn_opportunity(item_id=None): return _admin_create_edit_option(APNOpportunityOption, APNOpportunityOptionForm, 'admin_apn_opportunity_list', item_id)
@app.route('/admin/apn_opportunity/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_apn_opportunity(item_id): return _admin_delete_option(APNOpportunityOption, item_id, 'admin_apn_opportunity_list', related_ticket_attr_id='apn_opportunity_id')

@app.route('/admin/support_modals')
@admin_required
def admin_support_modal_list(): return _admin_list_options(SupportModalOption, 'admin/list_options.html', 'Manage Support Modals by Plan')
@app.route('/admin/support_modal/new', methods=['GET', 'POST'])
@app.route('/admin/support_modal/<int:item_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_create_edit_support_modal(item_id=None): return _admin_create_edit_option(SupportModalOption, SupportModalOptionForm, 'admin_support_modal_list', item_id)
@app.route('/admin/support_modal/<int:item_id>/delete', methods=['POST'])
@admin_required
def admin_delete_support_modal(item_id): return _admin_delete_option(SupportModalOption, item_id, 'admin_support_modal_list', related_ticket_attr_id='support_modal_id')
# In app.py

@app.route('/admin/tickets')
@admin_required
def admin_all_tickets():
    page = request.args.get('page', 1, type=int)
    filters = {k: v for k, v in request.args.items() if k != 'page' and v and v != ""}
    query = Ticket.query

    if filters.get('status'):
        query = query.filter(Ticket.status == filters['status'])
    if filters.get('priority'):
        query = query.filter(Ticket.priority == filters['priority'])
    if filters.get('category_id') and filters['category_id'] != '0':
        query = query.filter(Ticket.category_id == int(filters['category_id']))
    
    assignee_id_filter = filters.get('assigned_to_id')
    if assignee_id_filter:
        if assignee_id_filter == '0': 
            query = query.filter(Ticket.assigned_to_id.is_(None))
        elif assignee_id_filter.isdigit():
            query = query.filter(Ticket.assigned_to_id == int(assignee_id_filter))
            
    organization_id_filter = filters.get('organization_id')
    if organization_id_filter and organization_id_filter.isdigit() and organization_id_filter != '0':
        query = query.filter(Ticket.organization_id == int(organization_id_filter))
        
        # --- NEW: Handle Department Filter (only if organization is also selected) ---
        department_id_filter = filters.get('department_id')
        if department_id_filter and department_id_filter.isdigit() and department_id_filter != '0':
            # Ensure the department belongs to the selected organization for data integrity,
            # though the dynamic dropdown should prevent mismatches.
            query = query.filter(Ticket.department_id == int(department_id_filter))
    # --- END: Handle Department Filter ---

    tickets_pagination = query.order_by(Ticket.updated_at.desc()).paginate(page=page, per_page=10, error_out=False)
    
    categories_for_filter = Category.query.order_by('name').all()
    agents_for_filter = User.query.filter(User.role.in_(['agent', 'admin'])).order_by('username').all()
    organizations_for_filter = OrganizationOption.query.filter_by(is_active=True).order_by('name').all()
    
    # For pre-populating department filter if an org is already selected
    departments_for_filter = []
    if organization_id_filter and organization_id_filter.isdigit() and organization_id_filter != '0':
        departments_for_filter = Department.query.filter_by(
            organization_id=int(organization_id_filter), 
            is_active=True
        ).order_by(Department.name).all()


    return render_template('admin/all_tickets.html', 
                           title='All Tickets Overview', 
                           tickets_pagination=tickets_pagination, 
                           statuses=TICKET_STATUS_CHOICES, 
                           priorities=TICKET_PRIORITY_CHOICES, 
                           categories=categories_for_filter, 
                           agents=agents_for_filter,
                           organizations_for_filter=organizations_for_filter,
                           departments_for_filter=departments_for_filter, # Pass departments
                           current_filters=filters)
# --- NEW FEATURE ROUTES ---



# In app.py (ensure this existing route has appropriate permissions or create a new one)

@app.route('/api/departments_for_organization/<int:organization_id>')
@login_required # Change to @admin_required if this endpoint is specifically for admin use
def api_departments_for_organization(organization_id):
    # Allow '0' to mean "no organization" or "all departments" if that's a use case.
    # For admin filter, "0" usually means don't filter by department for this org.
    if organization_id == 0: 
        # For the admin filter, if no org is selected, the dept dropdown should be disabled
        # or show "Select Org First". If an org IS selected, then 0 for dept means "All Depts in Org".
        # This API will be called when an org IS selected.
        return jsonify([{'id': 0, 'name': '--- All Departments in Org ---'}]) 
        
    departments = Department.query.filter_by(organization_id=organization_id, is_active=True).order_by(Department.name).all()
    dept_list = [{'id': dept.id, 'name': dept.name} for dept in departments]
    
    # For the filter, "0" can mean "All departments in this specific organization"
    # The first "--- Select ---" option will be added by JS
    return jsonify([{'id': 0, 'name': '--- All Departments in Org ---'}] + dept_list if dept_list else [{'id': 0, 'name': '--- No Departments in this Org ---'}])

@app.route('/api/ai/generate_ticket_description', methods=['POST'])
@login_required
def ai_generate_ticket_description():
    if not app.config['GEMINI_API_KEY']:
        return jsonify({"error": "AI service unavailable. Key not configured."}), 503

    data = request.get_json()
    if not data:
        app.logger.error("AI Description: No JSON payload received.")
        return jsonify({"error": "Invalid payload. Expecting JSON."}), 400

    title = data.get('title', '').strip()
    current_description = data.get('current_description', '').strip() 
    category_id = data.get('category_id')
    severity_name = data.get('severity_name', '').strip()

    category_name = ""
    if category_id:
        try:
            cat_id_int = int(category_id)
            if cat_id_int != 0: 
                category_obj = db.session.get(Category, cat_id_int)
                if category_obj:
                    category_name = category_obj.name
        except ValueError:
            app.logger.warning(f"Invalid category_id format: {category_id}")
    
    prompt_parts = ["You are an expert technical writer for a ticketing system. Your task is to refine or generate a bug/issue description."]
    context_provided = False
    if current_description:
        prompt_parts.append(f"Given the following user-submitted description:\n'''\n{current_description}\n'''")
        context_provided = True
    if title:
        prompt_parts.append(f"The ticket title is: \"{title}\"")
        context_provided = True
    if category_name:
        prompt_parts.append(f"The ticket category is: \"{category_name}\"")
        context_provided = True
    if severity_name:
        prompt_parts.append(f"The ticket severity is: \"{severity_name}\"")
        context_provided = True

    if not context_provided:
        return jsonify({"error": "Please provide at least a title, current description, category, or severity to generate an AI description."}), 400

    if current_description:
        prompt_parts.append("\nRewrite this into a clear, professional, and concise description suitable for a technical support ticket. Focus on clarity and completeness. If possible, infer context, potential steps to reproduce (if logical from the input), expected behavior, and actual behavior. Output only the refined description text. Do not add any preamble like 'Here is the rewritten description:'.")
    else:
        prompt_parts.append("\nBased on the provided ticket information (title, category, severity), generate a detailed and professional bug/issue description. If the information is too generic, try to expand on common issues related to such a context. Include potential steps to reproduce, expected behavior, and actual behavior where appropriate. Output only the generated description text. Do not add any preamble.")
    
    prompt = "\n\n".join(prompt_parts)
    app.logger.info(f"AI Ticket Description Prompt (first 100 chars): {prompt[:100]}...")

    try:
        model = genai.GenerativeModel('gemini-1.5-flash-latest') 
        response = model.generate_content(
            prompt,
            generation_config=genai.types.GenerationConfig(
                max_output_tokens=800,
                temperature=0.5      
            )
        )
        generated_text = "".join(part.text for part in response.candidates[0].content.parts).strip() if response.candidates and response.candidates[0].content.parts else ""
        
        if not generated_text:
            if response.prompt_feedback and response.prompt_feedback.block_reason:
                block_msg = response.prompt_feedback.block_reason_message or "Content generation was limited."
                app.logger.warning(f"Gemini blocked AI ticket description generation: {block_msg}")
                return jsonify({"error": f"AI content generation was blocked: {block_msg}. Please rephrase your input or try again."}), 400
            app.logger.warning(f"Gemini returned empty content for AI ticket description. Prompt was: {prompt[:200]}...")
            return jsonify({"generated_description": "AI could not generate a description for the provided input. Please try rephrasing or adding more details."})
            
        return jsonify({"generated_description": generated_text})

    except Exception as e:
        app.logger.error(f"Gemini API error during AI ticket description generation: {str(e)}", exc_info=True)
        return jsonify({"error": "An error occurred while communicating with the AI service."}), 500

@app.route('/tools/gdoc_import', methods=['GET'])
@login_required
def gdoc_importer_page():
    return render_template('tools/gdoc_import.html', 
                           username=current_user.username,
                           title="Import Ticket from Google Doc")

@app.route('/api/extract_gdoc_content', methods=['POST'])
@login_required
def api_extract_gdoc_content():
    data = request.get_json()
    if not data or 'gdoc_url' not in data:
        return jsonify({"success": False, "detail": "Google Doc URL is required."}), 400
    
    gdoc_url = data.get('gdoc_url').strip()
    if not gdoc_url:
        return jsonify({"success": False, "detail": "Google Doc URL cannot be empty."}), 400

    is_likely_edit_url = "/edit" in gdoc_url.split('?')[0].split('#')[0] 
    is_likely_publish_url = ("/d/e/" in gdoc_url or "/pub" in gdoc_url)

    if is_likely_edit_url and not is_likely_publish_url : 
        app.logger.warning(f"GDoc Importer: Received likely editor URL: {gdoc_url}")
        return jsonify({"success": False, "detail": "Please use a 'Published to the web' link, not the direct editor link. (File > Share > Publish to web)"}), 400
    
    if not is_likely_publish_url:
        app.logger.warning(f"GDoc Importer: URL doesn't match typical publish patterns: {gdoc_url}. Allowing attempt.")

    app.logger.info(f"Attempting to extract content from GDoc URL: {gdoc_url}")
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        response = requests.get(gdoc_url, timeout=20, headers=headers) 
        
        if 'text/html' not in response.headers.get('Content-Type', '').lower():
            app.logger.error(f"GDoc URL did not return HTML. Content-Type: {response.headers.get('Content-Type')}. URL: {gdoc_url}")
            return jsonify({"success": False, "detail": "The URL did not return an HTML document. Check if it's correctly published and accessible."}), 400

        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        content_div = soup.find('div', id='contents')
        if not content_div: 
            body_content = soup.body
            if body_content:
                for tag_to_remove in body_content(['header', 'footer', 'nav', 'aside', 'script', 'style']):
                    tag_to_remove.decompose()
                content_div = body_content
            if not content_div:
                app.logger.warning(f"Could not find main content area (div#contents or body) in GDoc: {gdoc_url}")
                return jsonify({"success": False, "detail": "Could not find main content area in the document. Structure might be unexpected."}), 400
        
        for s_or_s_tag in content_div(['script', 'style']):
            s_or_s_tag.decompose()
        extracted_text = content_div.get_text(separator='\n', strip=True)
        
        if not extracted_text.strip():
            app.logger.warning(f"No text content found in the GDoc after cleaning: {gdoc_url}")
            return jsonify({"success": False, "detail": "No text content found in the document after cleaning."}), 400
        return jsonify({"success": True, "content": extracted_text})
    except requests.exceptions.Timeout:
        app.logger.error(f"Timeout fetching GDoc URL: {gdoc_url}")
        return jsonify({"success": False, "detail": "The request to Google Docs timed out. Please try again."}), 504
    except requests.exceptions.HTTPError as http_err:
        app.logger.error(f"HTTP error fetching GDoc {gdoc_url}: {http_err}")
        if http_err.response.status_code == 404:
            return jsonify({"success": False, "detail": "Document not found (404). Check the URL or publish settings."}), 404
        return jsonify({"success": False, "detail": f"Error fetching document (status {http_err.response.status_code}). Check if the document is 'Published to web' and accessible without login."}), http_err.response.status_code
    except requests.exceptions.RequestException as req_err:
        app.logger.error(f"Request error fetching GDoc {gdoc_url}: {req_err}")
        return jsonify({"success": False, "detail": f"Network error fetching document: {req_err}"}), 500
    except Exception as e:
        app.logger.error(f"Error parsing GDoc {gdoc_url}: {e}", exc_info=True)
        return jsonify({"success": False, "detail": "An error occurred while parsing the document content."}), 500

@app.route('/api/create_ticket_from_gdoc_data', methods=['POST'])
@login_required
def api_create_ticket_from_gdoc_data():
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "detail": "Invalid JSON payload"}), 400

    title = data.get('title')
    description = data.get('remedies') 
    category_id_str = data.get('category_id')
    severity_name = data.get('severity_name')
    customer_name = data.get('customer_name')

    errors = []
    if not title: errors.append("Title is required.")
    if not description: errors.append("Description (from GDoc content) is required.")
    if not category_id_str or category_id_str == "0": errors.append("Category is required.")
    if not severity_name: errors.append("Severity is required.")
    if not customer_name: errors.append("Customer Name is required.")

    if errors:
        return jsonify({"success": False, "detail": " ".join(errors)}), 400

    category_id = int(category_id_str)
    category = db.session.get(Category, category_id)
    severity_obj = SeverityOption.query.filter_by(name=severity_name, is_active=True).first()

    if not category:
        return jsonify({"success": False, "detail": "Invalid Category selected."}), 400
    if not severity_obj:
        return jsonify({"success": False, "detail": "Invalid Severity selected."}), 400

    ticket_creator = current_user
    ticket_org_id_to_save = None
    if current_user.is_client and current_user.organization_id:
        ticket_org_id_to_save = current_user.organization_id
    elif current_user.organization:
        ticket_org_id_to_save = current_user.organization_id
    
    ticket = Ticket(
        title=title, description=description, created_by_id=ticket_creator.id,
        category_id=category.id, severity=severity_obj.name,
        customer_name=customer_name, organization_id=ticket_org_id_to_save,
        status='Open', priority='Medium' 
    )
    if "Critical" in severity_obj.name or "Urgent" in severity_obj.name: ticket.priority = "Urgent"
    elif "High" in severity_obj.name: ticket.priority = "High"
    elif "Medium" in severity_obj.name: ticket.priority = "Medium"
    else: ticket.priority = "Low"

    db.session.add(ticket)
    try:
        db.session.flush() 
        log_interaction(ticket.id, 'TICKET_CREATED_GDOC', user_id=ticket_creator.id, details={'title': ticket.title, 'source': 'Google Doc Import'}, timestamp_override=ticket.created_at)
        db.session.commit()
        try:
            admin_and_agent_emails = list(set(([app.config['ADMIN_EMAIL']] if app.config['ADMIN_EMAIL'] else []) + [user.email for user in User.query.filter(User.role.in_(['admin', 'agent'])).all() if user.email]))
            if admin_and_agent_emails:
                mail.send(Message(subject=f"New Ticket (GDoc): #{ticket.id} - {ticket.title}", recipients=admin_and_agent_emails, body=render_template('email/new_ticket_admin_notification.txt', ticket=ticket, submitter=ticket_creator, ticket_url=url_for('view_ticket', ticket_id=ticket.id, _external=True))))
            creator_email_list = [ticket_creator.email] if ticket_creator.email else []
            if creator_email_list:
                 mail.send(Message(subject=f"Confirmation (GDoc): Your Ticket #{ticket.id} - {ticket.title}", recipients=creator_email_list, body=render_template('email/ticket_info_recipient.txt', ticket=ticket, submitter=ticket_creator, ticket_url=url_for('view_ticket', ticket_id=ticket.id, _external=True))))
        except Exception as e:
            app.logger.error(f"Failed to send email notifications for GDoc ticket #{ticket.id}: {e}", exc_info=True)
        trigger_priority_call_alert(ticket, old_severity=None)
        flash(f'Ticket #{ticket.id} created successfully from Google Doc!', 'success')
        return jsonify({"success": True, "message": f"Ticket #{ticket.id} created successfully!", "ticket_id": ticket.id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"GDoc Ticket creation DB error: {e}", exc_info=True)
        return jsonify({"success": False, "detail": f"Database error creating ticket: {str(e)[:150]}"}), 500

# --- Chatbot ---
DIRECT_AI_SIGNAL_TYPE = "DIRECT_AI_FOR_GENERAL_QUERY_TYPE" 
TICKET_DETAILS_TYPE = "TICKET_DETAILS_TYPE"

def query_database_for_chatbot_sqlalchemy(user_message_lower):
    response_data = "Sorry, I couldn't find specific information related to your query in our ticket system. You can try asking about: 'ticket id 123', 'tickets by user@example.com', 'open tickets', 'search tickets for [keyword]', 'how many open tickets are there?', or 'show me the latest 3 tickets'."
    found_specific_query = False
    try:
        ticket_id_phrases = ["ticket id ", "show ticket ", "details for ticket ", "ticket #", "ticket#", "ticket "]
        matched_phrase = next((p for p in ticket_id_phrases if user_message_lower.startswith(p)), None)
        tid_str = None
        if matched_phrase:
            potential_id = user_message_lower[len(matched_phrase):].strip().split(" ")[0]
            if potential_id.isdigit(): tid_str = potential_id
        elif user_message_lower.split()[-1].isdigit() and any(kw in user_message_lower for kw in ["ticket", "id"]):
            tid_str = user_message_lower.split()[-1]

        if tid_str:
            found_specific_query = True
            try:
                tid = int(tid_str)
                ticket = db.session.get(Ticket, tid)
                if ticket:
                    text_parts = [
                        f"Ticket ID: {ticket.id}", f"Title: {ticket.title}", f"Status: {ticket.status}",
                        f"Priority: {ticket.priority}", f"Severity: {ticket.severity or 'N/A'}"
                    ]
                    if ticket.description and ticket.description.strip(): text_parts.append(f"Description: {ticket.description[:200]}...")
                    text_parts.append(f"Created By: {ticket.creator.username} on {ticket.created_at.strftime('%Y-%m-%d %H:%M')}")
                    if ticket.assignee: text_parts.append(f"Assigned To: {ticket.assignee.username}")
                    
                    ticket_data_for_response = {
                        "type": TICKET_DETAILS_TYPE, "id": ticket.id, "title": ticket.title, "status": ticket.status,
                        "description": ticket.description, "created_by": ticket.creator.username,
                        "created_at": ticket.created_at.strftime('%Y-%m-%d %H:%M:%S'), 
                        "severity": ticket.severity, "priority": ticket.priority,
                        "assignee": ticket.assignee.username if ticket.assignee else "Unassigned",
                        "category": ticket.category_ref.name if ticket.category_ref else "N/A",
                        "attachments_info": [],
                        "summary_text_for_ai": "\n".join(text_parts)
                    }
                    attachments = ticket.ticket_attachments.all()
                    if attachments:
                        for att in attachments:
                            try:
                                ticket_data_for_response["attachments_info"].append({
                                    "name": att.filename,
                                    "url": url_for('uploaded_file', filename=att.stored_filename, _external=False),
                                    "type": "attachment"
                                })
                                text_parts.append(f"Attachment: {att.filename}")
                            except Exception as e_url:
                                app.logger.error(f"Error generating URL for attachment {att.filename} in chatbot: {e_url}")
                        ticket_data_for_response["summary_text_for_ai"] = "\n".join(text_parts)
                    response_data = ticket_data_for_response
                else: response_data = f"Sorry, I couldn't find any ticket with ID {tid}."
            except ValueError: response_data = f"The ticket ID '{tid_str}' doesn't seem to be a valid number."
            except Exception as e:
                app.logger.error(f"Chatbot error fetching ticket ID {tid_str} (SQLAlchemy): {e}", exc_info=True)
                response_data = "I encountered an error trying to fetch the ticket details."
            
        elif any(user_message_lower.startswith(p) for p in ["tickets by ", "show tickets for user ", "tickets for "]):
            found_specific_query = True
            s_query = user_message_lower.split(" ", 2)[-1].strip() 
            if not s_query: response_data = "Please specify a username or email to search for (e.g., 'tickets by agent1')."
            else:
                tickets = Ticket.query.join(User, Ticket.created_by_id == User.id)\
                               .filter(User.username.ilike(f"%{s_query}%"))\
                               .order_by(Ticket.created_at.desc()).limit(5).all()
                if tickets: response_data = f"Here are the latest 5 tickets for users matching '{s_query}':\n" + "\n".join([f"- ID {t.id}: {t.title} (Status: {t.status}, Created: {t.created_at.strftime('%Y-%m-%d')})" for t in tickets])
                else: response_data = f"No tickets found for users matching '{s_query}'."
        
        elif any(keyword in user_message_lower for keyword in [" tickets", " status is "]) and \
             any(status_keyword in user_message_lower for status_keyword in TICKET_STATUS_CHOICES_FLAT + ['pending']): 
            found_specific_query = True
            status_to_find = None
            for status_key in TICKET_STATUS_CHOICES_FLAT:
                if status_key in user_message_lower:
                    status_to_find = next((s_disp for s_val, s_disp in TICKET_STATUS_CHOICES if s_val.lower() == status_key), None)
                    break
            if 'pending' in user_message_lower and not status_to_find : 
                 on_hold_value = next((s_val for s_val, s_disp in TICKET_STATUS_CHOICES if s_disp.lower() == 'on hold'), None)
                 if on_hold_value : status_to_find = on_hold_value
            if status_to_find:
                tickets = Ticket.query.filter(Ticket.status == status_to_find)\
                               .order_by(Ticket.created_at.desc()).limit(5).all()
                if tickets: response_data = f"Here are the latest 5 '{status_to_find}' tickets:\n" + "\n".join([f"- ID {t.id}: {t.title} (By: {t.creator.username})" for t in tickets])
                else: response_data = f"No '{status_to_find}' tickets found currently."
            else: response_data = f"Which status are you interested in (e.g., {', '.join(TICKET_STATUS_CHOICES_FLAT)})?"

        elif user_message_lower.startswith("search tickets for ") or user_message_lower.startswith("find tickets about "):
            found_specific_query = True
            search_term = user_message_lower.replace("search tickets for ", "").replace("find tickets about ","").strip()
            if search_term:
                like_term = f"%{search_term}%"
                tickets = Ticket.query.filter(
                                or_(Ticket.title.ilike(like_term), 
                                    Ticket.description.ilike(like_term),
                                    Ticket.tags.ilike(like_term))
                               ).order_by(Ticket.created_at.desc()).limit(5).all()
                if tickets: response_data = f"Found up to 5 tickets matching '{search_term}':\n" + "\n".join([f"- ID {t.id}: {t.title} (Status: {t.status})" for t in tickets])
                else: response_data = f"No tickets found matching '{search_term}'."
            else: response_data = "Please specify what you want to search for (e.g., 'search tickets for login issue')."

        elif user_message_lower.startswith("how many tickets are ") or user_message_lower.startswith("count of "):
            found_specific_query = True
            status_to_count = None
            for status_key in TICKET_STATUS_CHOICES_FLAT:
                if status_key in user_message_lower:
                    status_to_count = next((s_disp for s_val, s_disp in TICKET_STATUS_CHOICES if s_val.lower() == status_key), None)
                    break
            if 'pending' in user_message_lower and not status_to_count :
                 on_hold_value = next((s_val for s_val, s_disp in TICKET_STATUS_CHOICES if s_disp.lower() == 'on hold'), None)
                 if on_hold_value : status_to_count = on_hold_value
            if status_to_count:
                count = Ticket.query.filter(Ticket.status == status_to_count).count()
                response_data = f"There are {count} ticket(s) with status '{status_to_count}'."
            elif "total tickets" in user_message_lower or "all tickets" in user_message_lower:
                count = Ticket.query.count()
                response_data = f"There are a total of {count} ticket(s) in the system."
            else: response_data = "Which status count are you interested in (e.g., 'how many tickets are open')?"
        
        elif "latest " in user_message_lower and " tickets" in user_message_lower:
            found_specific_query = True
            try:
                parts = user_message_lower.split()
                num_tickets = None
                for i, part in enumerate(parts):
                    if part == "latest" and i + 1 < len(parts) and parts[i+1].isdigit():
                        num_tickets = int(parts[i+1]); break
                if num_tickets is not None:
                    num_tickets = min(num_tickets, 10) 
                    tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(num_tickets).all()
                    if tickets: response_data = f"Here are the latest {len(tickets)} tickets:\n" + "\n".join([f"- ID {t.id}: {t.title} (Status: {t.status})" for t in tickets])
                    else: response_data = "No tickets found."
                else: response_data = "Please specify how many latest tickets you want (e.g., 'latest 5 tickets')."
            except ValueError: response_data = "Please specify a valid number for latest tickets."
            except Exception as e:
                app.logger.error(f"Chatbot error fetching latest tickets (SQLAlchemy): {e}", exc_info=True)
                response_data = "I encountered an error trying to fetch the latest tickets."

        elif any(user_message_lower.startswith(p) for p in ["who created ticket ", "creator of ticket "]):
            found_specific_query = True
            try:
                creator_tid_str = None
                for p in ["who created ticket ", "creator of ticket "]:
                    if user_message_lower.startswith(p):
                        potential_id = user_message_lower[len(p):].strip().split(" ")[0]
                        if potential_id.isdigit(): creator_tid_str = potential_id; break
                if not creator_tid_str and user_message_lower.split()[-1].isdigit() and any(kw in user_message_lower for kw in ["ticket", "id", "creator"]):
                     creator_tid_str = user_message_lower.split()[-1]
                if not creator_tid_str: response_data = "Please provide a valid ticket ID to find its creator (e.g., 'creator of ticket 123')."
                else:
                    tid = int(creator_tid_str)
                    ticket = db.session.get(Ticket, tid)
                    if ticket: response_data = f"Ticket ID {tid} (\"{ticket.title}\") was created by: {ticket.creator.username}."
                    else: response_data = f"No ticket found with ID {tid}."
            except ValueError: response_data = "Please provide a valid ticket ID (must be a number)."
            except Exception as e:
                app.logger.error(f"Chatbot error fetching ticket creator for ID '{creator_tid_str}' (SQLAlchemy): {e}", exc_info=True)
                response_data = "I encountered an error trying to find the ticket creator."

        if not found_specific_query and len(user_message_lower.split()) > 1:
            like_term = f"%{user_message_lower}%"
            tickets = Ticket.query.join(User, Ticket.created_by_id == User.id).filter(
                or_(Ticket.title.ilike(like_term), Ticket.description.ilike(like_term), User.username.ilike(like_term))
            ).order_by(Ticket.created_at.desc()).limit(3).all()
            if tickets:
                response_data = f"I found these tickets that might be related to '{user_message_lower}':\n" + "\n".join([f"- ID {t.id}: {t.title} (Status: {t.status}, By: {t.creator.username})" for t in tickets])
                response_data += "\n\nCould you be more specific if this isn't what you're looking for, or ask a general question?"
                found_specific_query = True
    except Exception as e:
        app.logger.error(f"Chatbot unexpected error during DB query for '{user_message_lower}' (SQLAlchemy): {e}", exc_info=True)
        response_data = "I encountered an unexpected issue while trying to understand your request. Please try again."
        found_specific_query = True
    if not found_specific_query:
        app.logger.info(f"Chatbot: No specific DB query matched for '{user_message_lower}'. Signaling for direct AI processing.")
        return {"type": DIRECT_AI_SIGNAL_TYPE, "original_query": user_message_lower}
    return response_data

def generate_ai_chat_response_gemini(user_message, db_query_or_signal):
    if not app.config['GEMINI_API_KEY']:
        app.logger.warning("Chatbot: Gemini API Key is missing. AI responses will be limited.")
        if isinstance(db_query_or_signal, dict) and db_query_or_signal.get("type") == DIRECT_AI_SIGNAL_TYPE:
            return "My AI capabilities are currently unavailable for general questions. Please try asking about specific tickets."
        if isinstance(db_query_or_signal, dict) and db_query_or_signal.get("type") == TICKET_DETAILS_TYPE:
            return db_query_or_signal.get("summary_text_for_ai", "Ticket information is available, but AI summarization is currently offline.")
        return str(db_query_or_signal) 
    prompt = ""
    if isinstance(db_query_or_signal, dict) and db_query_or_signal.get("type") == DIRECT_AI_SIGNAL_TYPE:
        original_query = db_query_or_signal.get("original_query", user_message)
        app.logger.info(f"Chatbot: Using direct AI prompt for query: '{original_query}'")
        prompt = f"""The user asked: "{original_query}"
Please provide a helpful and general response. You do not have access to specific database information for this question.
Answer as a helpful assistant. If the question seems like a command you cannot fulfill (e.g. 'delete ticket 5'), politely explain you are an informational assistant and cannot perform actions.
If the query is vague, ask for clarification. If it's a greeting, respond politely.
Chatbot's Answer:"""
    elif isinstance(db_query_or_signal, dict) and db_query_or_signal.get("type") == TICKET_DETAILS_TYPE:
        ticket_summary = db_query_or_signal.get("summary_text_for_ai", "Found details for a ticket.")
        app.logger.info(f"Chatbot: Using DB-contextualized (TICKET_DETAILS_TYPE) AI prompt for query: '{user_message}'")
        prompt = f"""User asked: "{user_message}"
Based *only* on this ticket information, provide a friendly and concise summary or answer related to the user's question.
Do not invent information not present in the ticket details. If the ticket description is long, summarize the key points relevant to the user's query.
Ticket Information:
{ticket_summary}

Chatbot's Answer:"""
    else: 
        app.logger.info(f"Chatbot: Using DB-contextualized (string) AI prompt for query: '{user_message}'")
        prompt = f"""User asked: "{user_message}"
Based *only* on the following database information, provide a friendly and concise answer. 
If the database info is a list of items, summarize it or list key items. 
If the database info indicates "no ticket/s found" or a similar negative result, state that politely. 
If it's an error message or a help message like "Please specify...", rephrase that helpfully for the user.
Do not add any information not present in the database result.
Database Result:
```{str(db_query_or_signal)}```
Chatbot's Answer:"""
    try:
        model = genai.GenerativeModel('gemini-1.5-flash-latest') 
        response = model.generate_content(
            prompt, generation_config=genai.types.GenerationConfig(max_output_tokens=500, temperature=0.5)
        )
        ai_text = "".join(part.text for part in response.candidates[0].content.parts).strip() if response.candidates and response.candidates[0].content.parts else ""
        if not ai_text: 
            if response.prompt_feedback and response.prompt_feedback.block_reason:
                block_msg = response.prompt_feedback.block_reason_message or "Content generation was limited due to safety settings."
                app.logger.warning(f"Gemini blocked AI response: {block_msg}. Prompt: {prompt[:200]}")
                return f"My response was limited by content policy ({block_msg}). Could you please rephrase your query or ask something different?"
            app.logger.info(f"Chatbot AI returned empty content. Fallback based on db_query_or_signal. Prompt: {prompt[:200]}")
            if isinstance(db_query_or_signal, dict) and db_query_or_signal.get("type") == DIRECT_AI_SIGNAL_TYPE: return "I'm sorry, I couldn't generate a specific response for that right now. Please try rephrasing your question."
            if isinstance(db_query_or_signal, dict) and db_query_or_signal.get("type") == TICKET_DETAILS_TYPE: return db_query_or_signal.get("summary_text_for_ai", "AI could not summarize the ticket, but its details were found.")
            return "The AI assistant couldn't phrase a response. Here's the direct information I found:\n" + str(db_query_or_signal)
        return ai_text
    except Exception as e:
        app.logger.error(f"Gemini API call error for chatbot (prompt: {prompt[:200]}...): {str(e)}", exc_info=True)
        if isinstance(db_query_or_signal, dict) and db_query_or_signal.get("type") == DIRECT_AI_SIGNAL_TYPE: return "An AI service error occurred. I can't process general queries right now. Try asking about specific tickets."
        if isinstance(db_query_or_signal, dict) and db_query_or_signal.get("type") == TICKET_DETAILS_TYPE: return db_query_or_signal.get("summary_text_for_ai", "An AI service error occurred, but ticket details were found.")
        return f"An AI service error occurred. Here's the direct information I found:\n{str(db_query_or_signal)}"

@app.route('/tools/chatbot', methods=['GET'])
@login_required
def chatbot_page_render(): 
    return render_template('tools/chatbot.html', username=current_user.username, title="AI Assistant")

@app.route('/api/chat', methods=['POST'])
@login_required
def api_chat():
    data = request.get_json()
    if not data: return jsonify({"error": "Invalid JSON payload"}), 400
    msg = data.get('message','').strip()
    mode = data.get('mode', 'ticket_assistant')
    if not msg: return jsonify({"reply": "Please type a message to start the chat."})
    app.logger.info(f"Chat API: mode='{mode}', message='{msg}' from user: {current_user.username}")
    final_reply_text = ""; relevant_docs_list = []; is_direct_ai_response = False; research_topic_suggestion = msg 
    db_result_for_ai_processing = None 
    if mode == 'general_ai':
        db_result_for_ai_processing = {"type": DIRECT_AI_SIGNAL_TYPE, "original_query": msg}
        is_direct_ai_response = True
    else:
        db_query_output = query_database_for_chatbot_sqlalchemy(msg.lower())
        if isinstance(db_query_output, dict):
            if db_query_output.get("type") == TICKET_DETAILS_TYPE:
                ticket_info = db_query_output
                db_result_for_ai_processing = ticket_info 
                research_topic_suggestion = ticket_info.get("title", msg) 
                attachments_from_db = ticket_info.get("attachments_info", [])
                relevant_docs_list.extend(attachments_from_db)
                is_direct_ai_response = False
            elif db_query_output.get("type") == DIRECT_AI_SIGNAL_TYPE: 
                db_result_for_ai_processing = db_query_output    
                is_direct_ai_response = True
            else:
                db_result_for_ai_processing = "Error: Received an unexpected data structure from database query."
                is_direct_ai_response = False 
        else:
            db_result_for_ai_processing = db_query_output
            is_direct_ai_response = False 
    final_reply_text = generate_ai_chat_response_gemini(msg, db_result_for_ai_processing)
    return jsonify({
        "reply": final_reply_text, "relevant_docs": relevant_docs_list,
        "is_direct_ai": is_direct_ai_response, "research_topic_suggestion": research_topic_suggestion,
        "aws_docs": []
    })

@app.route('/api/aws_doc_search', methods=['POST'])
@login_required
def api_aws_doc_search():
    data = request.get_json()
    if not data: return jsonify({"error": "Invalid JSON payload"}), 400
    research_topic = data.get('research_topic', '').strip()
    if not research_topic: return jsonify({"error": "Research topic is required."}), 400
    app.logger.info(f"AWS Doc Search API request for topic: '{research_topic}' by {current_user.username}")
    aws_documentation_links = []
    if not app.config['GEMINI_API_KEY']:
        app.logger.warning("AWS Doc Search: Gemini API Key not configured.")
        return jsonify({"error": "AI service for AWS Doc Search is unavailable."}), 503
    try:
        aws_search_prompt = f"""
        Please find up to 3-4 highly relevant official AWS documentation links related to the following topic: "{research_topic}". 
        For each link, provide a concise title (max 10 words) and the full URL.
        Focus on official AWS documentation (docs.aws.amazon.com, aws.amazon.com blogs, whitepapers, workshops.aws).
        Format your response STRICTLY as a JSON list of objects, where each object has "title" and "url" keys.
        Example:
        [
          {{"title": "Getting Started with Amazon S3", "url": "https://docs.aws.amazon.com/AmazonS3/latest/userguide/GetStartedWithS3.html"}},
          {{"title": "EC2 Instance Types Overview", "url": "https://aws.amazon.com/ec2/instance-types/"}}
        ]
        If no specific official AWS docs are found, return an empty list ([]). Do not invent links. Do not add any other text before or after the JSON list.
        """
        model = genai.GenerativeModel('gemini-1.5-flash-latest') 
        response = model.generate_content(aws_search_prompt, generation_config=genai.types.GenerationConfig(max_output_tokens=1500,temperature=0.2))
        ai_response_text = "".join(part.text for part in response.candidates[0].content.parts).strip() if response.candidates and response.candidates[0].content.parts else ""
        if ai_response_text:
            app.logger.debug(f"AWS Doc Search Gemini Raw Response: {ai_response_text}")
            try:
                if ai_response_text.startswith("```json"): ai_response_text = ai_response_text.split("```json\n", 1)[1].rsplit("\n```", 1)[0]
                elif ai_response_text.startswith("```"): ai_response_text = ai_response_text.split("```\n", 1)[1].rsplit("\n```", 1)[0]
                parsed_links = json.loads(ai_response_text) 
                if isinstance(parsed_links, list):
                    for link_obj in parsed_links:
                        if isinstance(link_obj, dict) and "title" in link_obj and "url" in link_obj:
                            if isinstance(link_obj["url"], str) and (link_obj["url"].startswith("http://") or link_obj["url"].startswith("https://")): aws_documentation_links.append({"title": str(link_obj["title"]), "url": link_obj["url"]})
                            else: app.logger.warning(f"Skipping invalid URL from Gemini for AWS docs: {link_obj.get('url')}")
                        else: app.logger.warning(f"Gemini returned non-dict or malformed link object for AWS docs: {link_obj}")
                else: app.logger.warning(f"Gemini did not return a list for AWS docs, got: {type(parsed_links)}. Raw: {ai_response_text}")
            except json.JSONDecodeError as json_e:
                app.logger.error(f"Failed to parse Gemini JSON response for AWS docs: {json_e}. Raw response: {ai_response_text}")
                if "http" not in ai_response_text: aws_documentation_links = []
            except Exception as e_parse:
                 app.logger.error(f"Unexpected error processing Gemini response for AWS docs: {e_parse}. Raw response: {ai_response_text}")
                 aws_documentation_links = [] 
        if not aws_documentation_links: app.logger.info(f"No valid AWS docs links extracted by AI for topic: {research_topic}")
        return jsonify({"aws_docs": aws_documentation_links, "message": "AWS documentation search complete." if aws_documentation_links else "No relevant AWS documentation found by AI for this topic."})
    except Exception as e:
        app.logger.error(f"Error during AWS doc search API for topic '{research_topic}': {e}", exc_info=True)
        return jsonify({"error": "Failed to search for AWS documentation due to an internal error."}), 500
    
    
#kb diagnostic

# app.py (add to your Models section)

class KBCategory(db.Model):
    __tablename__ = 'kb_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    description = db.Column(db.String(255), nullable=True)
    slug = db.Column(db.String(120), unique=True, nullable=False) # For user-friendly URLs
    parent_id = db.Column(db.Integer, db.ForeignKey('kb_categories.id'), nullable=True) # For subcategories
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    articles = db.relationship('KBArticle', backref='kb_category_ref', lazy='dynamic')
    parent = db.relationship('KBCategory', remote_side=[id], backref='children')

    def __init__(self, *args, **kwargs):
        if not kwargs.get('slug') and kwargs.get('name'):
            kwargs['slug'] = self._generate_slug(kwargs['name'])
        super().__init__(*args, **kwargs)

    def _generate_slug(self, name):
        # Basic slug generation, consider a more robust library for production
        slug = name.lower().strip().replace(' ', '-')
        slug = re.sub(r'[^\w-]', '', slug) # Remove non-alphanumeric except hyphens
        # Ensure uniqueness (simple append number if not unique)
        original_slug = slug
        count = 1
        while KBCategory.query.filter_by(slug=slug).first() and (not self.id or KBCategory.query.filter_by(slug=slug).first().id != self.id):
            slug = f"{original_slug}-{count}"
            count += 1
        return slug
    
    def set_name(self, name):
        self.name = name
        self.slug = self._generate_slug(name)


    def __repr__(self):
        return f'<KBCategory {self.name}>'

class KBArticle(db.Model):
    __tablename__ = 'kb_articles'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    slug = db.Column(db.String(220), unique=True, nullable=False) # For user-friendly URLs
    content = db.Column(db.Text, nullable=False) # Store as Markdown or HTML
    kb_category_id = db.Column(db.Integer, db.ForeignKey('kb_categories.id'), nullable=False)
    author_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True) # Agent/Admin who wrote it
    status = db.Column(db.String(20), default='Draft', nullable=False)  # Draft, Published, Archived
    views = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    published_at = db.Column(db.DateTime, nullable=True)
    tags = db.Column(db.Text, nullable=True) # Comma-separated tags

    author = db.relationship('User', backref='kb_articles_authored')

    def __init__(self, *args, **kwargs):
        if not kwargs.get('slug') and kwargs.get('title'):
            kwargs['slug'] = self._generate_slug(kwargs['title'])
        super().__init__(*args, **kwargs)

    def _generate_slug(self, title):
        # Basic slug generation
        slug = title.lower().strip().replace(' ', '-')
        slug = re.sub(r'[^\w-]', '', slug)
        original_slug = slug
        count = 1
        # Check for uniqueness, making sure to exclude self during an update
        while KBArticle.query.filter_by(slug=slug).first() and (not self.id or KBArticle.query.filter_by(slug=slug).first().id != self.id):
            slug = f"{original_slug}-{count}"
            count += 1
        return slug

    def set_title(self, title):
        self.title = title
        self.slug = self._generate_slug(title)

    def __repr__(self):
        return f'<KBArticle {self.title}>'



class KBCategoryForm(FlaskForm):
    name = StringField('Category Name', validators=[DataRequired(), Length(max=100)])
    description = TextAreaField('Description', validators=[Optional(), Length(max=255)])
    parent_id = SelectField('Parent Category (Optional)', coerce=int, validators=[Optional()])
    submit = SubmitField('Save Category')

    def __init__(self, *args, **kwargs):
        super(KBCategoryForm, self).__init__(*args, **kwargs)
        self.parent_id.choices = [(0, '--- No Parent ---')] + \
                                 [(c.id, c.name) for c in KBCategory.query.order_by('name').all()]
                                 # Add logic to prevent self-parenting if editing

class KBArticleForm(FlaskForm):
    title = StringField('Title', validators=[DataRequired(), Length(max=200)])
    content = TextAreaField('Content (Markdown supported)', validators=[DataRequired()], render_kw={'rows': 15})
    kb_category_id = SelectField('Category', coerce=int, validators=[DataRequired(message="Please select a category.")])
    status = SelectField('Status', choices=[('Draft', 'Draft'), ('Published', 'Published'), ('Archived', 'Archived')],
                         validators=[DataRequired()])
    tags = StringField('Tags (comma-separated)', validators=[Optional(), Length(max=255)])
    submit = SubmitField('Save Article')

    def __init__(self, *args, **kwargs):
        super(KBArticleForm, self).__init__(*args, **kwargs)
        self.kb_category_id.choices = [(0, '--- Select Category ---')] + \
                                      [(c.id, c.name) for c in KBCategory.query.order_by('name').all()]









# report download

def get_filtered_report_tickets(args):
    query = Ticket.query
    query = _apply_common_filters(query, args, model_to_filter=Ticket) 
    
    sort_by = args.get('sort_by', 'created_at')
    sort_order = args.get('sort_order', 'desc')
    if hasattr(Ticket, sort_by):
        column_to_sort = getattr(Ticket, sort_by)
        if sort_order == 'asc':
            query = query.order_by(column_to_sort.asc())
        else:
            query = query.order_by(column_to_sort.desc())
    else:
        query = query.order_by(Ticket.created_at.desc())
    return query.all()



@app.route('/reports/download/<format>')
@admin_required
def download_report_file(format):
    # request.args will contain the filters applied on the reports page
    tickets = get_filtered_report_tickets(request.args)
    
    filename_base = f"ticket_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    
    if not tickets:
        flash("No data found for the selected filters to download.", "warning")
        return redirect(url_for('reports_overview', **request.args))

    if format == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        # Header
        headers = ['ID', 'Title', 'Status', 'Priority', 'Category', 'Created By', 'Assigned To', 'Organization', 'Created At', 'Updated At', 'Resolved At']
        cw.writerow(headers)
        # Data
        for ticket in tickets:
            cw.writerow([
                ticket.id, ticket.title, ticket.status, ticket.priority,
                ticket.category_ref.name if ticket.category_ref else '',
                ticket.creator.username if ticket.creator else '',
                ticket.assignee.username if ticket.assignee else 'Unassigned',
                ticket.organization_option_ref.name if ticket.organization_option_ref else '',
                ticket.created_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.created_at else '',
                ticket.updated_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.updated_at else '',
                ticket.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.resolved_at else ''
            ])
        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename={filename_base}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    elif format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ticket Report"
        
        headers = ['ID', 'Title', 'Status', 'Priority', 'Category', 'Created By', 'Assigned To', 'Organization', 'Created At', 'Updated At', 'Resolved At']
        ws.append(headers)
        for col_num, header_title in enumerate(headers, 1): # Bold headers
            cell = ws.cell(row=1, column=col_num)
            cell.font = openpyxl.styles.Font(bold=True)
            # Auto-adjust column width (basic)
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(len(str(header_title)) + 2, 15)


        for ticket in tickets:
            ws.append([
                ticket.id, ticket.title, ticket.status, ticket.priority,
                ticket.category_ref.name if ticket.category_ref else '',
                ticket.creator.username if ticket.creator else '',
                ticket.assignee.username if ticket.assignee else 'Unassigned',
                ticket.organization_option_ref.name if ticket.organization_option_ref else '',
                ticket.created_at.replace(tzinfo=None) if ticket.created_at else '', # openpyxl prefers naive datetimes
                ticket.updated_at.replace(tzinfo=None) if ticket.updated_at else '',
                ticket.resolved_at.replace(tzinfo=None) if ticket.resolved_at else ''
            ])
        
        # Save to an in-memory stream
        xlsx_stream = io.BytesIO()
        wb.save(xlsx_stream)
        xlsx_stream.seek(0)
        
        output = make_response(xlsx_stream.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename={filename_base}.xlsx"
        output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return output

    elif format == 'pdf':
        # This is a simplified example using WeasyPrint.
        # You'd need to install WeasyPrint and its dependencies (Pango, Cairo, etc.)
        # `pip install WeasyPrint`
        try:
            from weasyprint import HTML, CSS # Keep import local to this block
            # Render a simple HTML table for the PDF
            # You might create a dedicated template: 'reports/pdf_ticket_list.html'
            html_string = render_template('reports/ticket_list_for_pdf.html', tickets=tickets, report_title=f"Ticket Report - {datetime.utcnow().strftime('%Y-%m-%d')}")
            
            # Optional: Add some basic CSS for the PDF
            # css_string = """ body { font-family: sans-serif; } table { border-collapse: collapse; width: 100%; } th, td { border: 1px solid #ccc; padding: 4px; text-align: left; } th { background-color: #f0f0f0;} """
            # pdf_file = HTML(string=html_string).write_pdf(stylesheets=[CSS(string=css_string)])
            
            pdf_file = HTML(string=html_string).write_pdf()


            output = make_response(pdf_file)
            output.headers["Content-Disposition"] = f"attachment; filename={filename_base}.pdf"
            output.headers["Content-type"] = "application/pdf"
            return output
        except ImportError:
            flash("PDF generation library (WeasyPrint) not installed or configured.", "danger")
            app.logger.error("WeasyPrint not found for PDF generation.")
            return redirect(url_for('reports_overview', **request.args))
        except Exception as e:
            flash(f"Error generating PDF: {str(e)}", "danger")
            app.logger.error(f"PDF generation error: {e}", exc_info=True)
            return redirect(url_for('reports_overview', **request.args))


    else:
        flash("Invalid download format requested.", "danger")
        return redirect(url_for('reports_overview'))

@app.route('/reports/share_email', methods=['POST'])
@admin_required
def share_report_by_email():
    data = request.get_json()
    recipient_emails_str = data.get('recipient_email')
    subject = data.get('email_subject') or "Ticket System Report"
    message_body = data.get('email_message') or "Please find the attached ticket report."
    attachment_format = data.get('email_format', 'csv')
    report_filters_str = data.get('report_filters', '')

    # Parse report_filters_str back into a dict-like structure for get_filtered_report_tickets
    from urllib.parse import parse_qs
    filter_args_dict = {k: v[0] for k, v in parse_qs(report_filters_str).items()}

    if not recipient_emails_str:
        return jsonify({'success': False, 'message': 'Recipient email is required.'}), 400
    
    recipient_list = [email.strip() for email in recipient_emails_str.split(',') if email.strip()]
    if not recipient_list:
        return jsonify({'success': False, 'message': 'No valid recipient emails provided.'}), 400

    tickets = get_filtered_report_tickets(filter_args_dict)
    if not tickets:
         return jsonify({'success': False, 'message': 'No data found for the selected filters to share.'}), 400

    # Generate attachment in memory
    attachment_content = None
    attachment_filename = f"ticket_report_{datetime.utcnow().strftime('%Y%m%d')}.{attachment_format}"
    mimetype = 'application/octet-stream'

    if attachment_format == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        headers = ['ID', 'Title', 'Status', 'Priority', 'Category', 'Created By', 'Assigned To', 'Organization', 'Created At', 'Updated At', 'Resolved At']
        cw.writerow(headers)
        for ticket in tickets: cw.writerow([ticket.id, ticket.title, ticket.status, ticket.priority, ticket.category_ref.name if ticket.category_ref else '', ticket.creator.username if ticket.creator else '', ticket.assignee.username if ticket.assignee else 'Unassigned', ticket.organization_option_ref.name if ticket.organization_option_ref else '', ticket.created_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.created_at else '', ticket.updated_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.updated_at else '', ticket.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.resolved_at else ''])
        attachment_content = si.getvalue().encode('utf-8')
        mimetype = 'text/csv'
    elif attachment_format == 'xlsx':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Ticket Report"
        headers = ['ID', 'Title', 'Status', 'Priority', 'Category', 'Created By', 'Assigned To', 'Organization', 'Created At', 'Updated At', 'Resolved At']
        ws.append(headers)
        for ticket in tickets: ws.append([ticket.id, ticket.title, ticket.status, ticket.priority, ticket.category_ref.name if ticket.category_ref else '', ticket.creator.username if ticket.creator else '', ticket.assignee.username if ticket.assignee else 'Unassigned', ticket.organization_option_ref.name if ticket.organization_option_ref else '', ticket.created_at.replace(tzinfo=None) if ticket.created_at else '', ticket.updated_at.replace(tzinfo=None) if ticket.updated_at else '', ticket.resolved_at.replace(tzinfo=None) if ticket.resolved_at else ''])
        xlsx_stream = io.BytesIO(); wb.save(xlsx_stream); xlsx_stream.seek(0)
        attachment_content = xlsx_stream.getvalue()
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif attachment_format == 'pdf':
        try:
            from weasyprint import HTML # Local import
            html_string = render_template('reports/ticket_list_for_pdf.html', tickets=tickets, report_title=subject)
            attachment_content = HTML(string=html_string).write_pdf()
            mimetype = 'application/pdf'
        except ImportError:
            return jsonify({'success': False, 'message': 'PDF generation library not available on server.'}), 500
        except Exception as e_pdf:
            app.logger.error(f"PDF generation for email error: {e_pdf}")
            return jsonify({'success': False, 'message': f'Error generating PDF for email: {str(e_pdf)}'}), 500
    else:
        return jsonify({'success': False, 'message': 'Invalid attachment format.'}), 400

    if attachment_content:
        try:
            msg = Message(subject=subject, recipients=recipient_list, body=message_body)
            msg.attach(attachment_filename, mimetype, attachment_content)
            mail.send(msg)
            return jsonify({'success': True, 'message': 'Report shared successfully!'})
        except Exception as e:
            app.logger.error(f"Failed to send report email: {e}", exc_info=True)
            return jsonify({'success': False, 'message': f'Failed to send email: {str(e)}'}), 500
    else:
        return jsonify({'success': False, 'message': 'Could not generate report attachment.'}), 500




# --- KB Category Management ---
# app.py

# ... (other imports and code) ...

# --- KB Category Management ---
@app.route('/admin/kb/categories')
@agent_required # Changed from @admin_required
def admin_kb_category_list():
    categories = KBCategory.query.order_by('name').all()
    return render_template('admin/kb/category_list.html', title='Manage KB Categories', categories=categories)

@app.route('/admin/kb/category/new', methods=['GET', 'POST'])
@app.route('/admin/kb/category/<int:category_id>/edit', methods=['GET', 'POST'])
@agent_required # Changed from @admin_required
def admin_create_edit_kb_category(category_id=None):
    category = KBCategory.query.get_or_404(category_id) if category_id else None
    form = KBCategoryForm(obj=category)
    
    if category:
        form.parent_id.choices = [(0, '--- No Parent ---')]
        form.parent_id.choices += [(c.id, c.name) for c in KBCategory.query.filter(KBCategory.id != category.id).order_by('name').all()]
    
    if form.validate_on_submit():
        is_new = category is None
        if is_new:
            category = KBCategory()
        
        existing_cat = KBCategory.query.filter(KBCategory.name == form.name.data, KBCategory.id != (category.id if category.id else -1) ).first()
        if existing_cat:
            form.name.errors.append("A category with this name already exists.")
        else:
            category.set_name(form.name.data) 
            category.description = form.description.data
            parent_id_val = form.parent_id.data
            category.parent_id = parent_id_val if parent_id_val != 0 else None
            
            if is_new:
                db.session.add(category)
            try:
                db.session.commit()
                flash(f'KB Category "{category.name}" saved.', 'success')
                return redirect(url_for('admin_kb_category_list'))
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"Error saving KB Category: {e}")
                flash(f'Error saving KB category: {str(e)}', 'danger')
    
    return render_template('admin/kb/create_edit_category.html', 
                           title='Edit KB Category' if category else 'New KB Category', 
                           form=form, category=category)

@app.route('/admin/kb/category/<int:category_id>/delete', methods=['POST'])
@agent_required # Changed from @admin_required
def admin_delete_kb_category(category_id):
    category = KBCategory.query.get_or_404(category_id)
    if category.articles.count() > 0 or KBCategory.query.filter_by(parent_id=category.id).count() > 0:
        flash(f'Cannot delete category "{category.name}" as it contains articles or subcategories.', 'danger')
    else:
        try:
            db.session.delete(category)
            db.session.commit()
            flash(f'KB Category "{category.name}" deleted.', 'success')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting KB Category: {e}")
            flash(f'Error deleting KB category: {str(e)}', 'danger')
    return redirect(url_for('admin_kb_category_list'))


# --- KB Article Management ---
@app.route('/admin/kb/articles')
@app.route('/admin/kb/articles/category/<slug>')
@agent_required # Changed from @admin_required
def admin_kb_article_list(slug=None):
    page = request.args.get('page', 1, type=int)
    query = KBArticle.query
    category_filter = None
    if slug:
        category_filter = KBCategory.query.filter_by(slug=slug).first_or_404()
        query = query.filter_by(kb_category_id=category_filter.id)
        
    articles_pagination = query.order_by(KBArticle.updated_at.desc()).paginate(page=page, per_page=15)
    categories = KBCategory.query.order_by('name').all()
    return render_template('admin/kb/article_list.html', 
                           title='Manage KB Articles' + (f' in {category_filter.name}' if category_filter else ''), 
                           articles_pagination=articles_pagination,
                           categories=categories,
                           current_category_slug=slug)

@app.route('/admin/kb/article/new', methods=['GET', 'POST'])
@app.route('/admin/kb/article/<int:article_id>/edit', methods=['GET', 'POST'])
@agent_required # This was already @agent_required, which is correct
def admin_create_edit_kb_article(article_id=None):
    article = KBArticle.query.get_or_404(article_id) if article_id else None
    form = KBArticleForm(obj=article)

    if form.validate_on_submit():
        is_new = article is None
        if is_new:
            article = KBArticle(author_id=current_user.id)
        
        existing_article = KBArticle.query.filter(KBArticle.title == form.title.data, KBArticle.id != (article.id if article.id else -1)).first()
        if existing_article:
            form.title.errors.append("An article with this title already exists.")
        else:
            article.set_title(form.title.data) 
            article.content = form.content.data
            article.kb_category_id = form.kb_category_id.data
            article.status = form.status.data
            article.tags = form.tags.data.strip() if form.tags.data else None
            
            if article.status == 'Published' and not article.published_at:
                article.published_at = datetime.utcnow()
            elif article.status != 'Published':
                 article.published_at = None 

            if is_new:
                db.session.add(article)
            try:
                db.session.commit()
                flash(f'KB Article "{article.title}" saved.', 'success')
                return redirect(url_for('admin_kb_article_list'))
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"Error saving KB Article: {e}")
                flash(f'Error saving KB article: {str(e)}', 'danger')
    elif request.method == "GET" and not article: 
        category_id_from_url = request.args.get('category_id', type=int)
        if category_id_from_url:
            form.kb_category_id.data = category_id_from_url


    return render_template('admin/kb/create_edit_article.html', 
                           title='Edit KB Article' if article else 'New KB Article', 
                           form=form, article=article)

@app.route('/admin/kb/article/<int:article_id>/delete', methods=['POST'])
@agent_required # This was already @agent_required, which is correct
def admin_delete_kb_article(article_id):
    article = KBArticle.query.get_or_404(article_id)
    try:
        db.session.delete(article)
        db.session.commit()
        flash(f'KB Article "{article.title}" deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting KB Article: {e}")
        flash(f'Error deleting KB article: {str(e)}', 'danger')
    return redirect(url_for('admin_kb_article_list'))



# app.py (add to your public/client-facing routes)
import markdown # For rendering Markdown content

# Initialize Markdown extension (once, globally or in app factory)
md = markdown.Markdown(extensions=['fenced_code', 'tables', 'attr_list', 'nl2br', 'toc'])


@app.context_processor
def inject_kb_categories_for_layout():
    top_level_categories = []
    if current_user.is_authenticated: # Only fetch if user is logged in, as KB viewing requires login
        top_level_categories = KBCategory.query.filter(KBCategory.parent_id.is_(None))\
            .join(KBArticle, KBCategory.id == KBArticle.kb_category_id)\
            .filter(KBArticle.status == 'Published')\
            .distinct()\
            .order_by(KBCategory.name).all()
    return dict(kb_nav_categories=top_level_categories)


@app.route('/kb')
@app.route('/kb/category/<slug>')
@login_required # Added for consistency, ensuring all KB viewing requires login
def kb_category_view(slug=None):
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('q', '').strip()
    current_category = None
    query = KBArticle.query.filter_by(status='Published')

    if slug:
        current_category = KBCategory.query.filter_by(slug=slug).first_or_404()
        category_ids_to_include = [current_category.id]
        children_cats = KBCategory.query.filter_by(parent_id=current_category.id).all()
        for child_cat in children_cats:
            category_ids_to_include.append(child_cat.id)
        query = query.filter(KBArticle.kb_category_id.in_(category_ids_to_include))
        
    if search_query:
        query = query.filter(
            or_(
                KBArticle.title.ilike(f'%{search_query}%'),
                KBArticle.content.ilike(f'%{search_query}%'),
                KBArticle.tags.ilike(f'%{search_query}%')
            )
        )

    articles_pagination = query.order_by(KBArticle.published_at.desc(), KBArticle.views.desc()).paginate(page=page, per_page=10)
    
    all_display_categories = KBCategory.query.order_by('name').all()


    return render_template('kb/category_view.html', 
                           title=current_category.name if current_category else "Knowledge Base",
                           articles_pagination=articles_pagination,
                           current_category=current_category,
                           all_display_categories=all_display_categories,
                           search_query=search_query)

@app.route('/kb/article/<slug>')
@login_required # This already had @login_required
def kb_article_view(slug):
    article = KBArticle.query.filter_by(slug=slug, status='Published').first_or_404()
    
    # Increment views only if the current user is not the author (or if you want to count all views)
    # For simplicity, counting all views for now.
    article.views = (article.views or 0) + 1 
    db.session.commit()
    
    html_content = Markup(md.convert(article.content))

    breadcrumbs = []
    cat_temp = article.kb_category_ref
    while cat_temp:
        breadcrumbs.append(cat_temp)
        cat_temp = cat_temp.parent
    breadcrumbs.reverse()


    return render_template('kb/article_view.html', 
                           title=article.title, 
                           article=article, 
                           html_content=html_content,
                           breadcrumbs=breadcrumbs)

# ... (rest of app.py)

# Helper to convert Markdown in Jinja templates if needed for other parts
@app.template_filter('markdown_to_html')
def markdown_to_html_filter(s):
    return Markup(md.convert(s)) if s else ''




# In app.py

# In app.py

# In app.py

def _apply_common_filters(query, args, model_to_filter=Ticket):
    """
    Helper to apply common filters to a ticket query or a query that can be joined with Ticket.
    model_to_filter: The SQLAlchemy model/alias to apply filters on (e.g., Ticket or an alias).
                     If the query is on a different model (e.g., Category) but needs filtering
                     based on associated Ticket attributes, ensure appropriate joins are made
                     *before* calling this helper, or adapt the helper.
    """
    start_date_str = args.get('start_date')
    end_date_str = args.get('end_date')
    status = args.get('status')
    priority = args.get('priority')
    category_id_str = args.get('category_id')
    organization_id_str = args.get('organization_id')
    assigned_to_id_str = args.get('assigned_to_id')
    department_id_str = args.get('department_id')

    # Date filters (applied to model_to_filter.created_at)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            query = query.filter(model_to_filter.created_at >= start_date)
        except ValueError: current_app.logger.warning(f"Invalid start_date format: {start_date_str}")
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            # Ensure end_date is inclusive by going to the start of the next day
            query = query.filter(model_to_filter.created_at < (end_date + timedelta(days=1)))
        except ValueError: current_app.logger.warning(f"Invalid end_date format: {end_date_str}")
    
    # Standard string/ID based filters
    if status and status != 'all': query = query.filter(model_to_filter.status == status)
    if priority and priority != 'all': query = query.filter(model_to_filter.priority == priority)
    
    if category_id_str and category_id_str != 'all' and category_id_str.isdigit() and category_id_str != '0':
        query = query.filter(model_to_filter.category_id == int(category_id_str))
        
    if organization_id_str and organization_id_str != 'all' and organization_id_str.isdigit() and organization_id_str != '0':
        query = query.filter(model_to_filter.organization_id == int(organization_id_str))
        # Apply department filter ONLY if organization is also selected and valid
        if department_id_str and department_id_str != 'all' and department_id_str.isdigit() and department_id_str != '0':
            query = query.filter(model_to_filter.department_id == int(department_id_str))
            
    if assigned_to_id_str and assigned_to_id_str != 'all':
        if assigned_to_id_str == 'unassigned' or assigned_to_id_str == '0': 
            query = query.filter(model_to_filter.assigned_to_id.is_(None))
        elif assigned_to_id_str.isdigit(): 
            query = query.filter(model_to_filter.assigned_to_id == int(assigned_to_id_str))
    
    return query





# In app.py - analytics_dashboard_page route

@app.route('/tools/analytics_dashboard')
@login_required 
def analytics_dashboard_page():
    all_statuses_data = [{'value': s[0], 'display': s[1]} for s in TICKET_STATUS_CHOICES]
    all_priorities_data = [{'value': p[0], 'display': p[1]} for p in TICKET_PRIORITY_CHOICES]
    all_categories_data = Category.query.order_by(Category.name).all()
    all_organizations_data = OrganizationOption.query.filter_by(is_active=True).order_by(OrganizationOption.name).all()
    all_agents_data = User.query.filter(User.role.in_(['agent', 'admin'])).order_by(User.username).all()

    selected_organization_id_str = request.args.get('organization_id')
    departments_for_filter_initial = []
    if selected_organization_id_str and selected_organization_id_str.isdigit() and selected_organization_id_str != 'all' and selected_organization_id_str != '0':
        try:
            org_id = int(selected_organization_id_str)
            departments_for_filter_initial = Department.query.filter_by(
                organization_id=org_id, 
                is_active=True
            ).order_by(Department.name).all()
        except ValueError:
            pass

    return render_template(
        'tools/analytics_dashboard.html', 
        title="Analytics Dashboard",
        all_statuses=all_statuses_data,
        all_priorities=all_priorities_data,
        all_categories=all_categories_data,
        all_organizations=all_organizations_data,
        all_agents=all_agents_data,
        initial_departments_for_filter=departments_for_filter_initial # For pre-populating dept filter
    )

@analytics_api_bp.route('/ticket_status_distribution')
@admin_required
def ticket_status_distribution():
    # ... your implementation ...
    query = db.session.query(Ticket.status, func.count(Ticket.id).label('count'))
    query = _apply_common_filters(query, request.args, model_to_filter=Ticket)
    results = query.group_by(Ticket.status).all()
    labels = [r.status for r in results]
    data = [r.count for r in results]
    return jsonify(labels=labels, data=data)

@analytics_api_bp.route('/ticket_volume_over_time')
@admin_required
def ticket_volume_over_time():
    # ... your implementation ...
    end_date = datetime.utcnow().date(); start_date = end_date - timedelta(days=29)
    start_date_str = request.args.get('start_date'); end_date_str = request.args.get('end_date')
    if start_date_str: 
        try: start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date(); 
        except ValueError: pass
    if end_date_str: 
        try: end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date(); 
        except ValueError: pass
    if start_date > end_date: start_date = end_date - timedelta(days=29)
    all_dates_in_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
    date_labels = [d.strftime('%Y-%m-%d') for d in all_dates_in_range]; ticket_counts_map = {label: 0 for label in date_labels}
    base_ticket_query = Ticket.query; filtered_ticket_pool = _apply_common_filters(base_ticket_query, request.args, model_to_filter=Ticket)
    results = db.session.query(cast(Ticket.created_at, Date).label('creation_date'), func.count(Ticket.id).label('count')).select_from(filtered_ticket_pool.subquery()).filter(cast(Ticket.created_at, Date) >= start_date).filter(cast(Ticket.created_at, Date) <= end_date).group_by('creation_date').order_by('creation_date').all()
    for r in results: date_str = r.creation_date.strftime('%Y-%m-%d'); ticket_counts_map[date_str] = r.count
    final_counts = [ticket_counts_map[label] for label in date_labels]
    return jsonify(labels=date_labels, data=final_counts)


@analytics_api_bp.route('/tickets_by_category')
@admin_required
def tickets_by_category():
    # ... your implementation ...
    query = db.session.query(Category.name, func.count(Ticket.id).label('count')).join(Ticket, Ticket.category_id == Category.id)
    query = _apply_common_filters(query, request.args, model_to_filter=Ticket) 
    results = query.group_by(Category.name).order_by(func.count(Ticket.id).desc()).all()
    labels = [r.name if r.name else "Uncategorized" for r in results]; data = [r.count for r in results]
    return jsonify(labels=labels, data=data)

@analytics_api_bp.route('/tickets_by_priority')
@admin_required
def tickets_by_priority():
    # ... your implementation ...
    query = db.session.query(Ticket.priority, func.count(Ticket.id).label('count'))
    query = _apply_common_filters(query, request.args, model_to_filter=Ticket)
    results = query.group_by(Ticket.priority).order_by(func.count(Ticket.id).desc()).all()
    labels = [r.priority if r.priority else "No Priority" for r in results]; data = [r.count for r in results]
    return jsonify(labels=labels, data=data)


@analytics_api_bp.route('/tickets_by_agent')
@admin_required
def tickets_by_agent():
    # ... your implementation ...
    AssigneeUser = aliased(User, name="assignee_user")
    query = db.session.query(AssigneeUser.username, func.count(Ticket.id).label('count')).outerjoin(AssigneeUser, Ticket.assigned_to_id == AssigneeUser.id)
    query = _apply_common_filters(query, request.args, model_to_filter=Ticket)
    results = query.group_by(AssigneeUser.username).order_by(func.count(Ticket.id).desc()).all()
    labels = [r.username if r.username else "Unassigned" for r in results]; data = [r.count for r in results]
    return jsonify(labels=labels, data=data)


@analytics_api_bp.route('/severity_distribution')
@admin_required
def severity_distribution():
    # ... your implementation ...
    query = db.session.query(Ticket.severity, func.count(Ticket.id).label('count'))
    query = _apply_common_filters(query, request.args, model_to_filter=Ticket)
    results_raw = query.group_by(Ticket.severity).all()
    severity_options_ordered = SeverityOption.query.order_by(SeverityOption.order).all()
    severity_order_map = {sev.name: idx for idx, sev in enumerate(severity_options_ordered)}
    results = sorted(results_raw, key=lambda r: severity_order_map.get(r.severity, 999))
    labels = [r.severity if r.severity else "No Severity" for r in results]; data = [r.count for r in results]
    return jsonify(labels=labels, data=data)

# THIS IS THE SINGLE DEFINITION FOR avg_resolution_time
@analytics_api_bp.route('/avg_resolution_time')
@admin_required 
def avg_resolution_time():
    if not hasattr(Ticket, 'resolved_at'): 
        current_app.logger.error("Ticket model is missing 'resolved_at' attribute.")
        return jsonify({"error": "Server configuration error"}), 500
    base_query = db.session.query(Ticket.created_at, Ticket.resolved_at).filter(
        Ticket.status.in_(['Resolved', 'Closed']),
        Ticket.resolved_at.isnot(None), 
        Ticket.created_at.isnot(None)  
    )
    filtered_query = _apply_common_filters(base_query, request.args, model_to_filter=Ticket)
    tickets_for_resolution = filtered_query.all()
    total_resolution_seconds = 0; resolved_ticket_count = 0 
    for created_at_ts, resolved_at_ts in tickets_for_resolution:
        if resolved_at_ts and created_at_ts: 
            resolution_duration = resolved_at_ts - created_at_ts
            if resolution_duration.total_seconds() >= 0: 
                total_resolution_seconds += resolution_duration.total_seconds()
                resolved_ticket_count +=1
    average_resolution_time_str = "N/A"; avg_res_seconds = 0
    if resolved_ticket_count > 0:
        average_seconds = total_resolution_seconds / resolved_ticket_count
        avg_res_seconds = int(average_seconds)
        days = int(average_seconds // (24 * 3600)); hours = int((average_seconds % (24 * 3600)) // 3600); minutes = int((average_seconds % 3600) // 60)
        parts = []; 
        if days > 0: parts.append(f"{days}d"); 
        if hours > 0: parts.append(f"{hours}h"); 
        if minutes > 0: parts.append(f"{minutes}m")
        average_resolution_time_str = " ".join(parts) if parts else "<1m"
    return jsonify({'average_resolution_time_str': average_resolution_time_str, 'average_resolution_seconds': avg_res_seconds, 'resolved_ticket_count': resolved_ticket_count})

@analytics_api_bp.route('/cumulative_tickets_over_time')
@admin_required
def cumulative_tickets_over_time():
    # ... your implementation ...
    args = request.args; end_date_max = datetime.utcnow().date(); start_date_min = end_date_max - timedelta(days=29)
    start_date_filter_str = args.get('start_date'); end_date_filter_str = args.get('end_date')
    if start_date_filter_str: 
        try: start_date_min = datetime.strptime(start_date_filter_str, '%Y-%m-%d').date(); 
        except ValueError: pass
    if end_date_filter_str: 
        try: end_date_max = datetime.strptime(end_date_filter_str, '%Y-%m-%d').date(); 
        except ValueError: pass
    if start_date_min > end_date_max: start_date_min = end_date_max - timedelta(days=29)
    all_dates_in_range = [start_date_min + timedelta(days=x) for x in range((end_date_max - start_date_min).days + 1)]
    date_labels = [d.strftime('%Y-%m-%d') for d in all_dates_in_range]
    base_ticket_pool = Ticket.query; filtered_ticket_pool = _apply_common_filters(base_ticket_pool, args, model_to_filter=Ticket)
    daily_counts_results = db.session.query(cast(Ticket.created_at, Date).label('creation_date'), func.count(Ticket.id).label('daily_count')).select_from(filtered_ticket_pool.subquery()).filter(cast(Ticket.created_at, Date).between(start_date_min, end_date_max)).group_by(cast(Ticket.created_at, Date)).order_by(cast(Ticket.created_at, Date).asc()).all()
    daily_counts_map = {result.creation_date.strftime('%Y-%m-%d'): result.daily_count for result in daily_counts_results}
    cumulative_sum = 0; final_cumulative_counts = []
    for date_str in date_labels: cumulative_sum += daily_counts_map.get(date_str, 0); final_cumulative_counts.append(cumulative_sum)
    return jsonify(labels=date_labels, data=final_cumulative_counts)

@analytics_api_bp.route('/priority_status_radar_for_category')
@admin_required
def priority_status_radar_for_category():
    # ... your implementation (as previously corrected) ...
    args = request.args; category_id_filter_str = args.get('category_id'); target_category_id = None; category_name_for_chart = "Overall" 
    if category_id_filter_str and category_id_filter_str != 'all' and category_id_filter_str.isdigit() and category_id_filter_str != '0':
        target_category_id = int(category_id_filter_str); cat_obj = db.session.get(Category, target_category_id)
        if cat_obj: category_name_for_chart = cat_obj.name; 
        else: target_category_id = None 
    if not target_category_id: 
        most_active_cat_query_base = db.session.query(Ticket.category_id, func.count(Ticket.id).label('ticket_count'))
        temp_args_for_most_active = args.copy(); temp_args_for_most_active.pop('category_id', None) 
        filtered_cat_counts_query = _apply_common_filters(most_active_cat_query_base, temp_args_for_most_active, model_to_filter=Ticket)
        most_active_cat = filtered_cat_counts_query.group_by(Ticket.category_id).order_by(func.count(Ticket.id).desc()).first()
        if most_active_cat and most_active_cat.category_id:
            target_category_id = most_active_cat.category_id; cat_obj = db.session.get(Category, target_category_id)
            if cat_obj: category_name_for_chart = f"{cat_obj.name} (Most Active)"; 
            else: return jsonify(labels=[], datasets=[], category_name="Error: Category Not Found")
        else: return jsonify(labels=[], datasets=[], category_name="No Ticket Data for Radar")
    if not target_category_id: return jsonify(labels=[], datasets=[], category_name="No Category Data")
    radar_statuses = ['Open', 'In Progress']; radar_priorities_config = TICKET_PRIORITY_CHOICES; priority_labels = [p[1] for p in radar_priorities_config]; datasets_data = []
    for status_val, status_display in [(s[0], s[1]) for s in TICKET_STATUS_CHOICES if s[0] in radar_statuses]:
        query = db.session.query(Ticket.priority, func.count(Ticket.id).label('count')).filter(Ticket.category_id == target_category_id).filter(Ticket.status == status_val)
        temp_args_for_dataset = args.copy(); temp_args_for_dataset.pop('status', None); temp_args_for_dataset.pop('priority', None); temp_args_for_dataset.pop('category_id', None) 
        final_query_for_status = _apply_common_filters(query, temp_args_for_dataset, model_to_filter=Ticket)
        status_priority_counts_raw = final_query_for_status.group_by(Ticket.priority).all()
        status_priority_map = {r.priority: r.count for r in status_priority_counts_raw}
        data_for_status = [status_priority_map.get(p_val[0], 0) for p_val in radar_priorities_config]
        datasets_data.append({ 'label': status_display, 'data': data_for_status, 'borderWidth': 1.5 })
    return jsonify(labels=priority_labels, datasets=datasets_data, category_name=category_name_for_chart)

@analytics_api_bp.route('/total_tickets_in_period') # This was defined separately, include it here
def total_tickets_in_period():
    query = Ticket.query
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    start_date, end_date = None, None # Initialize
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
             return jsonify({"error": "Invalid date format. Use YYYY-MM-DD."}), 400
    
    # Apply common filters, then specific date filter if provided
    query = _apply_common_filters(query, request.args, model_to_filter=Ticket)
    if start_date and end_date: # Apply date range if valid dates were parsed
        query = query.filter(Ticket.created_at >= start_date, Ticket.created_at <= end_date)

    count = query.count()
    period_desc = "Selected Period" # Default description
    # You might want to refine period_desc based on whether dates were actually applied.
    # This is just an example if you still want to show it.
    if start_date and end_date:
        period_desc = f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
    elif start_date_str: # Check original string if date parsing failed but string exists
        period_desc = f"From {start_date_str}"
    elif end_date_str:
        period_desc = f"Until {end_date_str}"

    return jsonify({'total_tickets': count, 'period_description': period_desc})

@analytics_api_bp.route('/open_urgent_high_tickets') # This was defined separately, include it here
def open_urgent_high_tickets():
    open_statuses = ['Open', 'In Progress', 'On Hold'] # Adjusted 'Pending' to 'On Hold' if that's your actual status
    urgent_high_priorities = ['Urgent', 'High']
    query = Ticket.query.filter(
        Ticket.status.in_(open_statuses),
        Ticket.priority.in_(urgent_high_priorities)
    )
    query = _apply_common_filters(query, request.args, model_to_filter=Ticket) # Apply other common filters
    count = query.count()
    return jsonify({'open_urgent_high_count': count})




app.register_blueprint(analytics_api_bp)


# --- CLI Commands ---
@app.cli.command('init-db')
def init_db_command():
    try:
        with app.app_context():
            engine = db.engine
            if engine.name == 'mysql': app.logger.info("MySQL detected. Attempting to disable FOREIGN_KEY_CHECKS."); connection = engine.connect(); connection.execute(db.text('SET FOREIGN_KEY_CHECKS = 0;')); connection.commit(); connection.close()
            app.logger.info("Dropping all tables..."); db.drop_all(); app.logger.info("All tables dropped successfully.")
            app.logger.info("Creating all tables..."); db.create_all(); app.logger.info("All tables created successfully.")
            if engine.name == 'mysql': app.logger.info("Attempting to re-enable FOREIGN_KEY_CHECKS for MySQL."); connection = engine.connect(); connection.execute(db.text('SET FOREIGN_KEY_CHECKS = 1;')); connection.commit(); connection.close()
        print('Database tables dropped and recreated successfully.')
    except Exception as e:
        app.logger.error(f"Error during init-db CLI command: {e}", exc_info=True); print(f"Error during init-db: {e}")
        if 'engine' in locals() and engine.name == 'mysql':
            try:
                with app.app_context(): connection = db.engine.connect(); connection.execute(db.text('SET FOREIGN_KEY_CHECKS = 1;')); connection.commit(); connection.close()
            except Exception as fk_e: app.logger.error(f"Failed to re-enable FOREIGN_KEY_CHECKS after error: {fk_e}")


@app.cli.command('create-initial-data')
def create_initial_data_command():
    with app.app_context():
        # 1. Create Organizations First
        org_ck_name = 'CloudKeeper (CK)'
        org_jiet_name = 'JIET Jodhpur' # Example, adjust if not your actual mapped domain
        org_other_name = 'Other Demo Org' # For a user not tied to CK/JIET domain

        org_ck = OrganizationOption.query.filter_by(name=org_ck_name).first()
        if not org_ck:
            org_ck = OrganizationOption(name=org_ck_name, is_active=True)
            db.session.add(org_ck)
            print(f"Created Organization: {org_ck_name}")

        org_jiet = OrganizationOption.query.filter_by(name=org_jiet_name).first()
        if not org_jiet: # Only create if you have users intended for it
            email_for_jiet_user = 'user@jietjodhpur.ac.in' # Example
            if any(u['email'].endswith('@jietjodhpur.ac.in') for u in users_data_definitions): # Check if needed
                org_jiet = OrganizationOption(name=org_jiet_name, is_active=True)
                db.session.add(org_jiet)
                print(f"Created Organization: {org_jiet_name}")

        org_other = OrganizationOption.query.filter_by(name=org_other_name).first()
        if not org_other:
            org_other = OrganizationOption(name=org_other_name, is_active=True)
            db.session.add(org_other)
            print(f"Created Organization: {org_other_name}")

        db.session.commit() # Commit organizations to ensure they have IDs

        # Re-fetch organizations to ensure IDs are populated for department creation
        org_ck = OrganizationOption.query.filter_by(name=org_ck_name).first()
        if org_jiet_name and not org_jiet: # Re-fetch only if it was meant to be created
            org_jiet = OrganizationOption.query.filter_by(name=org_jiet_name).first()
        if not org_other:
             org_other = OrganizationOption.query.filter_by(name=org_other_name).first()


        # 2. Create Departments (optional, can be None for some users)
        dept_ck_support_name = 'CK Support Dept'
        dept_ck_support = None
        if org_ck: # Only create dept if org exists
            dept_ck_support = Department.query.filter_by(name=dept_ck_support_name, organization_id=org_ck.id).first()
            if not dept_ck_support:
                dept_ck_support = Department(name=dept_ck_support_name, organization_id=org_ck.id, is_active=True)
                db.session.add(dept_ck_support)
                print(f"Created Department: {dept_ck_support_name} in {org_ck.name}")
            db.session.commit() # Commit department
            if not dept_ck_support.id: # re-fetch if newly created
                 dept_ck_support = Department.query.filter_by(name=dept_ck_support_name, organization_id=org_ck.id).first()


        # 3. Define User Data using actual Organization and Department objects (or their IDs)
        users_data_definitions_template = [
            {'username': 'admin', 'email_suffix': '@cloudkeeper.com', 'role': 'admin', 'password': 'adminpass', 'org_name': 'CloudKeeper (CK)', 'dept_name': None},
            {'username': 'agent_ck', 'email_suffix': '+1@cloudkeeper.com', 'role': 'agent', 'password': 'agentpass', 'org_name': 'CloudKeeper (CK)', 'dept_name': None},
            {'username': 'client_ck_support', 'email_suffix': '+2@cloudkeeper.com', 'role': 'client', 'password': 'clientpass', 'org_name': 'CloudKeeper (CK)', 'dept_name': 'CK Support Dept'},
            {'username': 'org_client_ck', 'email_suffix': '+3@cloudkeeper.com', 'role': 'organization_client', 'password': 'clientpass', 'org_name': 'CloudKeeper (CK)', 'dept_name': None},
            {'username': 'client_jiet_it', 'email_suffix': '+4@jietjodhpur.ac.in', 'role': 'client', 'password': 'clientpass', 'org_name': 'JIET Jodhpur', 'dept_name': 'JIET IT Support'},
            {'username': 'client_ck_sales', 'email_suffix': '+5@cloudkeeper.com', 'role': 'client', 'password': 'clientpass', 'org_name': 'CloudKeeper (CK)', 'dept_name': 'CK Sales Dept'},
            {'username': 'client_demo_x', 'email_suffix': '+6@example.com', 'role': 'client', 'password': 'clientpass', 'org_name': 'Demo Company X', 'dept_name': None},
        ]
        
        base_email_user = 'monish.jodha' # Centralize the base email user part

        users_data_definitions = []
        for u_template in users_data_definitions_template:
            user_def = u_template.copy()
            user_def['email'] = f"{base_email_user}{u_template['email_suffix']}"
            # Use actual ENV VARS for emails and passwords if available, else use defaults
            user_def['email'] = os.environ.get(f"{u_template['username'].upper()}_USER_EMAIL", user_def['email'])
            user_def['password'] = os.environ.get(f"{u_template['username'].upper()}_USER_PASSWORD", u_template['password'])
            users_data_definitions.append(user_def)


        # 1. Ensure all necessary Organizations exist based on user definitions
        print("Ensuring necessary organizations exist...")
        org_objects = {}
        distinct_org_names = set(u['org_name'] for u in users_data_definitions if u['org_name'])
        for org_name in distinct_org_names:
            org = OrganizationOption.query.filter_by(name=org_name).first()
            if not org:
                org = OrganizationOption(name=org_name, is_active=True)
                db.session.add(org)
                print(f"  Created Organization: {org_name}")
            org_objects[org_name] = org
        db.session.commit() # Commit orgs to get IDs
        # Re-fetch to ensure objects have IDs
        for org_name in org_objects:
            org_objects[org_name] = OrganizationOption.query.filter_by(name=org_name).first()


        # 2. Ensure all necessary Departments exist
        print("\nEnsuring necessary departments exist...")
        dept_objects = {}
        distinct_dept_defs = set((u['dept_name'], u['org_name']) for u in users_data_definitions if u['dept_name'] and u['org_name'])
        for dept_name, org_name_for_dept in distinct_dept_defs:
            org_obj_for_dept = org_objects.get(org_name_for_dept)
            if not org_obj_for_dept:
                print(f"  Skipping department '{dept_name}' creation because its organization '{org_name_for_dept}' was not found/created.")
                continue
            
            dept_key = (dept_name, org_obj_for_dept.id) # Use (name, org_id) as a unique key for the dict
            dept = Department.query.filter_by(name=dept_name, organization_id=org_obj_for_dept.id).first()
            if not dept:
                dept = Department(name=dept_name, organization_id=org_obj_for_dept.id, is_active=True)
                db.session.add(dept)
                print(f"  Created Department: {dept_name} in {org_name_for_dept}")
            dept_objects[dept_key] = dept
        db.session.commit() # Commit depts to get IDs
        # Re-fetch to ensure objects have IDs
        for dept_key_tuple in dept_objects:
            dept_name_lookup, org_id_lookup = dept_key_tuple
            dept_objects[dept_key_tuple] = Department.query.filter_by(name=dept_name_lookup, organization_id=org_id_lookup).first()


        # 3. Create Users
        print("\nAttempting to create initial users...")
        created_user_ids = {}
        for u_data in users_data_definitions:
            user_exists_check = User.query.filter(
                (User.username == u_data['username']) | (User.email == u_data['email'])
            ).first()

            if not user_exists_check:
                org_obj_for_user = org_objects.get(u_data['org_name'])
                org_id_to_assign = org_obj_for_user.id if org_obj_for_user else None
                
                dept_obj_for_user = None
                if u_data['dept_name'] and org_obj_for_user:
                    dept_obj_for_user = dept_objects.get((u_data['dept_name'], org_obj_for_user.id))
                dept_id_to_assign = dept_obj_for_user.id if dept_obj_for_user and org_id_to_assign else None
                
                # Auto-assign org based on email domain IF no org explicitly defined in u_data for this user
                # AND the user is a client type.
                if u_data['role'] in ['client', 'organization_client'] and not org_id_to_assign and not org_obj_for_user:
                    auto_assigned_org = get_organization_by_email_domain(u_data['email'], auto_create=True)
                    if auto_assigned_org:
                        org_id_to_assign = auto_assigned_org.id
                        print(f"  Auto-assigning organization '{auto_assigned_org.name}' to {u_data['username']} based on email domain.")
                        # If an org was auto-assigned, check if a department with the same name was meant for this *new* org.
                        # This is a bit more complex logic - for now, auto-assignment will not also auto-assign department.
                        # Department is usually more explicit.


                user = User(username=u_data['username'], email=u_data['email'], role=u_data['role'],
                            organization_id=org_id_to_assign,
                            department_id=dept_id_to_assign)
                user.set_password(u_data['password'])
                db.session.add(user)
                db.session.flush() 
                created_user_ids[user.username] = user.id
                print(f"  User '{user.username}' created with ID: {user.id} (Role: {user.role}, Org: {org_id_to_assign}, Dept: {dept_id_to_assign}).")
            else:
                found_by_username = (user_exists_check.username == u_data['username'])
                found_by_email = (user_exists_check.email == u_data['email'])
                match_reason = []
                if found_by_username: match_reason.append("username")
                if found_by_email: match_reason.append("email")
                
                created_user_ids[user_exists_check.username] = user_exists_check.id
                print(f"  User '{u_data['username']}' (or email '{u_data['email']}') matches existing User ID: {user_exists_check.id} (found by {', '.join(match_reason)}). Skipping creation.")
        
        db.session.commit()
        print("\n--- CREATED/EXISTING USER IDs (for reference) ---")
        for username_key, user_id_val in created_user_ids.items():
            user_obj = db.session.get(User, user_id_val)
            if user_obj:
                org_info = f"Org: {user_obj.organization.name if user_obj.organization else 'N/A'}"
                dept_info = f"Dept: {user_obj.department.name if user_obj.department else 'N/A'}"
                print(f"  User: {user_obj.username} (ID: {user_id_val}) - Role: {user_obj.role} - {org_info} - {dept_info}")
            else:
                print(f"  Warning: Could not retrieve user details for ID {user_id_val} (originally specified for username key: {username_key})")
        print("-----------------------------------------")

        # 4. Standard Options (Categories, Severities, etc.) - this part was okay previously
        options_map = {
            Category: ['Technical Support', 'Billing Inquiry', 'General Question', 'Feature Request', 'Hardware Issue', 'Software Bug'],
            CloudProviderOption: ['AWS', 'Azure', 'GCP', 'On-Premise', 'Other Cloud', 'Not Applicable'],
            EnvironmentOption: ['Production', 'Staging', 'Development', 'Test', 'QA', 'UAT', 'Shared'],
            FormTypeOption: ['Incident Report', 'Service Request', 'Change Request', 'Information Query'],
            APNOpportunityOption: ['MAP Funding', 'Well-Architected Review Lead', 'New Service PoC', 'Migration Project', 'Training Need'],
            SupportModalOption: ['Basic SLA', 'Standard SLA', 'Premium SLA', 'Enterprise Support', '24/7 Critical']
        }
        print("\nEnsuring default standard options...")
        for model_class, names in options_map.items():
            for name_val in names:
                if not model_class.query.filter_by(name=name_val).first():
                    if model_class == Category:
                        instance_args = {
                            'name': name_val,
                            'description': f"Category for {name_val}"
                        }
                    else:
                        instance_args = {'name': name_val, 'is_active': True}
                        if hasattr(model_class, 'description'):
                             instance_args['description'] = f"Default description for {name_val}"
                    
                    db.session.add(model_class(**instance_args))
                    print(f"  Added {model_class.__name__}: {name_val}")

        severities_data = [
            {'name': 'Severity 1 (Critical)', 'order': 1, 'description': 'System down, critical impact.'},
            {'name': 'Severity 2 (High)', 'order': 2, 'description': 'Major functionality impacted.'},
            {'name': 'Severity 3 (Medium)', 'order': 3, 'description': 'Minor functionality impacted, workaround available.'},
            {'name': 'Severity 4 (Low)', 'order': 4, 'description': 'Cosmetic issue or informational request.'}
        ]
        for sev_data in severities_data:
            if not SeverityOption.query.filter_by(name=sev_data['name']).first():
                db.session.add(SeverityOption(name=sev_data['name'], order=sev_data['order'], description=sev_data['description'], is_active=True))
                print(f"  Added Severity: {sev_data['name']}")
        db.session.commit()


        # 5. Create Sample Tickets
        print("\nCreating sample tickets...")
        cat_tech = Category.query.filter_by(name='Technical Support').first()
        cat_billing = Category.query.filter_by(name='Billing Inquiry').first()
        sev_high = SeverityOption.query.filter_by(name='Severity 2 (High)').first()
        sev_medium = SeverityOption.query.filter_by(name='Severity 3 (Medium)').first()

        # Re-fetch critical objects to ensure they are in the current session with IDs
        org_ck_final = org_objects.get('CloudKeeper (CK)')
        dept_ck_support_final = dept_objects.get(('CK Support Dept', org_ck_final.id)) if org_ck_final else None
        
        org_jiet_final = org_objects.get('JIET Jodhpur')
        dept_jiet_it_final = dept_objects.get(('JIET IT Support', org_jiet_final.id)) if org_jiet_final else None

        user_client_ck_support = User.query.get(created_user_ids.get('client_ck_support'))
        user_org_client_ck = User.query.get(created_user_ids.get('org_client_ck'))
        user_client_jiet_it = User.query.get(created_user_ids.get('client_jiet_it'))


        if user_client_ck_support and cat_tech and sev_high and org_ck_final and dept_ck_support_final:
            if not Ticket.query.filter_by(title="CK Support: VPN Connection Dropping").first():
                t_vpn = Ticket(title="CK Support: VPN Connection Dropping", description="Intermittent VPN drops for CK Support team.",
                               created_by_id=user_client_ck_support.id, category_id=cat_tech.id, severity=sev_high.name, priority="High",
                               organization_id=org_ck_final.id, department_id=dept_ck_support_final.id, customer_name=org_ck_final.name)
                db.session.add(t_vpn); db.session.flush()
                log_interaction(t_vpn.id, 'TICKET_CREATED', user_id=user_client_ck_support.id, details={'title': t_vpn.title})
                print(f"  Ticket added for CK Support (VPN) by User ID {user_client_ck_support.id}")

        if user_org_client_ck and cat_billing and sev_medium and org_ck_final:
            if not Ticket.query.filter_by(title="CK Org: Query on Annual Subscription").first():
                t_billing = Ticket(title="CK Org: Query on Annual Subscription", description="Need clarification on annual billing terms for entire CK org.",
                                   created_by_id=user_org_client_ck.id, category_id=cat_billing.id, severity=sev_medium.name, priority="Medium",
                                   organization_id=org_ck_final.id, department_id=None, customer_name=org_ck_final.name)
                db.session.add(t_billing); db.session.flush()
                log_interaction(t_billing.id, 'TICKET_CREATED', user_id=user_org_client_ck.id, details={'title': t_billing.title})
                print(f"  Ticket added by CK Org Client (ID {user_org_client_ck.id}) - Org Wide")
        
        if user_client_jiet_it and cat_tech and sev_high and org_jiet_final and dept_jiet_it_final:
            if not Ticket.query.filter_by(title="JIET IT: Email Server Lag").first():
                t_email = Ticket(title="JIET IT: Email Server Lag", description="JIET staff reporting slow email delivery.",
                                 created_by_id=user_client_jiet_it.id, category_id=cat_tech.id, severity=sev_high.name, priority="High",
                                 organization_id=org_jiet_final.id, department_id=dept_jiet_it_final.id, customer_name=org_jiet_final.name)
                db.session.add(t_email); db.session.flush()
                log_interaction(t_email.id, 'TICKET_CREATED', user_id=user_client_jiet_it.id, details={'title': t_email.title})
                print(f"  Ticket added for JIET IT (Email) by user ID {user_client_jiet_it.id}")

        db.session.commit()
        print("Sample tickets committed.")
        print("\nInitial data setup process complete. Please check the printed user IDs for reference.")

# ... (Rest of your app.py, especially the __main__ block for running the app) ...

if __name__ == '__main__':
    current_upload_folder = app.config.get('UPLOAD_FOLDER')
    if not os.path.isabs(current_upload_folder):
        current_upload_folder = os.path.join(app.root_path, current_upload_folder)
        app.config['UPLOAD_FOLDER'] = current_upload_folder

    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        try:
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            app.logger.info(f"Startup: Upload folder ensured/created at: {app.config['UPLOAD_FOLDER']}")
        except OSError as e:
            app.logger.critical(f"CRITICAL STARTUP FAILURE: Could not create upload folder {app.config['UPLOAD_FOLDER']}: {e}. Check path and permissions. Application might not work correctly.", exc_info=True)

    app.run(debug=True, host='0.0.0.0', port=5000)